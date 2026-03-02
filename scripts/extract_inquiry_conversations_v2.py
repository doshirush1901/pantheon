#!/usr/bin/env python3
"""
Extract ALL conversations from Inquiry Form leads - V2
More thorough extraction that captures more Q&A pairs.
"""

import json
import re
import base64
from pathlib import Path
from datetime import datetime
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).parent.parent

# Auto-reply patterns
AUTO_REPLY_PATTERNS = [
    r'out of (the )?office', r'automatic reply', r'auto[- ]?reply',
    r'vacation', r'mailer-daemon', r'postmaster', r'undeliverable',
]
AUTO_REPLY_REGEX = re.compile('|'.join(AUTO_REPLY_PATTERNS), re.IGNORECASE)


def clean_email(email):
    """Clean email string."""
    if not email:
        return []
    # Handle multiple emails separated by /
    emails = []
    for e in str(email).replace(' ', '').split('/'):
        e = e.strip().lower()
        if '@' in e and '>' in e:
            e = e.split('<')[-1].replace('>', '')
        if '@' in e:
            emails.append(e)
    return emails


def is_genuine(msg):
    """Check if message is genuine."""
    subject = msg.get('subject', '').lower()
    body = msg.get('body', '')
    
    if AUTO_REPLY_REGEX.search(subject):
        return False
    if len(body.strip()) < 30:
        return False
    return True


def get_message_body(payload):
    """Extract body from payload."""
    body = ""
    
    if 'body' in payload and payload['body'].get('data'):
        body = base64.urlsafe_b64decode(payload['body']['data']).decode('utf-8', errors='ignore')
    elif 'parts' in payload:
        for part in payload['parts']:
            if part.get('mimeType') == 'text/plain' and part.get('body', {}).get('data'):
                body = base64.urlsafe_b64decode(part['body']['data']).decode('utf-8', errors='ignore')
                break
            elif 'parts' in part:
                for subpart in part['parts']:
                    if subpart.get('mimeType') == 'text/plain' and subpart.get('body', {}).get('data'):
                        body = base64.urlsafe_b64decode(subpart['body']['data']).decode('utf-8', errors='ignore')
                        break
    
    # Clean body
    body = re.sub(r'\[image:.*?\]', '', body)
    body = re.sub(r'\n{3,}', '\n\n', body)
    return body.strip()


def get_gmail_service():
    """Initialize Gmail API."""
    token_path = PROJECT_ROOT / "token_rushabh.json"
    creds = Credentials.from_authorized_user_file(str(token_path))
    return build('gmail', 'v1', credentials=creds)


def search_all_threads(service, email, max_results=20):
    """Search for ALL threads with this email."""
    try:
        # Broader search
        query = f"from:{email} OR to:{email}"
        result = service.users().threads().list(
            userId='me', q=query, maxResults=max_results
        ).execute()
        return result.get('threads', [])
    except Exception as e:
        return []


def extract_thread_conversations(service, thread_id, lead_emails):
    """Extract all Q&A pairs from a thread."""
    pairs = []
    
    try:
        thread = service.users().threads().get(
            userId='me', id=thread_id, format='full'
        ).execute()
        
        messages = []
        for msg in thread.get('messages', []):
            headers = {h['name'].lower(): h['value'] for h in msg['payload'].get('headers', [])}
            
            msg_data = {
                'from': headers.get('from', ''),
                'to': headers.get('to', ''),
                'subject': headers.get('subject', ''),
                'date': headers.get('date', ''),
                'body': get_message_body(msg['payload']),
            }
            
            if is_genuine(msg_data):
                messages.append(msg_data)
        
        if len(messages) < 2:
            return pairs
        
        # Categorize messages
        customer_msgs = []
        rushabh_msgs = []
        
        for msg in messages:
            from_addr = msg['from'].lower()
            
            if 'machinecraft' in from_addr:
                rushabh_msgs.append(msg)
            else:
                # Check if from any of the lead emails
                for lead_email in lead_emails:
                    if lead_email in from_addr:
                        customer_msgs.append(msg)
                        break
        
        # Create pairs - match customer message with next Rushabh response
        for cust_msg in customer_msgs:
            cust_body = cust_msg['body']
            if len(cust_body) < 50:
                continue
                
            # Find any Rushabh response in this thread
            for rush_msg in rushabh_msgs:
                rush_body = rush_msg['body']
                if len(rush_body) < 50:
                    continue
                    
                pairs.append({
                    'customer_question': cust_body[:2000],
                    'rushabh_response': rush_body[:2000],
                    'subject': cust_msg['subject'],
                })
                break  # One pair per customer message
        
        # Also pair Rushabh outreach with any customer response
        for rush_msg in rushabh_msgs:
            rush_body = rush_msg['body']
            if len(rush_body) < 100:
                continue
                
            for cust_msg in customer_msgs:
                cust_body = cust_msg['body']
                if len(cust_body) < 30:
                    continue
                    
                # Check if this pair already exists
                already_exists = any(
                    p['customer_question'][:100] == cust_body[:100] 
                    for p in pairs
                )
                if not already_exists:
                    pairs.append({
                        'customer_question': cust_body[:2000],
                        'rushabh_response': rush_body[:2000],
                        'subject': rush_msg['subject'],
                    })
                    break
                    
    except Exception as e:
        pass
    
    return pairs


def main():
    print("=" * 70)
    print("INQUIRY FORM CONVERSATION EXTRACTION V2")
    print("=" * 70)
    
    # Load leads from Excel
    wb = load_workbook(PROJECT_ROOT / 'data/imports/Single Station Inquiry Form (Responses) (2).xlsx')
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            data = dict(zip(headers, row))
            email = data.get('Client Email') or data.get('Email Address')
            if email:
                leads.append({
                    'emails': clean_email(email),
                    'name': data.get('Client Name', ''),
                    'company': data.get('Company Name', ''),
                    'interest': data.get('Customer interest', ''),
                    'forming_area': data.get('Max. Forming Area', ''),
                    'depth': data.get('Max Forming Depth (Draw)', ''),
                    'thickness': data.get('Max Sheet Thickness', ''),
                    'materials': data.get('Materials to Process', ''),
                })
    
    print(f"Loaded {len(leads)} leads from inquiry form")
    
    # Initialize Gmail
    service = get_gmail_service()
    
    all_pairs = []
    leads_with_conversations = 0
    
    for i, lead in enumerate(leads):
        emails = lead['emails']
        if not emails:
            continue
            
        print(f"\n[{i+1}/{len(leads)}] {lead['company']} - {emails[0][:30]}...")
        
        # Search for threads with any of the lead's emails
        all_threads = []
        for email in emails:
            threads = search_all_threads(service, email, max_results=10)
            all_threads.extend(threads)
        
        # Deduplicate threads
        seen_thread_ids = set()
        unique_threads = []
        for t in all_threads:
            if t['id'] not in seen_thread_ids:
                seen_thread_ids.add(t['id'])
                unique_threads.append(t)
        
        if not unique_threads:
            print("  No threads found")
            continue
        
        print(f"  Found {len(unique_threads)} thread(s)")
        
        # Extract conversations from threads
        lead_pairs = []
        for thread in unique_threads[:5]:  # Limit to 5 threads per lead
            pairs = extract_thread_conversations(service, thread['id'], emails)
            lead_pairs.extend(pairs)
        
        if lead_pairs:
            leads_with_conversations += 1
            print(f"  Extracted {len(lead_pairs)} Q&A pair(s)")
            
            for pair in lead_pairs:
                all_pairs.append({
                    'lead': lead,
                    **pair
                })
    
    # Deduplicate pairs by customer question (first 200 chars)
    unique_pairs = []
    seen = set()
    for pair in all_pairs:
        key = pair['customer_question'][:200]
        if key not in seen:
            seen.add(key)
            unique_pairs.append(pair)
    
    print("\n" + "=" * 70)
    print("RESULTS")
    print("=" * 70)
    print(f"Total leads: {len(leads)}")
    print(f"Leads with conversations: {leads_with_conversations}")
    print(f"Total Q&A pairs (before dedup): {len(all_pairs)}")
    print(f"Unique Q&A pairs: {len(unique_pairs)}")
    
    # Save results
    output = {
        'extracted_at': datetime.now().isoformat(),
        'total_leads': len(leads),
        'leads_with_conversations': leads_with_conversations,
        'total_pairs': len(unique_pairs),
        'pairs': unique_pairs,
    }
    
    output_path = PROJECT_ROOT / 'data/training/inquiry_form_conversations_v2.json'
    with open(output_path, 'w') as f:
        json.dump(output, f, indent=2, default=str)
    
    print(f"\nSaved to {output_path}")
    
    # Show sample
    if unique_pairs:
        print("\n" + "=" * 70)
        print("SAMPLE CONVERSATIONS (first 3)")
        print("=" * 70)
        for pair in unique_pairs[:3]:
            lead = pair['lead']
            print(f"\n--- {lead['company']} ({lead['emails'][0] if lead['emails'] else 'N/A'}) ---")
            print(f"Interest: {lead.get('interest', 'N/A')}")
            print(f"Subject: {pair['subject'][:60]}...")
            print(f"\nCustomer: {pair['customer_question'][:200]}...")
            print(f"\nRushabh: {pair['rushabh_response'][:200]}...")


if __name__ == "__main__":
    main()
