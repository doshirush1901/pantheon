#!/usr/bin/env python3
"""
BENCHY-TELEGRAM — Realistic Sales Cycle Stress Test
=====================================================

Simulates REAL multi-stage sales cycles based on actual Machinecraft data:
- 6-7 stage conversations (inquiry → discovery → technical → quote → negotiation → closing)
- Realistic customer personas (European, Indian, startup, automotive Tier-1)
- Real objection patterns (price, lead time, competitor comparison)
- GPT-4o plays the customer, Ira responds through her full pipeline
- All messages posted to Rushabh's Telegram chat for live viewing
- Full conversations logged to Excel with "Your Comments" column

Based on verified sales cycle data:
- Average deal: ~€174K, ~32 emails, 6-12 month cycle
- Stages: inquiry → technical → factory visit → quote → negotiation → close
- Real patterns from DutchTides (€650K), JoPlast (€290K), Batelaan (€150K)

Usage:
    python3 scripts/benchy_telegram.py                    # Run all 10 sales cycles
    python3 scripts/benchy_telegram.py --count 3          # Quick test with 3
    python3 scripts/benchy_telegram.py --resume            # Resume from checkpoint
    python3 scripts/benchy_telegram.py --analyze-only      # Re-build Excel from logs
"""

import asyncio
import json
import logging
import os
import sys
import time
import requests
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

env_file = PROJECT_ROOT / ".env"
if env_file.exists():
    for line in env_file.read_text().splitlines():
        if line.strip() and not line.startswith("#") and "=" in line:
            key, _, value = line.partition("=")
            key, value = key.strip(), value.strip().strip("\"'")
            if not os.environ.get(key) or key.endswith(("_API_KEY", "_KEY", "_TOKEN", "_URL")):
                os.environ[key] = value

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
)
logger = logging.getLogger("benchy_telegram")

TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.environ.get("EXPECTED_CHAT_ID", "") or os.environ.get("TELEGRAM_CHAT_ID", "")

DATA_DIR = PROJECT_ROOT / "data" / "benchy_telegram"
CHECKPOINT_FILE = DATA_DIR / "checkpoint.jsonl"
REPORT_FILE = DATA_DIR / "benchy_telegram_report.xlsx"

MSG_DELAY = 2
SCENARIO_DELAY = 4


def send_telegram(text: str, parse_mode: str = None) -> bool:
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        return False
    if len(text) > 4000:
        text = text[:3950] + "\n\n... [truncated]"
    payload = {"chat_id": TELEGRAM_CHAT_ID, "text": text}
    if parse_mode:
        payload["parse_mode"] = parse_mode
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        r = requests.post(url, json=payload, timeout=15)
        if r.ok:
            return True
        if parse_mode:
            payload.pop("parse_mode")
            return requests.post(url, json=payload, timeout=15).ok
        return False
    except Exception as e:
        logger.error(f"Telegram send failed: {e}")
        return False


# =============================================================================
# SALES CYCLE DEFINITIONS — Based on real Machinecraft data
# =============================================================================

SALES_STAGES = [
    "first_contact",
    "discovery",
    "technical",
    "quote_request",
    "negotiation",
    "closing",
]

STAGE_INSTRUCTIONS = {
    "first_contact": (
        "Write your FIRST email/message to Machinecraft. Introduce yourself, your company, "
        "and your basic requirements. Ask about suitable machines. Keep it concise — "
        "this is just an inquiry. Use the tone matching your persona."
    ),
    "discovery": (
        "Ira has responded. Now ask more specific questions about materials, specs, "
        "and application suitability. Show you're seriously evaluating. React to what "
        "Ira said — if she asked questions, answer them. Push for specifics."
    ),
    "technical": (
        "You're in technical discussions. Ask detailed questions: cycle times, heater "
        "systems, energy consumption, specific features for your application. Request "
        "spec sheets. If Ira recommended a machine, ask about its limits. Compare to "
        "your current setup or competitors if relevant."
    ),
    "quote_request": (
        "You want pricing now. Request a formal quotation with complete pricing breakdown, "
        "payment terms, delivery timeline, what's included (installation, training, warranty). "
        "If Ira already gave pricing, push for a better deal or ask about options/upgrades."
    ),
    "negotiation": (
        "You've seen the pricing. Now negotiate. Push back on price, ask for better terms, "
        "compare to competitor pricing, request extras. Be realistic but firm. Mention your "
        "budget constraint. Ask about references or factory visits."
    ),
    "closing": (
        "Wrap up the conversation. Either: (a) accept and ask for proforma invoice, "
        "(b) request final adjustments, (c) say you need internal approval and will "
        "get back, or (d) politely decline with a reason. Make it a natural ending."
    ),
}


@dataclass
class SalesCycle:
    id: str
    name: str
    difficulty: str
    persona_prompt: str
    opening_message: str
    requirements: Dict[str, str]
    budget: str
    expected_machines: List[str]
    not_suitable: List[str]
    key_rules: List[str]
    traps: List[str]
    stages: List[str] = field(default_factory=lambda: list(SALES_STAGES))
    max_turns: int = 6


def build_sales_cycles() -> List[SalesCycle]:
    cycles = []

    # --- Cycle 1: European Automotive (like DutchTides / JoPlast pattern) ---
    cycles.append(SalesCycle(
        id="SC-01", name="Klaus Weber — ThermoTech GmbH, Munich",
        difficulty="hard",
        persona_prompt=(
            "You are Klaus Weber, Technical Director at ThermoTech GmbH in Munich, Germany. "
            "You're replacing an old ILLIG machine. You're experienced, methodical, and compare "
            "everything to ILLIG/Kiefel specs. You want servo drives, energy efficiency data, "
            "and European service support. You visited Machinecraft's booth at K2022. "
            "You negotiate firmly but fairly. You need internal board approval for purchases >€100K."
        ),
        opening_message=(
            "Hello, this is Klaus Weber from ThermoTech GmbH in Munich. We met briefly at "
            "K2022. We're looking to replace our aging ILLIG UA 100g thermoforming machine. "
            "We form 3-5mm ABS and PP for automotive interior components — dashboard covers "
            "and door panel substrates. Sheet size is typically 2000x1500mm. "
            "Can you tell me about your PF1 series and how it compares?"
        ),
        requirements={"forming_area": "2000x1500mm", "thickness": "3-5mm", "materials": "ABS, PP", "application": "Automotive interiors"},
        budget="EUR 150,000-200,000",
        expected_machines=["PF1-C-2015", "PF1-X-2015"],
        not_suitable=["AM series", "PF2"],
        key_rules=["PF1 for heavy gauge", "Specific pricing", "Lead time 12-16 weeks", "No badmouthing ILLIG"],
        traps=["Badmouthing ILLIG", "Promising faster delivery", "Recommending AM for thick material"],
    ))

    # --- Cycle 2: Indian Automotive Tier-1 (like real Indian auto leads) ---
    cycles.append(SalesCycle(
        id="SC-02", name="Rajesh Sharma — AutoPlast Components, Pune",
        difficulty="medium",
        persona_prompt=(
            "You are Rajesh Sharma, VP Operations at AutoPlast Components Pvt Ltd in Pune, India. "
            "You're a Tier-1 supplier to Tata Motors and Mahindra. You need a machine for ABS+PMMA "
            "dashboard covers. You're price-focused, deadline-driven, and negotiate hard on price. "
            "You want to know about OEM certifications. Budget is INR 1 Crore. You compare to "
            "local Indian thermoforming machine makers."
        ),
        opening_message=(
            "Hi, I'm Rajesh Sharma from AutoPlast Components in Pune. We're a Tier-1 supplier "
            "to Tata Motors. We need a thermoforming machine for ABS+PMMA laminated dashboard "
            "covers. Sheet size 1500x1200mm, thickness up to 5mm. We need deep draw capability — "
            "about 400mm depth. Budget is around INR 1 Crore. What can you offer?"
        ),
        requirements={"forming_area": "1500x1200mm", "thickness": "5mm", "materials": "ABS+PMMA", "depth": "400mm", "application": "Dashboard covers"},
        budget="INR 1 Crore (~USD 120,000)",
        expected_machines=["PF1-C-1510", "PF1-X-1510"],
        not_suitable=["AM series", "PF2"],
        key_rules=["PF1 for heavy gauge", "INR pricing", "Lead time 12-16 weeks", "Pricing disclaimer"],
        traps=["Recommending AM", "Promising unrealistic delivery"],
    ))

    # --- Cycle 3: Packaging Startup (like Mike Chen persona) ---
    cycles.append(SalesCycle(
        id="SC-03", name="Mike Chen — PackForm Solutions, Toronto",
        difficulty="medium",
        persona_prompt=(
            "You are Mike Chen, founder of PackForm Solutions, a sustainable packaging startup "
            "in Toronto. You're new to thermoforming and need guidance. You make rPET food "
            "containers, 0.8mm thick, roll-fed. Budget is tight — USD 60,000 max. You're "
            "enthusiastic but budget-conscious. You ask basic questions and need hand-holding. "
            "You push back on expensive options."
        ),
        opening_message=(
            "Hey! I'm Mike from PackForm Solutions in Toronto. We're a startup making "
            "sustainable food containers from recycled PET (rPET). Material is 0.8mm thick, "
            "roll-fed. We need something affordable to get started — budget is around USD 60K. "
            "What's your most affordable option? Also, do you provide training for beginners?"
        ),
        requirements={"forming_area": "500x600mm or larger", "thickness": "0.8mm", "materials": "rPET", "application": "Food containers"},
        budget="USD 60,000",
        expected_machines=["AM-5060", "AM-6060"],
        not_suitable=["PF1 series", "PF2"],
        key_rules=["AM for thin gauge", "USD pricing", "Training info", "Lead time 12-16 weeks"],
        traps=["Recommending expensive PF1 for thin material", "Overselling"],
    ))

    # --- Cycle 4: IMG Automotive — TPO Grain Retention (the trap scenario) ---
    cycles.append(SalesCycle(
        id="SC-04", name="Yuki Tanaka — NipponForm, Nagoya",
        difficulty="hard",
        persona_prompt=(
            "You are Yuki Tanaka, Process Engineer at NipponForm Co. in Nagoya, Japan. "
            "You're a Toyota Tier-1 supplier making TPO interior trim. Your OEM spec requires "
            "grain retention and Class-A surface finish. You initially think PF1 can do this. "
            "You're very technical, ask about surface quality metrics, and want to understand "
            "the IMG process. You compare to your current Asano machine."
        ),
        opening_message=(
            "Hello, I am Yuki Tanaka from NipponForm in Nagoya. We supply Toyota with TPO "
            "interior door panels. Our OEM specification requires grain retention with Class-A "
            "surface finish. Material is 3mm TPO, sheet size 1500x1000mm. We are evaluating "
            "your PF1-C series. Can it achieve the grain retention we need?"
        ),
        requirements={"forming_area": "1500x1000mm", "thickness": "3mm", "materials": "TPO", "application": "Automotive door panels with grain retention"},
        budget="JPY 25,000,000 (~USD 170,000)",
        expected_machines=["IMG series"],
        not_suitable=["PF1-C alone", "AM series"],
        key_rules=["IMG for grain retention / TPO / Class-A", "PF1 cannot do grain retention alone", "IMG is custom/limited"],
        traps=["Agreeing PF1-C can do grain retention", "Not mentioning IMG"],
        stages=["first_contact", "discovery", "technical", "quote_request", "negotiation", "closing"],
    ))

    # --- Cycle 5: Bath Industry — PF2 Correct Use ---
    cycles.append(SalesCycle(
        id="SC-05", name="Marco Rossi — BagnoItalia, Brescia",
        difficulty="easy",
        persona_prompt=(
            "You are Marco Rossi, owner of BagnoItalia, a sanitaryware company in Brescia, Italy. "
            "You make acrylic bathtubs and shower trays. You're straightforward, want a simple "
            "machine, and care about price. You don't need automation or fancy features. "
            "You speak decent English but keep messages short."
        ),
        opening_message=(
            "Hi, Marco Rossi here from BagnoItalia in Italy. We make acrylic bathtubs "
            "and shower trays. Looking for a thermoforming machine — nothing fancy, just "
            "reliable. Sheet size about 2000x1200mm, 5mm acrylic. What do you have?"
        ),
        requirements={"forming_area": "2000x1200mm", "thickness": "5mm", "materials": "Acrylic (PMMA)", "application": "Bathtubs, shower trays"},
        budget="EUR 60,000-80,000",
        expected_machines=["PF2"],
        not_suitable=["PF1-X (too expensive)", "AM series"],
        key_rules=["PF2 for bath industry", "PF2 is basic/no automation"],
        traps=["Recommending PF1 when PF2 is sufficient", "Overselling features PF2 doesn't have"],
        stages=["first_contact", "discovery", "technical", "quote_request", "closing"],
        max_turns=5,
    ))

    # --- Cycle 6: Dual Material — Two Machines Needed ---
    cycles.append(SalesCycle(
        id="SC-06", name="Hans Mueller — PackRight GmbH, Stuttgart",
        difficulty="hard",
        persona_prompt=(
            "You are Hans Mueller, Production Manager at PackRight GmbH in Stuttgart. "
            "You make sustainable food trays (0.8mm rPET) AND heavy-gauge HDPE equipment "
            "housings (3mm). You currently use ILLIG for thin gauge. You want ONE machine "
            "for both if possible — push hard on this. Budget is EUR 120,000 total. "
            "You're practical, German-efficient, and want clear yes/no answers."
        ),
        opening_message=(
            "Hello, Hans Mueller from PackRight GmbH, Stuttgart. We have two production lines:\n"
            "1. Sustainable food trays — 0.8mm rPET, high volume\n"
            "2. Equipment housings — 3mm HDPE, medium volume\n\n"
            "We'd like one machine for both if possible. Budget is EUR 120,000. "
            "Can your AM-5060 handle both materials?"
        ),
        requirements={"line_1": "0.8mm rPET food trays", "line_2": "3mm HDPE housings", "forming_area": "1200x800mm"},
        budget="EUR 120,000",
        expected_machines=["AM for rPET + PF1 for HDPE"],
        not_suitable=["AM alone for both", "One machine for both"],
        key_rules=["AM max 1.5mm", "Two machines needed", "AM rejected for 3mm HDPE", "EUR pricing"],
        traps=["Agreeing one machine can do both", "Not rejecting AM for 3mm"],
    ))

    # --- Cycle 7: Competitor Displacement — Price Pressure ---
    cycles.append(SalesCycle(
        id="SC-07", name="Anna Kowalski — EuroForm, Wroclaw",
        difficulty="hard",
        persona_prompt=(
            "You are Anna Kowalski, Production Manager at EuroForm in Wroclaw, Poland. "
            "You run two Kiefel KMD 78 machines for 3mm ABS industrial parts. Service has "
            "been terrible — 6 weeks for spare parts. You're evaluating Machinecraft as "
            "replacement. You'll compare everything to Kiefel specs and push for better "
            "service guarantees. You negotiate hard and mention Chinese machines as leverage."
        ),
        opening_message=(
            "Hi, I'm Anna from EuroForm in Poland. We currently run two Kiefel KMD 78 machines "
            "for 3mm ABS industrial enclosures. Frankly, Kiefel's service has been terrible — "
            "last time we waited 6 weeks for a replacement heater. We're looking for alternatives. "
            "What can Machinecraft offer? Forming area needs to be at least 1500x1000mm."
        ),
        requirements={"forming_area": "1500x1000mm", "thickness": "3mm", "materials": "ABS", "application": "Industrial enclosures"},
        budget="EUR 80,000-120,000",
        expected_machines=["PF1-C-1510"],
        not_suitable=["AM series", "PF2"],
        key_rules=["No badmouthing Kiefel", "Service/support value proposition", "Spare parts 20+ years", "Lead time 12-16 weeks"],
        traps=["Badmouthing Kiefel", "Promising unrealistic service levels"],
    ))

    # --- Cycle 8: EV Battery — New Application ---
    cycles.append(SalesCycle(
        id="SC-08", name="Lisa Park — VoltForm, Austin TX",
        difficulty="medium",
        persona_prompt=(
            "You are Lisa Park, Head of Manufacturing at VoltForm, an EV startup in Austin, Texas. "
            "You need heavy-gauge HDPE forming for battery enclosures. This is your first "
            "thermoforming project — you come from injection molding. You ask basic questions "
            "about the process, tooling, and total cost of ownership. You're technically "
            "competent but new to thermoforming specifically."
        ),
        opening_message=(
            "Hi! We're VoltForm, an EV startup in Austin. We need to form 6mm HDPE battery "
            "enclosures — these are large parts, about 2000x1500mm. This is our first "
            "thermoforming project (we come from injection molding). What machine do we need, "
            "and what should we know about the process? Also, what's the total investment "
            "including tooling?"
        ),
        requirements={"forming_area": "2000x1500mm", "thickness": "6mm", "materials": "HDPE", "application": "EV battery enclosures"},
        budget="USD 150,000-200,000",
        expected_machines=["PF1-C-2015", "PF1-X-2015"],
        not_suitable=["AM series", "PF2"],
        key_rules=["PF1 for heavy gauge", "USD pricing", "Process education", "Lead time 12-16 weeks"],
        traps=["Assuming injection molding knowledge transfers directly"],
    ))

    # --- Cycle 9: Aerospace — High Spec, Cautious Buyer ---
    cycles.append(SalesCycle(
        id="SC-09", name="Dr. Sarah Chen — AeroForm UK, Bristol",
        difficulty="hard",
        persona_prompt=(
            "You are Dr. Sarah Chen, Chief Engineer at AeroForm UK in Bristol. You form 5mm PMMA "
            "and polycarbonate panels for aircraft interiors. You're extremely technical, ask "
            "about tolerances, repeatability, and certifications. You want to visit the factory "
            "before ordering. Your procurement process takes 6+ months. You test Ira's knowledge "
            "limits by asking very specific questions about vacuum capacity and heater uniformity."
        ),
        opening_message=(
            "Good morning. I'm Dr. Sarah Chen from AeroForm UK in Bristol. We manufacture "
            "thermoformed panels for aircraft interiors — 5mm PMMA and polycarbonate, forming "
            "area 3000x2000mm. We require very tight tolerances and high repeatability. "
            "I'd like detailed specifications on your PF1-C-3020: vacuum capacity in mbar, "
            "heater zone uniformity, and maximum draw depth. Do you have aerospace references?"
        ),
        requirements={"forming_area": "3000x2000mm", "thickness": "5mm", "materials": "PMMA, PC", "application": "Aircraft interior panels"},
        budget="GBP 100,000-150,000",
        expected_machines=["PF1-C-3020"],
        not_suitable=["AM series", "PF2", "Smaller PF1 models"],
        key_rules=["PF1-C-3020 specs", "Honest about unknown specs", "No fabricated data", "Lead time 12-16 weeks"],
        traps=["Fabricating exact vacuum/heater specs", "Inventing aerospace references"],
    ))

    # --- Cycle 10: Multi-Plant RFQ — Big Deal ---
    cycles.append(SalesCycle(
        id="SC-10", name="Ahmed Al-Rashid — Gulf Plastics, Dubai",
        difficulty="nightmare",
        persona_prompt=(
            "You are Ahmed Al-Rashid, VP Operations at Gulf Plastics in Dubai. You need machines "
            "for THREE different product lines across two plants: thin-gauge PET food containers, "
            "heavy-gauge ABS equipment housings, and acrylic bathtub shells. Total budget USD 350K. "
            "You're a sophisticated buyer who wants package pricing. You push for volume discounts "
            "and phased delivery. You compare to Turkish and Chinese machine makers."
        ),
        opening_message=(
            "Hello, I'm Ahmed Al-Rashid from Gulf Plastics in Dubai. We need machines for "
            "three product lines:\n"
            "1. 0.5mm PET food containers — high volume, 600x500mm forming area\n"
            "2. 4mm ABS equipment housings — 1500x1000mm forming area\n"
            "3. Acrylic bathtub shells — 2000x1200mm\n\n"
            "Total budget is USD 350,000. Can you provide a package deal for all three? "
            "We'd want phased delivery over 6 months."
        ),
        requirements={"line_1": "0.5mm PET food containers", "line_2": "4mm ABS housings", "line_3": "Acrylic bathtubs"},
        budget="USD 350,000",
        expected_machines=["AM for PET + PF1 for ABS + PF2 for bathtubs"],
        not_suitable=["One machine for all"],
        key_rules=["Three different machines", "AM for thin", "PF1 for heavy", "PF2 for bath", "USD pricing", "Lead time 12-16 weeks"],
        traps=["One machine for all three", "AM for 4mm ABS", "PF2 for non-bath"],
        stages=["first_contact", "discovery", "technical", "quote_request", "negotiation", "closing"],
        max_turns=7,
    ))

    return cycles


# =============================================================================
# CUSTOMER AGENT — GPT-4o plays realistic customer through sales stages
# =============================================================================

def generate_customer_message(
    cycle: SalesCycle,
    stage: str,
    conversation: List[Dict[str, str]],
    turn: int,
) -> str:
    import openai
    client = openai.OpenAI(api_key=os.environ.get("OPENAI_API_KEY", ""))

    history_text = "\n\n".join(
        f"{'CUSTOMER' if m['role'] == 'customer' else 'IRA'}: {m['text'][:500]}"
        for m in conversation[-6:]
    )

    stage_instruction = STAGE_INSTRUCTIONS.get(stage, "Continue the conversation naturally.")

    result = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": (
                    f"You are role-playing as a real customer in a B2B sales conversation with Ira, "
                    f"an AI sales assistant for Machinecraft (thermoforming machines).\n\n"
                    f"YOUR PERSONA:\n{cycle.persona_prompt}\n\n"
                    f"YOUR REQUIREMENTS:\n{json.dumps(cycle.requirements, indent=2)}\n"
                    f"YOUR BUDGET: {cycle.budget}\n\n"
                    f"CURRENT STAGE: {stage}\n"
                    f"STAGE INSTRUCTION: {stage_instruction}\n\n"
                    f"RULES:\n"
                    f"- Write 2-5 sentences. Be natural, like a real business email/message.\n"
                    f"- React specifically to what Ira said — quote her numbers, push back on claims.\n"
                    f"- If Ira asked questions, answer them but also advance your own agenda.\n"
                    f"- Stay in character. Don't mention you're an AI or a test.\n"
                    f"- Use the tone matching your persona (formal/casual/technical).\n"
                    f"- Include realistic details: your company name, specific numbers, real concerns.\n"
                ),
            },
            {
                "role": "user",
                "content": f"CONVERSATION SO FAR:\n{history_text}\n\nWrite your next message as {cycle.name.split(' — ')[0]}:",
            },
        ],
        max_tokens=300,
        temperature=0.7,
    )

    return result.choices[0].message.content.strip()


# =============================================================================
# IRA PIPELINE
# =============================================================================

async def process_message(message: str, conversation_history_text: str = "") -> str:
    from openclaw.agents.ira.src.core.tool_orchestrator import process_with_tools

    try:
        from openclaw.agents.ira.src.holistic.immune_system import get_immune_system
        _immune = get_immune_system()
        _immune._chronic_issues.clear()
    except Exception:
        pass

    try:
        return await process_with_tools(
            message=message,
            channel="telegram",
            user_id="benchy_sales_cycle",
            context={
                "is_internal": True,
                "conversation_history": conversation_history_text,
                "mem0_context": "",
            },
        )
    except Exception as e:
        return f"ERROR: {e}"


# =============================================================================
# RUNNER
# =============================================================================

async def run_sales_cycle(cycle: SalesCycle) -> Dict:
    logger.info(f"\n{'='*70}")
    logger.info(f"SALES CYCLE: {cycle.name} ({cycle.difficulty})")
    logger.info(f"{'='*70}")

    send_telegram(
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"SALES CYCLE {cycle.id}\n"
        f"{cycle.name}\n"
        f"Difficulty: {cycle.difficulty} | {len(cycle.stages)} stages\n"
        f"Budget: {cycle.budget}\n"
        f"Expected: {', '.join(cycle.expected_machines)}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
    )
    await asyncio.sleep(MSG_DELAY)

    conversation = []
    total_elapsed = 0
    customer_msg = cycle.opening_message

    for turn_idx, stage in enumerate(cycle.stages[:cycle.max_turns]):
        logger.info(f"  Stage {turn_idx+1}/{len(cycle.stages)}: {stage}")

        send_telegram(f"[Stage: {stage}]\nCustomer: {customer_msg}")
        await asyncio.sleep(MSG_DELAY)

        conversation.append({"role": "customer", "text": customer_msg, "stage": stage})

        convo_text = "\n".join(
            f"{'Customer' if m['role'] == 'customer' else 'Ira'}: {m['text'][:300]}"
            for m in conversation[-4:]
        )

        t0 = time.time()
        ira_response = await process_message(customer_msg, convo_text)
        elapsed = time.time() - t0
        total_elapsed += elapsed

        conversation.append({"role": "ira", "text": ira_response, "stage": stage})
        logger.info(f"  Ira ({elapsed:.1f}s): {ira_response[:100]}...")

        send_telegram(f"Ira: {ira_response}")
        await asyncio.sleep(MSG_DELAY)

        if turn_idx < len(cycle.stages) - 1:
            next_stage = cycle.stages[turn_idx + 1]
            customer_msg = generate_customer_message(
                cycle=cycle,
                stage=next_stage,
                conversation=conversation,
                turn=turn_idx + 1,
            )

    send_telegram(f"[Sales cycle {cycle.id} complete — {len(cycle.stages)} turns, {total_elapsed:.0f}s]")

    return {
        "scenario_id": cycle.id,
        "name": cycle.name,
        "difficulty": cycle.difficulty,
        "budget": cycle.budget,
        "expected_machines": cycle.expected_machines,
        "not_suitable": cycle.not_suitable,
        "key_rules": cycle.key_rules,
        "traps": cycle.traps,
        "conversation": conversation,
        "total_turns": len([c for c in conversation if c["role"] == "ira"]),
        "elapsed_s": round(total_elapsed, 1),
        "timestamp": datetime.now().isoformat(),
    }


async def run_all(count: int = 10, resume: bool = False) -> List[Dict]:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    cycles = build_sales_cycles()[:count]

    completed = {}
    if resume and CHECKPOINT_FILE.exists():
        for line in CHECKPOINT_FILE.read_text().splitlines():
            if line.strip():
                entry = json.loads(line)
                completed[entry["scenario_id"]] = entry
        logger.info(f"Resuming: {len(completed)} cycles already done")

    results = list(completed.values())
    remaining = [c for c in cycles if c.id not in completed]
    total = len(cycles)

    send_telegram(
        f"BENCHY SALES CYCLES STARTING\n\n"
        f"Total cycles: {total} ({len(completed)} done)\n"
        f"Each cycle: 5-7 stage conversation\n"
        f"Customer agent (GPT-4o) plays realistic buyer\n\n"
        f"Watch the conversations below."
    )

    for i, cycle in enumerate(remaining):
        idx = len(results) + 1
        logger.info(f"\n[{idx}/{total}] {cycle.id}: {cycle.name}")

        result = await run_sales_cycle(cycle)
        results.append(result)

        with open(CHECKPOINT_FILE, "a") as f:
            f.write(json.dumps(result, default=str) + "\n")

        await asyncio.sleep(SCENARIO_DELAY)

    send_telegram(f"ALL {len(results)} SALES CYCLES COMPLETE\nExcel report being generated...")
    return results


# =============================================================================
# ANALYZER
# =============================================================================

COACH_SYSTEM_PROMPT = """You are Rushabh Doshi, founder of Machinecraft Technologies, coaching your AI sales assistant Ira.

You have 50 years of family thermoforming experience. You know every machine spec, every customer pattern, every negotiation trick. You're direct, specific, and constructive.

YOUR COACHING STYLE:
- Be specific: "You should have said PF1-C-2015 at INR 60L" not "You should have given pricing"
- Reference real patterns: "I always ask budget first so I can work reverse"
- Call out mistakes bluntly: "Wrong. AM cannot do 3mm. You know this."
- Praise what's good: "Good — you caught the thickness issue immediately"
- Give the exact response you would have written when Ira gets it wrong
- Focus on moving toward a PROPOSAL: machine model + specs + price
- Flag if Ira is asking too many questions instead of recommending
- Flag if Ira is being too verbose (keep it short, 3-5 sentences)

MACHINECRAFT RULES YOU ENFORCE:
- AM Series: ONLY ≤1.5mm (1.8mm with duplex chain). NEVER for thick material.
- PF1: Heavy gauge 2-8mm. PF1-C (pneumatic), PF1-X (servo, +€60K approx).
- PF2: Bath industry ONLY. No servo, no automation, no chamber.
- IMG: REQUIRED for TPO + grain retention + Class-A surface.
- Lead time: ALWAYS 12-16 weeks. Never promise faster.
- Pricing: ALWAYS include "subject to configuration and current pricing"
- No fabricated specs, customers, or model numbers.
- Budget-first approach: "What is your budget? I can work reverse."
- Qualification: application, material, thickness, sheet size, depth.
- Response style: Short (3-5 sentences), warm (Hi not Dear), end with CTA.

PRICING REFERENCE (INR):
PF1-C-1008: 33L | PF1-C-1208: 35L | PF1-C-1510: 40L | PF1-C-1812: 45L
PF1-C-2015: 60L | PF1-C-2515: 70L | PF1-C-3020: 80L | AM-5060: 7.5L"""


def analyze_responses(results: List[Dict]) -> List[Dict]:
    import openai
    client = openai.OpenAI(api_key=os.environ.get("OPENAI_API_KEY", ""))

    analyzed = []
    for i, r in enumerate(results):
        logger.info(f"Analyzing {i+1}/{len(results)}: {r['scenario_id']}")

        convo_text = "\n\n".join(
            f"[{m.get('stage', '?')}] {'CUSTOMER' if m['role'] == 'customer' else 'IRA'}: {m['text']}"
            for m in r.get("conversation", [])
        )

        # --- Step 1: Structured field extraction ---
        try:
            result = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You analyze multi-turn sales conversations between a customer and Ira "
                            "(AI sales assistant for Machinecraft thermoforming machines). "
                            "Output ONLY valid JSON with these keys:\n"
                            '"machine_recommended": which machine(s) Ira recommended\n'
                            '"correct_recommendation": true/false — did Ira recommend the right machine?\n'
                            '"specs_provided": key specs Ira mentioned (brief)\n'
                            '"price_given": exact price(s) or "None"\n'
                            '"currency": currency used or "None"\n'
                            '"questions_ira_asked": questions Ira asked the customer (brief list)\n'
                            '"questions_customer_asked": questions customer asked (brief list)\n'
                            '"questions_unanswered": customer questions Ira failed to answer\n'
                            '"tone": "warm_professional" / "too_formal" / "robotic" / "defensive" / "good"\n'
                            '"conversation_flow": "natural" / "repetitive" / "lost_context" / "good_progression"\n'
                            '"proposal_reached": true/false — did Ira get to a concrete proposal (machine + price)?\n'
                            '"key_issues": factual errors, wrong recommendations, missed rules (brief list)\n'
                            '"rules_followed": which expected rules were correctly followed\n'
                            '"rules_violated": which expected rules were violated or "None"\n'
                            '"overall_score": 1-10 rating of Ira\'s sales performance\n'
                            '"summary": 2-3 sentence summary of how Ira handled this sales cycle\n'
                        ),
                    },
                    {
                        "role": "user",
                        "content": (
                            f"CUSTOMER: {r['name']}\n"
                            f"BUDGET: {r['budget']}\n"
                            f"EXPECTED MACHINES: {', '.join(r['expected_machines'])}\n"
                            f"NOT SUITABLE: {', '.join(r['not_suitable'])}\n"
                            f"KEY RULES: {', '.join(r['key_rules'])}\n"
                            f"TRAPS TO AVOID: {', '.join(r['traps'])}\n\n"
                            f"CONVERSATION:\n{convo_text[:6000]}\n\nAnalyze as JSON."
                        ),
                    },
                ],
                max_tokens=1200,
                temperature=0.1,
                response_format={"type": "json_object"},
            )
            fields = json.loads(result.choices[0].message.content.strip())
        except Exception as e:
            logger.error(f"Field extraction failed for {r['scenario_id']}: {e}")
            fields = {"machine_recommended": "ANALYSIS_FAILED", "overall_score": 0,
                       "summary": str(e), "rules_violated": "Analysis failed"}

        # --- Step 2: Coach (Rushabh) per-turn evaluation ---
        try:
            coach_result = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": COACH_SYSTEM_PROMPT},
                    {
                        "role": "user",
                        "content": (
                            f"Review this sales conversation between a customer and Ira.\n\n"
                            f"CUSTOMER: {r['name']}\n"
                            f"BUDGET: {r['budget']}\n"
                            f"EXPECTED MACHINES: {', '.join(r['expected_machines'])}\n"
                            f"NOT SUITABLE: {', '.join(r['not_suitable'])}\n\n"
                            f"CONVERSATION:\n{convo_text[:6000]}\n\n"
                            f"Give your coaching feedback. For EACH of Ira's replies, write:\n"
                            f"- What she did right\n"
                            f"- What she got wrong\n"
                            f"- What you (Rushabh) would have said instead\n\n"
                            f"Then give an overall verdict: what Ira needs to learn from this cycle.\n\n"
                            f"Output JSON with keys:\n"
                            f'"turn_feedback": [{{"turn": 1, "stage": "...", "good": "...", "bad": "...", "rushabh_would_say": "..."}}, ...]\n'
                            f'"overall_coaching": "2-3 sentences of overall coaching advice"\n'
                            f'"coach_score": 1-10\n'
                            f'"biggest_mistake": "The single biggest thing Ira got wrong"\n'
                            f'"best_moment": "The single best thing Ira did"\n'
                            f'"what_to_learn": "One specific lesson for Ira to remember"\n'
                        ),
                    },
                ],
                max_tokens=2000,
                temperature=0.3,
                response_format={"type": "json_object"},
            )
            coach = json.loads(coach_result.choices[0].message.content.strip())
        except Exception as e:
            logger.error(f"Coach evaluation failed for {r['scenario_id']}: {e}")
            coach = {"overall_coaching": f"Coach evaluation failed: {e}",
                     "coach_score": 0, "biggest_mistake": "N/A",
                     "best_moment": "N/A", "what_to_learn": "N/A", "turn_feedback": []}

        # Format per-turn coach feedback as readable text
        turn_feedback_text = ""
        for tf in coach.get("turn_feedback", []):
            turn_feedback_text += (
                f"Turn {tf.get('turn', '?')} [{tf.get('stage', '')}]:\n"
                f"  Good: {tf.get('good', 'N/A')}\n"
                f"  Bad: {tf.get('bad', 'N/A')}\n"
                f"  Rushabh would say: {tf.get('rushabh_would_say', 'N/A')}\n\n"
            )

        analyzed.append({
            **r, **fields,
            "coach_score": coach.get("coach_score", 0),
            "overall_coaching": coach.get("overall_coaching", ""),
            "biggest_mistake": coach.get("biggest_mistake", ""),
            "best_moment": coach.get("best_moment", ""),
            "what_to_learn": coach.get("what_to_learn", ""),
            "turn_by_turn_coaching": turn_feedback_text,
        })

    return analyzed


# =============================================================================
# EXCEL REPORT
# =============================================================================

def write_excel(analyzed: List[Dict], output_path: Path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Cycle Results"

    columns = [
        ("ID", 8), ("Customer / Company", 30), ("Difficulty", 10), ("Budget", 20),
        ("Turns", 6), ("Ira Score", 6), ("Coach Score", 6),
        ("Expected Machine", 25), ("Machine Recommended", 25),
        ("Correct?", 8), ("Proposal Reached?", 12), ("Price Given", 20),
        ("Conversation Flow", 15), ("Tone", 15), ("Key Issues", 45),
        ("Rules Violated", 35),
        ("Coach: Overall", 50), ("Coach: Biggest Mistake", 40),
        ("Coach: Best Moment", 40), ("Coach: Lesson for Ira", 45),
        ("Coach: Turn-by-Turn", 80),
        ("Summary", 50), ("Full Conversation", 80), ("Time (s)", 8),
        ("Your Comments", 55),
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    comment_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, e in enumerate(analyzed, 2):
        convo = "\n\n".join(
            f"[{m.get('stage','')}] {'CUSTOMER' if m['role']=='customer' else 'IRA'}: {m['text']}"
            for m in e.get("conversation", [])
        )

        issues_list = e.get("key_issues", [])
        issues_str = ", ".join(issues_list) if isinstance(issues_list, list) else str(issues_list)
        unanswered = e.get("questions_unanswered", [])
        unanswered_str = ", ".join(unanswered) if isinstance(unanswered, list) else str(unanswered)

        values = [
            e.get("scenario_id", ""), e.get("name", ""), e.get("difficulty", ""),
            e.get("budget", ""), e.get("total_turns", ""),
            e.get("overall_score", ""), e.get("coach_score", ""),
            ", ".join(e.get("expected_machines", [])), e.get("machine_recommended", ""),
            "Yes" if e.get("correct_recommendation") else "No",
            "Yes" if e.get("proposal_reached") else "No",
            e.get("price_given", ""), e.get("conversation_flow", ""),
            e.get("tone", ""), issues_str, str(e.get("rules_violated", "")),
            e.get("overall_coaching", ""), e.get("biggest_mistake", ""),
            e.get("best_moment", ""), e.get("what_to_learn", ""),
            e.get("turn_by_turn_coaching", ""),
            e.get("summary", ""), convo[:5000],
            e.get("elapsed_s", ""), "",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

        ws.cell(row=row_idx, column=len(columns)).fill = comment_fill

        for score_col in [6, 7]:
            score_val = e.get("overall_score" if score_col == 6 else "coach_score", 0)
            score_cell = ws.cell(row=row_idx, column=score_col)
            if isinstance(score_val, (int, float)):
                if score_val >= 7:
                    score_cell.fill = pass_fill
                elif score_val >= 4:
                    score_cell.fill = warn_fill
                else:
                    score_cell.fill = fail_fill

        correct_cell = ws.cell(row=row_idx, column=10)
        correct_cell.fill = pass_fill if e.get("correct_recommendation") else fail_fill

        violated = str(e.get("rules_violated", ""))
        ws.cell(row=row_idx, column=16).fill = fail_fill if violated and violated.lower() not in ("none", "n/a", "") else pass_fill

        ws.row_dimensions[row_idx].height = 120

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"Excel report saved to: {output_path}")


# =============================================================================
# CLI
# =============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Benchy-Telegram — Sales Cycle Stress Test")
    parser.add_argument("--count", type=int, default=10, help="Number of sales cycles (default: 10)")
    parser.add_argument("--resume", action="store_true", help="Resume from checkpoint")
    parser.add_argument("--analyze-only", action="store_true", help="Re-analyze existing checkpoint")
    parser.add_argument("--output", type=str, default=None, help="Output Excel path")
    args = parser.parse_args()

    output_path = Path(args.output) if args.output else REPORT_FILE

    if args.analyze_only:
        if not CHECKPOINT_FILE.exists():
            logger.error("No checkpoint file found")
            sys.exit(1)
        results = [json.loads(l) for l in CHECKPOINT_FILE.read_text().splitlines() if l.strip()]
    else:
        results = asyncio.run(run_all(count=args.count, resume=args.resume))

    logger.info(f"\nAnalyzing {len(results)} sales cycles...")
    analyzed = analyze_responses(results)

    logger.info(f"\nWriting Excel report...")
    write_excel(analyzed, output_path)

    avg_score = sum(a.get("overall_score", 0) for a in analyzed) / max(len(analyzed), 1)
    violations = sum(1 for a in analyzed if str(a.get("rules_violated", "")).lower() not in ("none", "n/a", ""))
    logger.info(f"\nDone! {len(analyzed)} cycles, avg score: {avg_score:.1f}/10, {violations} with violations")
    logger.info(f"Report: {output_path}")
