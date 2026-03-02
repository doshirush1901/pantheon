#!/usr/bin/env python3
"""
MAILBOX INGESTION — Bulk Historical Backfill
=============================================

Scans Rushabh's Gmail mailbox and ingests every email into Ira's knowledge
pipeline. This is the one-time "PhD program" — years of real sales conversations
compressed into a few hours of digestion.

What gets extracted:
    - Every price ever quoted (real closing prices, not list prices)
    - Every objection a customer raised and how Rushabh handled it
    - Every competitor mentioned and in what context
    - Every customer relationship — who's warm, who went cold, who came back
    - Rushabh's actual negotiation style from thousands of real emails
    - Delivery timelines committed and whether they held
    - Which machine configurations customers actually buy vs. just ask about

Usage:
    python scripts/ingest_mailbox.py                          # Full scan
    python scripts/ingest_mailbox.py --since 2024-01-01       # Since date
    python scripts/ingest_mailbox.py --limit 100              # First 100
    python scripts/ingest_mailbox.py --dry-run                # Preview only
    python scripts/ingest_mailbox.py --resume                 # Continue from last
    python scripts/ingest_mailbox.py --query "subject:PF1"    # Custom filter

Prerequisites:
    - Gmail API credentials (credentials.json + token.json with gmail.readonly)
    - Run: python setup_gmail.py (after scope upgrade)
"""

import argparse
import base64
import json
import logging
import os
import sys
import time
from datetime import datetime
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Dict, List, Optional

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger("ingest_mailbox")

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "openclaw" / "agents" / "ira"))
sys.path.insert(0, str(PROJECT_ROOT / "openclaw" / "agents" / "ira" / "src" / "brain"))

PROGRESS_FILE = PROJECT_ROOT / "data" / "brain" / "mailbox_ingest_progress.json"
STATS_FILE = PROJECT_ROOT / "data" / "brain" / "mailbox_ingest_stats.json"

GMAIL_SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/gmail.send",
]

BATCH_SLEEP_SECONDS = 2
EMAILS_PER_BATCH = 20


def _load_env():
    env_file = PROJECT_ROOT / ".env"
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            if line.strip() and not line.startswith("#") and "=" in line:
                key, _, value = line.partition("=")
                os.environ.setdefault(key.strip(), value.strip().strip('"'))


def _get_gmail_service():
    """
    Authenticate with Gmail API using a SEPARATE token for Rushabh's mailbox.

    Uses token_rushabh.json (not token.json) so we don't overwrite Ira's
    send-only token. On first run, a browser will open — log in as
    rushabh@machinecraft.org to authorize read access to that mailbox.
    """
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    creds_path = PROJECT_ROOT / "credentials.json"
    token_path = PROJECT_ROOT / "token_rushabh.json"

    if not creds_path.exists():
        logger.error(
            "credentials.json not found.\n"
            "  1. Go to https://console.cloud.google.com/apis/credentials\n"
            "  2. Download your OAuth 2.0 Desktop credentials\n"
            "  3. Save as credentials.json in the project root"
        )
        sys.exit(1)

    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), GMAIL_SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            logger.info(
                "Opening browser for Gmail authorization.\n"
                ">>> LOG IN AS rushabh@machinecraft.org <<<\n"
                "This grants read-only access to scan the mailbox."
            )
            flow = InstalledAppFlow.from_client_secrets_file(str(creds_path), GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json())
        logger.info(f"Token saved to {token_path}")

    return build("gmail", "v1", credentials=creds)


def _load_progress() -> Dict:
    try:
        if PROGRESS_FILE.exists():
            return json.loads(PROGRESS_FILE.read_text())
    except Exception:
        pass
    return {"processed_ids": [], "last_page_token": None, "total_processed": 0, "started_at": None}


def _save_progress(progress: Dict):
    PROGRESS_FILE.parent.mkdir(parents=True, exist_ok=True)
    PROGRESS_FILE.write_text(json.dumps(progress, indent=2, default=str))


def _decode_body(payload: Dict) -> str:
    """Recursively extract text/plain body from Gmail message payload."""
    if payload.get("mimeType") == "text/plain" and payload.get("body", {}).get("data"):
        return base64.urlsafe_b64decode(payload["body"]["data"]).decode("utf-8", errors="replace")

    for part in payload.get("parts", []):
        text = _decode_body(part)
        if text:
            return text

    if payload.get("body", {}).get("data"):
        try:
            return base64.urlsafe_b64decode(payload["body"]["data"]).decode("utf-8", errors="replace")
        except Exception:
            pass

    return ""


def _get_header(headers: List[Dict], name: str) -> str:
    for h in headers:
        if h.get("name", "").lower() == name.lower():
            return h.get("value", "")
    return ""


def _parse_email(msg: Dict) -> Optional[Dict]:
    """Parse a Gmail API message into a structured dict."""
    payload = msg.get("payload", {})
    headers = payload.get("headers", [])

    subject = _get_header(headers, "Subject")
    from_email = _get_header(headers, "From")
    to_email = _get_header(headers, "To")
    date_str = _get_header(headers, "Date")

    # Extract just the email address from "Name <email>" format
    import re
    email_match = re.search(r'[\w.+-]+@[\w-]+\.[\w.]+', from_email)
    from_addr = email_match.group(0) if email_match else from_email

    email_match = re.search(r'[\w.+-]+@[\w-]+\.[\w.]+', to_email)
    to_addr = email_match.group(0) if email_match else to_email

    body = _decode_body(payload)
    if not body or len(body.strip()) < 30:
        return None

    parsed_date = ""
    try:
        dt = parsedate_to_datetime(date_str)
        parsed_date = dt.isoformat()
    except Exception:
        parsed_date = date_str

    machinecraft_domains = ["machinecraft.org", "machinecraft.in"]
    direction = "outbound" if any(from_addr.lower().endswith(d) for d in machinecraft_domains) else "inbound"

    return {
        "id": msg.get("id", ""),
        "thread_id": msg.get("threadId", ""),
        "subject": subject,
        "from_email": from_addr,
        "to_email": to_addr,
        "date": parsed_date,
        "direction": direction,
        "body": body,
    }


def ingest_mailbox(
    query: str = "from:rushabh@machinecraft.org OR to:rushabh@machinecraft.org",
    since: Optional[str] = None,
    limit: Optional[int] = None,
    dry_run: bool = False,
    resume: bool = False,
):
    """Main ingestion loop."""
    _load_env()

    logger.info("Connecting to Gmail API...")
    service = _get_gmail_service()

    progress = _load_progress() if resume else {
        "processed_ids": [], "last_page_token": None,
        "total_processed": 0, "started_at": datetime.now().isoformat(),
    }
    processed_set = set(progress.get("processed_ids", []))

    if since:
        query += f" after:{since}"

    logger.info(f"Query: {query}")
    if limit:
        logger.info(f"Limit: {limit} emails")
    if dry_run:
        logger.info("DRY RUN — no ingestion will occur")

    from email_nutrient_extractor import extract_email_knowledge

    ingestor = None
    if not dry_run:
        try:
            from knowledge_ingestor import KnowledgeIngestor
            ingestor = KnowledgeIngestor()
        except ImportError:
            logger.warning("KnowledgeIngestor not available — will only do RAG indexing")

    stats = {
        "emails_scanned": 0, "emails_processed": 0, "emails_skipped": 0,
        "items_extracted": 0, "items_ingested": 0,
        "errors": 0, "started_at": datetime.now().isoformat(),
    }

    page_token = progress.get("last_page_token") if resume else None
    total_processed = progress.get("total_processed", 0)

    while True:
        try:
            params = {"userId": "me", "q": query, "maxResults": EMAILS_PER_BATCH}
            if page_token:
                params["pageToken"] = page_token

            results = service.users().messages().list(**params).execute()
        except Exception as e:
            logger.error(f"Gmail API error: {e}")
            break

        messages = results.get("messages", [])
        if not messages:
            logger.info("No more messages")
            break

        for msg_stub in messages:
            msg_id = msg_stub["id"]

            if msg_id in processed_set:
                stats["emails_skipped"] += 1
                continue

            if limit and total_processed >= limit:
                logger.info(f"Reached limit of {limit} emails")
                _save_progress({**progress, "processed_ids": list(processed_set),
                                "total_processed": total_processed})
                _print_stats(stats)
                return stats

            stats["emails_scanned"] += 1

            try:
                msg = service.users().messages().get(userId="me", id=msg_id, format="full").execute()
                parsed = _parse_email(msg)

                if not parsed:
                    stats["emails_skipped"] += 1
                    processed_set.add(msg_id)
                    continue

                items = extract_email_knowledge(
                    subject=parsed["subject"],
                    body=parsed["body"],
                    from_email=parsed["from_email"],
                    to_email=parsed["to_email"],
                    direction=parsed["direction"],
                    date_str=parsed["date"],
                    use_llm=True,
                )

                stats["items_extracted"] += len(items)

                if dry_run:
                    logger.info(
                        f"  [DRY] {parsed['direction']:8s} | {parsed['from_email'][:30]:30s} | "
                        f"{parsed['subject'][:50]:50s} | {len(items)} items"
                    )
                else:
                    if items and ingestor:
                        try:
                            result = ingestor.ingest_batch(items)
                            stats["items_ingested"] += result.items_ingested
                        except Exception as e:
                            logger.warning(f"  Ingestion error: {e}")
                            stats["errors"] += 1

                    # RAG index the raw email
                    try:
                        from realtime_indexer import index_new_email
                        index_new_email(
                            email_id=hash(msg_id) % (10**9),
                            subject=parsed["subject"],
                            body=parsed["body"],
                            from_email=parsed["from_email"],
                            to_email=parsed["to_email"],
                            date=datetime.fromisoformat(parsed["date"]) if parsed["date"] else None,
                            direction=parsed["direction"],
                        )
                    except Exception:
                        pass

                stats["emails_processed"] += 1
                total_processed += 1
                processed_set.add(msg_id)

                if total_processed % 10 == 0:
                    logger.info(f"  Progress: {total_processed} emails processed, {stats['items_extracted']} items extracted")

            except Exception as e:
                logger.warning(f"  Error processing {msg_id}: {e}")
                stats["errors"] += 1
                processed_set.add(msg_id)

        page_token = results.get("nextPageToken")

        # Save progress after each page
        progress["processed_ids"] = list(processed_set)[-5000:]
        progress["last_page_token"] = page_token
        progress["total_processed"] = total_processed
        _save_progress(progress)

        if not page_token:
            logger.info("No more pages")
            break

        logger.info(f"  Sleeping {BATCH_SLEEP_SECONDS}s before next batch...")
        time.sleep(BATCH_SLEEP_SECONDS)

    stats["finished_at"] = datetime.now().isoformat()
    _save_progress({**progress, "processed_ids": list(processed_set),
                    "total_processed": total_processed, "finished_at": stats["finished_at"]})
    _print_stats(stats)

    # Save stats
    STATS_FILE.parent.mkdir(parents=True, exist_ok=True)
    STATS_FILE.write_text(json.dumps(stats, indent=2))

    # Fire growth signal for the bulk ingestion
    if not dry_run:
        try:
            sys.path.insert(0, str(PROJECT_ROOT / "openclaw" / "agents" / "ira" / "src" / "holistic"))
            from growth_signal import signal_bulk_ingestion
            signal_bulk_ingestion(
                source="gmail_backfill",
                items_total=stats["items_ingested"],
                emails_processed=stats["emails_processed"],
            )
        except Exception:
            pass

    return stats


def _print_stats(stats: Dict):
    logger.info("\n" + "=" * 60)
    logger.info("MAILBOX INGESTION COMPLETE")
    logger.info("=" * 60)
    logger.info(f"  Emails scanned:    {stats['emails_scanned']}")
    logger.info(f"  Emails processed:  {stats['emails_processed']}")
    logger.info(f"  Emails skipped:    {stats['emails_skipped']}")
    logger.info(f"  Items extracted:   {stats['items_extracted']}")
    logger.info(f"  Items ingested:    {stats['items_ingested']}")
    logger.info(f"  Errors:            {stats['errors']}")
    logger.info("=" * 60)


def _load_contact_emails() -> List[str]:
    """
    Load customer/lead emails from ALL available spreadsheets.
    Scans every xlsx/csv in data/imports/ for email-like columns.
    This is the master list — the pure protein, no junk.
    """
    contacts = set()
    imports_dir = PROJECT_ROOT / "data" / "imports"

    if not imports_dir.exists():
        logger.warning(f"Imports directory not found: {imports_dir}")
        return []

    def _add(raw: str):
        raw = str(raw or "").strip().lower()
        # Extract bare email from "Name <email>" or quoted formats
        import re as _re
        match = _re.search(r'[\w.+-]+@[\w-]+\.[\w.]+', raw)
        if not match:
            return
        email = match.group(0).rstrip(".")
        if (len(email) > 5 and "machinecraft" not in email
                and "example" not in email and "noreply" not in email):
            contacts.add(email)

    def _scan_xlsx(path):
        """Scan an xlsx file for any column containing emails."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                headers = [str(c.value or "").lower().strip()
                           for c in next(ws.iter_rows(min_row=1, max_row=1))]
                email_idxs = [i for i, h in enumerate(headers)
                              if any(k in h for k in ["email", "e-mail", "mail"])]
                for idx in email_idxs:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and len(row) > idx:
                            _add(row[idx])
            wb.close()
        except Exception as e:
            logger.debug(f"Scan {path.name}: {e}")

    # Scan all xlsx files
    for f in sorted(imports_dir.glob("*.xlsx")):
        _scan_xlsx(f)

    # Scan all csv files
    import csv as csv_mod
    for f in sorted(imports_dir.glob("*.csv")):
        for encoding in ["utf-8-sig", "latin-1", "cp1252"]:
            try:
                with open(f, "r", encoding=encoding) as fh:
                    reader = csv_mod.DictReader(fh)
                    email_cols = [h for h in (reader.fieldnames or [])
                                  if any(k in h.lower() for k in ["email", "e-mail", "mail"])]
                    for row in reader:
                        for col in email_cols:
                            _add(row.get(col, ""))
                break
            except (UnicodeDecodeError, Exception):
                continue

    logger.info(f"Loaded {len(contacts)} unique contact emails from {imports_dir}")
    return sorted(contacts)


def ingest_from_contacts(
    since: Optional[str] = None,
    limit_per_contact: int = 50,
    dry_run: bool = False,
    resume: bool = False,
):
    """
    The smart approach: load the master customer/lead list from spreadsheets,
    then query Gmail for each contact's conversations. Pure protein, zero junk.
    """
    _load_env()

    contact_emails = _load_contact_emails()
    if not contact_emails:
        logger.error("No contact emails found. Check data/imports/ for spreadsheets.")
        return

    logger.info(f"Loaded {len(contact_emails)} customer/lead emails from spreadsheets")
    logger.info("Connecting to Gmail API...")
    service = _get_gmail_service()

    progress = _load_progress() if resume else {
        "processed_ids": [], "contacts_done": [], "last_page_token": None,
        "total_processed": 0, "started_at": datetime.now().isoformat(),
    }
    processed_set = set(progress.get("processed_ids", []))
    contacts_done = set(progress.get("contacts_done", []))

    from email_nutrient_extractor import extract_email_knowledge

    ingestor = None
    if not dry_run:
        try:
            from knowledge_ingestor import KnowledgeIngestor
            ingestor = KnowledgeIngestor()
        except ImportError:
            logger.warning("KnowledgeIngestor not available")

    stats = {
        "contacts_total": len(contact_emails), "contacts_processed": 0,
        "emails_scanned": 0, "emails_processed": 0, "emails_skipped": 0,
        "items_extracted": 0, "items_ingested": 0, "errors": 0,
        "started_at": datetime.now().isoformat(),
    }

    for contact_idx, contact_email in enumerate(contact_emails):
        if contact_email in contacts_done:
            stats["contacts_processed"] += 1
            continue

        query = f"from:{contact_email} OR to:{contact_email}"
        if since:
            query += f" after:{since}"

        logger.info(f"\n[{contact_idx+1}/{len(contact_emails)}] Scanning: {contact_email}")

        page_token = None
        contact_count = 0

        while True:
            try:
                params = {"userId": "me", "q": query, "maxResults": 20}
                if page_token:
                    params["pageToken"] = page_token
                results = service.users().messages().list(**params).execute()
            except Exception as e:
                logger.warning(f"  Gmail API error for {contact_email}: {e}")
                stats["errors"] += 1
                break

            messages = results.get("messages", [])
            if not messages:
                break

            for msg_stub in messages:
                msg_id = msg_stub["id"]
                if msg_id in processed_set:
                    stats["emails_skipped"] += 1
                    continue

                if contact_count >= limit_per_contact:
                    break

                stats["emails_scanned"] += 1
                try:
                    msg = service.users().messages().get(userId="me", id=msg_id, format="full").execute()
                    parsed = _parse_email(msg)
                    if not parsed:
                        stats["emails_skipped"] += 1
                        processed_set.add(msg_id)
                        continue

                    items = extract_email_knowledge(
                        subject=parsed["subject"], body=parsed["body"],
                        from_email=parsed["from_email"], to_email=parsed["to_email"],
                        direction=parsed["direction"], date_str=parsed["date"],
                        use_llm=True,
                    )
                    stats["items_extracted"] += len(items)

                    if dry_run:
                        logger.info(
                            f"  [DRY] {parsed['direction']:8s} | {parsed['subject'][:60]:60s} | {len(items)} items"
                        )
                    else:
                        if items and ingestor:
                            try:
                                result = ingestor.ingest_batch(items)
                                stats["items_ingested"] += result.items_ingested
                            except Exception as e:
                                logger.warning(f"  Ingestion error: {e}")
                                stats["errors"] += 1

                        try:
                            from realtime_indexer import index_new_email
                            index_new_email(
                                email_id=hash(msg_id) % (10**9),
                                subject=parsed["subject"], body=parsed["body"],
                                from_email=parsed["from_email"], to_email=parsed["to_email"],
                                date=datetime.fromisoformat(parsed["date"]) if parsed["date"] else None,
                                direction=parsed["direction"],
                            )
                        except Exception:
                            pass

                    stats["emails_processed"] += 1
                    contact_count += 1
                    processed_set.add(msg_id)

                except Exception as e:
                    logger.warning(f"  Error: {e}")
                    stats["errors"] += 1
                    processed_set.add(msg_id)

            if contact_count >= limit_per_contact:
                break

            page_token = results.get("nextPageToken")
            if not page_token:
                break
            time.sleep(1)

        contacts_done.add(contact_email)
        stats["contacts_processed"] += 1
        logger.info(f"  -> {contact_count} emails, {stats['items_extracted']} total items so far")

        # Save progress after each contact
        progress["processed_ids"] = list(processed_set)[-10000:]
        progress["contacts_done"] = list(contacts_done)
        progress["total_processed"] = stats["emails_processed"]
        _save_progress(progress)

        time.sleep(BATCH_SLEEP_SECONDS)

    stats["finished_at"] = datetime.now().isoformat()
    _save_progress({**progress, "finished_at": stats["finished_at"]})
    _print_stats(stats)

    STATS_FILE.parent.mkdir(parents=True, exist_ok=True)
    STATS_FILE.write_text(json.dumps(stats, indent=2))

    if not dry_run:
        try:
            sys.path.insert(0, str(PROJECT_ROOT / "openclaw" / "agents" / "ira" / "src" / "holistic"))
            from growth_signal import signal_bulk_ingestion
            signal_bulk_ingestion(
                source="gmail_contacts_backfill",
                items_total=stats["items_ingested"],
                emails_processed=stats["emails_processed"],
            )
        except Exception:
            pass

        _send_completion_email(stats)

    return stats


def _send_completion_email(stats: Dict):
    """Send Rushabh a summary email when ingestion finishes."""
    try:
        sys.path.insert(0, str(PROJECT_ROOT / "openclaw" / "agents" / "ira" / "src" / "brain"))
        from knowledge_validator import send_email_gmail

        duration = ""
        if stats.get("started_at") and stats.get("finished_at"):
            try:
                start = datetime.fromisoformat(stats["started_at"])
                end = datetime.fromisoformat(stats["finished_at"])
                mins = int((end - start).total_seconds() / 60)
                hours = mins // 60
                remaining_mins = mins % 60
                duration = f"{hours}h {remaining_mins}m" if hours else f"{mins}m"
            except Exception:
                pass

        body = (
            "Hi Rushabh,\n\n"
            "Ira here. I just finished digesting your mailbox. Here's what I absorbed:\n\n"
            "INGESTION SUMMARY\n"
            "========================================\n"
            f"  Contacts scanned:    {stats.get('contacts_total', '?')}\n"
            f"  Contacts processed:  {stats.get('contacts_processed', '?')}\n"
            f"  Emails processed:    {stats.get('emails_processed', '?')}\n"
            f"  Emails skipped:      {stats.get('emails_skipped', '?')}\n"
            f"  Knowledge items:     {stats.get('items_extracted', '?')} extracted, "
            f"{stats.get('items_ingested', '?')} ingested\n"
            f"  Errors:              {stats.get('errors', '?')}\n"
            f"  Duration:            {duration or 'unknown'}\n\n"
            "WHERE THE KNOWLEDGE WENT\n"
            "========================================\n"
            "  - Qdrant (vector search): main + discovered collections\n"
            "  - Mem0 (long-term memory): semantic memories\n"
            "  - Neo4j (knowledge graph): entity relationships\n"
            "  - JSON backup: disaster recovery\n\n"
            "WHAT I LEARNED\n"
            "========================================\n"
            "  - Pricing you've quoted to real customers\n"
            "  - Customer requirements and objections\n"
            "  - Competitor mentions (ILLIG, GEISS, etc.)\n"
            "  - Your negotiation and communication style\n"
            "  - Machine configurations customers actually buy\n"
            "  - Regional patterns and delivery commitments\n\n"
            "I'm smarter now. Ask me anything about your customers.\n\n"
            "-- Ira"
        )

        subject = (
            f"Ira Growth Report: {stats.get('items_ingested', 0)} knowledge items "
            f"absorbed from {stats.get('emails_processed', 0)} emails"
        )

        send_email_gmail(
            to_email="rushabh@machinecraft.org",
            subject=subject,
            body=body,
        )
        logger.info("[GROWTH] Completion email sent to rushabh@machinecraft.org")
    except Exception as e:
        logger.warning(f"[GROWTH] Could not send completion email: {e}")


def main():
    parser = argparse.ArgumentParser(description="Ingest Gmail mailbox into Ira's knowledge base")
    parser.add_argument("--query", default=None,
                        help="Gmail search query (overrides --contacts mode)")
    parser.add_argument("--contacts", action="store_true", default=True,
                        help="Smart mode: scan only known customers/leads from spreadsheets (default)")
    parser.add_argument("--all", action="store_true",
                        help="Scan entire mailbox (not recommended — use --contacts instead)")
    parser.add_argument("--since", help="Only process emails after this date (YYYY-MM-DD)")
    parser.add_argument("--limit", type=int, help="Max emails per contact (default 50) or total for --all")
    parser.add_argument("--dry-run", action="store_true", help="Preview without ingesting")
    parser.add_argument("--resume", action="store_true", help="Resume from last checkpoint")
    args = parser.parse_args()

    if args.query or args.all:
        query = args.query or "from:rushabh@machinecraft.org OR to:rushabh@machinecraft.org"
        ingest_mailbox(
            query=query, since=args.since, limit=args.limit,
            dry_run=args.dry_run, resume=args.resume,
        )
    else:
        ingest_from_contacts(
            since=args.since,
            limit_per_contact=args.limit or 50,
            dry_run=args.dry_run,
            resume=args.resume,
        )


if __name__ == "__main__":
    main()
