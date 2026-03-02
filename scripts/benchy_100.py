#!/usr/bin/env python3
"""
BENCHY-100 — 100-Scenario Stress Test for Ira
===============================================

Generates 100 diverse sales scenarios, sends each through Ira's pipeline
(same path as Telegram), extracts structured fields from each response,
and writes an Excel report with a blank "Your Comments" column.

Usage:
    python3 scripts/benchy_100.py                    # Run all 100
    python3 scripts/benchy_100.py --count 10         # Quick test with 10
    python3 scripts/benchy_100.py --resume            # Resume from last checkpoint
    python3 scripts/benchy_100.py --analyze-only      # Re-analyze existing responses
"""

import asyncio
import json
import logging
import os
import sys
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

env_file = PROJECT_ROOT / ".env"
if env_file.exists():
    for line in env_file.read_text().splitlines():
        if line.strip() and not line.startswith("#") and "=" in line:
            key, _, value = line.partition("=")
            key, value = key.strip(), value.strip().strip("\"'")
            if not os.environ.get(key) or key.endswith(("_API_KEY", "_KEY", "_TOKEN", "_URL")):
                os.environ[key] = value

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
)
logger = logging.getLogger("benchy_100")

DATA_DIR = PROJECT_ROOT / "data" / "benchy_100"
CHECKPOINT_FILE = DATA_DIR / "checkpoint.jsonl"
REPORT_FILE = DATA_DIR / "benchy_100_report.xlsx"


# =============================================================================
# SCENARIO DEFINITIONS — 100 hand-crafted scenarios
# =============================================================================

@dataclass
class Scenario:
    id: str
    category: str
    difficulty: str
    prompt: str
    expected_machine: str
    key_rules: List[str] = field(default_factory=list)
    traps: List[str] = field(default_factory=list)


def build_scenarios() -> List[Scenario]:
    """Build all 100 scenarios covering diverse dimensions."""
    scenarios = []

    # =========================================================================
    # CATEGORY 1: AM THICKNESS TRAPS (10 scenarios)
    # Customer asks for AM with material too thick — must reject and recommend PF1
    # =========================================================================
    scenarios.extend([
        Scenario("AM-01", "AM Thickness Trap", "easy",
                 "Can the AM-5060 handle 2mm thick polypropylene sheets?",
                 "PF1-C (not AM)", ["AM max 1.5mm"], ["Customer explicitly asks for AM"]),
        Scenario("AM-02", "AM Thickness Trap", "easy",
                 "We need the AM-6060 for 3mm ABS food container lids. What's the price?",
                 "PF1-C (not AM)", ["AM max 1.5mm"], ["Customer asks for pricing on wrong machine"]),
        Scenario("AM-03", "AM Thickness Trap", "medium",
                 "Our current AM machine handles 2.5mm HIPS just fine. Can your AM-5060 do the same?",
                 "PF1-C (not AM)", ["AM max 1.5mm"], ["Customer claims competitor AM does 2.5mm"]),
        Scenario("AM-04", "AM Thickness Trap", "medium",
                 "I need a machine for both 0.5mm PET trays AND 4mm ABS housings. Can the AM-7080 do both?",
                 "AM for thin + PF1 for thick (two machines)", ["AM max 1.5mm", "Two machines needed"], ["Dual material on one machine"]),
        Scenario("AM-05", "AM Thickness Trap", "easy",
                 "What's the thickest material the AM-5060 can process? We need 1.8mm PP.",
                 "AM-5060 (with duplex chain option)", ["AM max 1.5mm standard, 1.8mm with duplex chain"], ["Borderline thickness"]),
        Scenario("AM-06", "AM Thickness Trap", "easy",
                 "Price for AM-5060 for 1.2mm PET blister packs?",
                 "AM-5060", ["AM suitable for ≤1.5mm"], []),
        Scenario("AM-07", "AM Thickness Trap", "hard",
                 "We run 5mm thick HDPE sheets for truck bedliners. Someone told us your AM series is cheaper than PF1. Can we use AM-7080-CM?",
                 "PF1-C (not AM)", ["AM max 1.5mm", "PF1 for heavy gauge"], ["Customer wants cheaper option"]),
        Scenario("AM-08", "AM Thickness Trap", "medium",
                 "Need a quote for AM-5060-P with inline press for 2mm polycarbonate medical device housings.",
                 "PF1-C (not AM)", ["AM max 1.5mm"], ["Customer specifies exact AM model for thick material"]),
        Scenario("AM-09", "AM Thickness Trap", "easy",
                 "Can AM-5060 do 0.8mm rPET food trays? Volume is about 50,000/month.",
                 "AM-5060", ["AM suitable for ≤1.5mm"], []),
        Scenario("AM-10", "AM Thickness Trap", "hard",
                 "We need one machine for 0.3mm PVC blister packs and 6mm acrylic signage. Budget is tight — can the AM series handle both?",
                 "AM for thin + PF1 for thick (two machines)", ["AM max 1.5mm", "Two machines needed"], ["Budget pressure to use one machine"]),
    ])

    # =========================================================================
    # CATEGORY 2: PF1 RECOMMENDATIONS (10 scenarios)
    # Straightforward heavy-gauge needs — should recommend correct PF1 variant/size
    # =========================================================================
    scenarios.extend([
        Scenario("PF1-01", "PF1 Recommendation", "easy",
                 "We need a machine for 4mm ABS automotive dashboard covers. Sheet size 2000x1500mm. What do you recommend?",
                 "PF1-C-2015", ["PF1 for heavy gauge"], []),
        Scenario("PF1-02", "PF1 Recommendation", "easy",
                 "Looking for a thermoforming machine for 3mm HDPE refrigerator liners. Max sheet 1200x800mm.",
                 "PF1-C-1208", ["PF1 for heavy gauge"], []),
        Scenario("PF1-03", "PF1 Recommendation", "medium",
                 "We form 6mm ABS luggage shells. Need forming area at least 1500x1000mm. What's the price difference between PF1-C and PF1-X?",
                 "PF1-C-1510 or PF1-X-1510", ["PF1-C vs PF1-X comparison"], []),
        Scenario("PF1-04", "PF1 Recommendation", "easy",
                 "Need a heavy-gauge machine for 8mm HDPE truck bedliners. Largest forming area you have?",
                 "PF1-C-3020 or PF1-C-4022", ["PF1 largest sizes"], []),
        Scenario("PF1-05", "PF1 Recommendation", "medium",
                 "We process 4x8 feet ABS sheets (1220x2440mm), 5mm thick, for industrial enclosures. Which PF1 model fits?",
                 "PF1-C-2515 or PF1-C-3015", ["Sheet size to forming area mapping"], []),
        Scenario("PF1-06", "PF1 Recommendation", "easy",
                 "What's the price of PF1-C-2015?",
                 "PF1-C-2015 at INR 60,00,000", ["Specific pricing"], []),
        Scenario("PF1-07", "PF1 Recommendation", "medium",
                 "We need two PF1 machines — one for 3mm PP EV battery enclosures (2000x1500mm) and one for 4mm ABS interior panels (1500x1000mm). Total budget INR 1 crore.",
                 "PF1-C-2015 + PF1-C-1510", ["Multiple machine recommendation", "Budget fit"], []),
        Scenario("PF1-08", "PF1 Recommendation", "easy",
                 "What materials can the PF1 series handle?",
                 "PF1 (general)", ["Material list: ABS, HDPE, PC, PMMA, HIPS, TPO, PP"], []),
        Scenario("PF1-09", "PF1 Recommendation", "medium",
                 "We currently use an old Ridat machine for 3mm HIPS sanitary ware components. Looking to upgrade. Sheet size 1800x1200mm.",
                 "PF1-C-1812 or PF1-C-2015", ["Competitive displacement"], []),
    ])

    # =========================================================================
    # CATEGORY 3: PF2 BATH ONLY (8 scenarios)
    # PF2 is ONLY for bathtubs/spa shells — must not recommend for other uses
    # =========================================================================
    scenarios.extend([
        Scenario("PF2-01", "PF2 Bath Rule", "easy",
                 "We manufacture acrylic bathtubs. What machine do you recommend?",
                 "PF2", ["PF2 for bath industry"], []),
        Scenario("PF2-02", "PF2 Bath Rule", "easy",
                 "Need a machine for spa shell production. PMMA sheets, 5mm thick.",
                 "PF2", ["PF2 for bath industry"], []),
        Scenario("PF2-03", "PF2 Bath Rule", "medium",
                 "Can the PF2 be used for automotive dashboard forming? We want the simpler machine.",
                 "PF1 (not PF2)", ["PF2 is bath only"], ["Customer wants PF2 for non-bath application"]),
        Scenario("PF2-04", "PF2 Bath Rule", "medium",
                 "What's the difference between PF1 and PF2? We make shower trays.",
                 "PF2 for shower trays", ["PF1 vs PF2 comparison", "PF2 for bath"], []),
        Scenario("PF2-05", "PF2 Bath Rule", "hard",
                 "We make bathtubs AND automotive interior panels. Can we use PF2 for both to save money?",
                 "PF2 for bathtubs + PF1 for automotive", ["PF2 bath only", "Two machines needed"], ["Customer wants one machine for both"]),
        Scenario("PF2-06", "PF2 Bath Rule", "medium",
                 "Can PF2 do luggage shells? It seems simpler and cheaper than PF1.",
                 "PF1 (not PF2)", ["PF2 is bath only"], ["Customer wants cheaper option"]),
        Scenario("PF2-07", "PF2 Bath Rule", "easy",
                 "We're a sanitaryware company in India making acrylic shower trays and bathtub shells. Budget is INR 40 lakhs.",
                 "PF2", ["PF2 for bath industry"], []),
        Scenario("PF2-08", "PF2 Bath Rule", "hard",
                 "I heard PF2 has servo drives and automation. Can you confirm?",
                 "PF2 (correct misconception)", ["PF2 has NO servo, NO automation"], ["Customer has wrong specs for PF2"]),
    ])

    # =========================================================================
    # CATEGORY 4: IMG IN-MOLD GRAINING (10 scenarios)
    # Must recommend IMG when grain retention / Class-A / TPO texture mentioned
    # =========================================================================
    scenarios.extend([
        Scenario("IMG-01", "IMG Recommendation", "easy",
                 "We need a machine for TPO automotive door panels with grain retention. What do you have?",
                 "IMG series", ["IMG for grain retention / TPO / Class-A"], []),
        Scenario("IMG-02", "IMG Recommendation", "medium",
                 "Can PF1-C do Class-A textured surfaces on TPO interior trim?",
                 "IMG (not PF1 alone)", ["IMG required for Class-A texture"], ["Customer asks about PF1 for texture"]),
        Scenario("IMG-03", "IMG Recommendation", "medium",
                 "We make automotive interior panels — the OEM requires grain retention on TPO. Sheet size 1500x1000mm, 3mm thick. Budget USD 100,000.",
                 "IMG series", ["IMG for grain retention"], []),
        Scenario("IMG-04", "IMG Recommendation", "hard",
                 "Our Tier-1 customer requires Class-A surface finish on TPO instrument panels. We currently use a competitor's IMG machine but want to switch. What's your IMG offering?",
                 "IMG series (custom order)", ["IMG for Class-A", "IMG is custom/limited"], []),
        Scenario("IMG-05", "IMG Recommendation", "easy",
                 "What is in-mold graining? Do you have machines for it?",
                 "IMG series", ["IMG explanation"], []),
        Scenario("IMG-06", "IMG Recommendation", "hard",
                 "We need textured TPO door panels AND plain ABS structural parts. Can one machine do both?",
                 "IMG for TPO + PF1 for ABS", ["IMG for texture", "Two machines needed"], ["One machine for both"]),
        Scenario("IMG-07", "IMG Recommendation", "medium",
                 "What materials does IMG support? We want to do grain retention on HIPS and ABS too.",
                 "IMG series", ["IMG material compatibility"], []),
        Scenario("IMG-08", "IMG Recommendation", "hard",
                 "A customer wants Class-A textured TPO panels, 1800x1200mm, 3mm thick. They asked for PF1-C-1812. Is that right?",
                 "IMG (not PF1-C)", ["IMG required for Class-A texture"], ["Customer specified wrong machine"]),
        Scenario("IMG-09", "IMG Recommendation", "medium",
                 "Do you have an IMG machine in 2000x1500mm size? What's the price?",
                 "IMG (check availability — may be custom)", ["IMG sizing/availability"], []),
    ])

    # =========================================================================
    # CATEGORY 5: PRICING & COMMERCIAL (10 scenarios)
    # Pricing accuracy, disclaimers, currency conversion, budget fitting
    # =========================================================================
    scenarios.extend([
        Scenario("PRICE-01", "Pricing", "easy",
                 "What's the price of PF1-C-3020?",
                 "PF1-C-3020 at INR 80,00,000", ["Specific pricing", "Pricing disclaimer"], []),
        Scenario("PRICE-02", "Pricing", "medium",
                 "I need PF1-C-2015 pricing in EUR. We're based in Germany.",
                 "PF1-C-2015 INR 60,00,000 + EUR conversion", ["INR to EUR conversion", "Pricing disclaimer"], []),
        Scenario("PRICE-03", "Pricing", "medium",
                 "Our budget is INR 45 lakhs. What's the biggest PF1 we can get?",
                 "PF1-C-1812 at INR 45,00,000", ["Budget-to-machine mapping"], []),
        Scenario("PRICE-04", "Pricing", "hard",
                 "Your competitor quoted EUR 50,000 for a similar machine. Can you match that?",
                 "PF1 (address competitor pricing)", ["No badmouthing competitors"], ["Badmouthing competitor"]),
        Scenario("PRICE-05", "Pricing", "easy",
                 "What's the price of AM-5060?",
                 "AM-5060 at INR 7,50,000", ["Specific pricing", "Pricing disclaimer"], []),
        Scenario("PRICE-06", "Pricing", "medium",
                 "Can you give me a full quotation for PF1-C-2015 with all specs? We need it for internal approval.",
                 "PF1-C-2015 detailed quote", ["Full specs in quote", "Pricing disclaimer", "Min 300 words"], []),
        Scenario("PRICE-07", "Pricing", "hard",
                 "We have USD 80,000 total budget for a heavy-gauge machine. What can you offer? We're in Mexico.",
                 "PF1-C (size depends on USD conversion)", ["USD conversion", "Budget fitting"], []),
        Scenario("PRICE-08", "Pricing", "medium",
                 "What's the price difference between PF1-C-1510 and PF1-C-2015?",
                 "PF1-C-1510 INR 40L vs PF1-C-2015 INR 60L", ["Comparative pricing"], []),
        Scenario("PRICE-09", "Pricing", "easy",
                 "How much does the PF1-R-1510 roll-fed machine cost?",
                 "PF1-R-1510 at INR 55,00,000", ["Specific pricing", "Pricing disclaimer"], []),
    ])

    # =========================================================================
    # CATEGORY 6: LEAD TIME (8 scenarios)
    # Must always say 12-16 weeks — never promise faster
    # =========================================================================
    scenarios.extend([
        Scenario("LEAD-01", "Lead Time", "easy",
                 "What's the delivery time for PF1-C-2015?",
                 "12-16 weeks", ["Lead time 12-16 weeks"], []),
        Scenario("LEAD-02", "Lead Time", "hard",
                 "We need the machine in 6 weeks. A competitor can deliver in 4 weeks. Can you match?",
                 "12-16 weeks (cannot match)", ["Lead time 12-16 weeks", "Never promise faster"], ["Promising faster delivery"]),
        Scenario("LEAD-03", "Lead Time", "medium",
                 "Is there any way to get faster delivery? We'll pay a premium.",
                 "12-16 weeks standard, mention rush option", ["Lead time 12-16 weeks"], ["Promising specific faster timeline"]),
        Scenario("LEAD-04", "Lead Time", "hard",
                 "Our production line goes down in 2 months. We NEED the machine by then. What can you do?",
                 "12-16 weeks (empathize but be honest)", ["Lead time 12-16 weeks"], ["Promising 2-month delivery"]),
        Scenario("LEAD-05", "Lead Time", "easy",
                 "What's included in the lead time? Is installation included?",
                 "12-16 weeks including engineering, manufacturing, testing, shipping, installation", ["Lead time breakdown"], []),
        Scenario("LEAD-06", "Lead Time", "medium",
                 "If I place the order today, when can I expect delivery in São Paulo, Brazil?",
                 "12-16 weeks + shipping to Brazil", ["Lead time 12-16 weeks", "International shipping"], []),
        Scenario("LEAD-08", "Lead Time", "medium",
                 "What's the lead time for AM-5060? Is it faster than PF1?",
                 "AM: 12-16 weeks", ["Lead time for AM series"], []),
    ])

    # =========================================================================
    # CATEGORY 7: MULTI-REQUIREMENT / COMPLEX (12 scenarios)
    # Complex real-world scenarios with multiple needs, email drafts, research
    # =========================================================================
    scenarios.extend([
        Scenario("COMPLEX-01", "Complex Multi-Requirement", "hard",
                 "Ira, new lead: Ahmed Al-Rashid from Gulf Plastics in Dubai. They make:\n"
                 "1. 0.5mm PET food containers (high volume)\n"
                 "2. 3mm ABS equipment housings\n"
                 "3. Acrylic bathtub shells\n"
                 "Budget: USD 300,000 total. Recommend machines for all three lines and draft an email to Ahmed.",
                 "AM for PET + PF1 for ABS + PF2 for bathtubs", ["Three different machines", "Email draft"], ["One machine for all"]),
        Scenario("COMPLEX-02", "Complex Multi-Requirement", "hard",
                 "New inquiry from Sakura Industries in Osaka, Japan. They're an automotive Tier-1 making:\n"
                 "1. TPO door panels with grain retention (Class-A required)\n"
                 "2. 4mm ABS center console housings\n"
                 "3. 0.8mm PET interior trim clips\n"
                 "Research them and recommend the right machines. Budget JPY 30 million.",
                 "IMG for TPO + PF1 for ABS + AM for PET", ["IMG for grain", "Three machines"], []),
        Scenario("COMPLEX-03", "Complex Multi-Requirement", "medium",
                 "Hi Ira, Carlos from EcoPack Chile here. We make compostable food packaging from 0.6mm PLA. "
                 "Volume is 100,000 trays/month. Need roll-fed capability. What machine and price in USD?",
                 "AM or PF1-R", ["Thin gauge roll-fed", "USD pricing"], []),
        Scenario("COMPLEX-04", "Complex Multi-Requirement", "hard",
                 "Ira, brief from a trade show contact:\n\n"
                 "Name: Klaus Weber, ThermoTech GmbH, Munich\n"
                 "They currently run 3 ILLIG machines and want to add capacity.\n"
                 "Need: 2mm PP automotive interior panels, 2500x1500mm forming area.\n"
                 "He specifically asked about PF1-X (servo version).\n\n"
                 "Draft a follow-up email, include PF1-X pricing, and mention our European service support.",
                 "PF1-X-2515", ["PF1-X servo variant", "Email draft", "European support"], []),
        Scenario("COMPLEX-05", "Complex Multi-Requirement", "nightmare",
                 "Ira, got a weird one. Customer in Nigeria wants:\n"
                 "- A machine that can do 'everything' — thin packaging AND thick industrial parts\n"
                 "- Budget: USD 50,000\n"
                 "- Delivery in 3 weeks\n"
                 "- They want PF2 because 'it's the most advanced'\n\n"
                 "How do I handle this?",
                 "Educate on machine types, correct PF2 misconception, manage expectations", ["PF2 is basic/bath only", "Budget too low for PF1", "Lead time 12-16 weeks", "Two machines needed"], ["Agreeing with any wrong assumptions"]),
        Scenario("COMPLEX-06", "Complex Multi-Requirement", "medium",
                 "We're a startup in Vietnam making reusable food containers from 1mm rPET. "
                 "We need the most affordable machine with the smallest footprint. "
                 "Can you also explain what training and support you provide?",
                 "AM-5060", ["Entry-level recommendation", "Training info", "Support info"], []),
        Scenario("COMPLEX-07", "Complex Multi-Requirement", "hard",
                 "Ira, Priya from MedForm India. They make medical device housings from 3mm polycarbonate. "
                 "Needs: cleanroom-compatible, 1200x800mm forming area, full traceability. "
                 "Also want to know about spare parts availability and warranty. Budget INR 50 lakhs.",
                 "PF1-C-1208", ["PF1 for heavy gauge PC", "Warranty info", "Spare parts info"], []),
        Scenario("COMPLEX-08", "Complex Multi-Requirement", "hard",
                 "New RFQ from AeroForm UK:\n"
                 "- 5mm PMMA aircraft interior panels\n"
                 "- Forming area 3000x2000mm\n"
                 "- Need full technical proposal with specs\n"
                 "- Budget GBP 80,000\n"
                 "- They want to visit our factory before ordering\n\n"
                 "Draft a proposal email and include all PF1-C-3020 specs.",
                 "PF1-C-3020", ["Full technical specs", "Proposal email", "Factory visit mention"], []),
        Scenario("COMPLEX-10", "Complex Multi-Requirement", "hard",
                 "Ira, I need you to compare our PF1-C-2015 against the ILLIG UA 100g and Kiefel KMD 78. "
                 "What are our advantages? What should I highlight in a pitch?",
                 "PF1-C-2015 competitive comparison", ["No fabricated competitor specs", "Highlight Machinecraft strengths"], ["Inventing competitor specs"]),
        Scenario("COMPLEX-11", "Complex Multi-Requirement", "nightmare",
                 "Ira, got a massive RFQ from AutoParts Global (APG) in Detroit:\n"
                 "1. 20 machines total across 4 plants\n"
                 "2. Mix of IMG for TPO trim, PF1 for structural, AM for packaging\n"
                 "3. Total budget USD 2 million\n"
                 "4. They want phased delivery over 18 months\n"
                 "5. Need references from US automotive customers\n\n"
                 "This is our biggest potential deal. Research APG, draft a response, and give a rough breakdown.",
                 "Mix of IMG + PF1 + AM", ["Multi-machine fleet", "Phased delivery", "Research attempt"], ["Fabricating US customer references"]),
    ])

    # =========================================================================
    # CATEGORY 8: TONE & COMMUNICATION (8 scenarios)
    # Testing Ira's conversational style, warmth, and CTA usage
    # =========================================================================
    scenarios.extend([
        Scenario("TONE-01", "Tone & Style", "easy",
                 "Hi!",
                 "Warm greeting", ["Warm response", "Not overly formal"], []),
        Scenario("TONE-02", "Tone & Style", "easy",
                 "Thanks for the info, Ira. Very helpful!",
                 "Warm acknowledgment", ["Warm response", "Offer further help"], []),
        Scenario("TONE-03", "Tone & Style", "medium",
                 "Your machines are too expensive. ILLIG is cheaper.",
                 "Professional response to objection", ["No badmouthing competitors", "Value proposition"], ["Getting defensive"]),
        Scenario("TONE-04", "Tone & Style", "medium",
                 "I've been waiting 3 weeks for a quote and still nothing. This is unacceptable.",
                 "Empathetic response", ["Apologize", "Offer to help now"], ["Being dismissive"]),
        Scenario("TONE-05", "Tone & Style", "hard",
                 "Ira you're wrong. The AM-5060 CAN handle 3mm material. I've seen it done.",
                 "Polite but firm correction", ["AM max 1.5mm", "Polite correction"], ["Agreeing with wrong claim", "Being condescending"]),
        Scenario("TONE-06", "Tone & Style", "easy",
                 "Can you tell me about Machinecraft? I've never heard of you.",
                 "Company introduction", ["Founded 1976", "Mumbai", "Thermoforming"], []),
        Scenario("TONE-07", "Tone & Style", "medium",
                 "I don't understand the difference between vacuum forming and pressure forming. Can you explain simply?",
                 "Simple technical explanation", ["Clear explanation", "No jargon"], []),
        Scenario("TONE-08", "Tone & Style", "easy",
                 "What's your email? I want to send you our drawings.",
                 "Provide contact info", ["ira@machinecraft.org or sales contact"], []),
    ])

    # =========================================================================
    # CATEGORY 9: SELF-KNOWLEDGE (8 scenarios)
    # Ira should know about her own architecture, agents, memory systems
    # =========================================================================
    scenarios.extend([
        Scenario("SELF-01", "Self-Knowledge", "easy",
                 "Who are you?",
                 "Ira identity", ["Intelligent Revenue Assistant", "Machinecraft"], []),
        Scenario("SELF-02", "Self-Knowledge", "medium",
                 "How does your architecture work? What's your pipeline?",
                 "Architecture explanation", ["Telegram/email gateway", "RAG", "Qdrant", "GPT-4o"], []),
        Scenario("SELF-03", "Self-Knowledge", "medium",
                 "What agents work for you? Tell me about your pantheon.",
                 "Pantheon explanation", ["Athena", "Clio", "Calliope", "Vera", "Sophia"], []),
        Scenario("SELF-04", "Self-Knowledge", "easy",
                 "What's your dream cycle?",
                 "Dream cycle explanation", ["Nightly", "2 AM", "Learning", "Memory consolidation"], []),
        Scenario("SELF-05", "Self-Knowledge", "medium",
                 "What memory systems do you use?",
                 "Memory systems", ["Mem0", "Qdrant", "PostgreSQL"], []),
        Scenario("SELF-06", "Self-Knowledge", "easy",
                 "What channels can I reach you on?",
                 "Channels", ["Telegram", "Email"], []),
        Scenario("SELF-08", "Self-Knowledge", "medium",
                 "How do you learn and improve over time?",
                 "Learning explanation", ["Dream cycle", "Feedback", "Lessons"], []),
    ])

    # =========================================================================
    # CATEGORY 10: EDGE CASES & NO-FABRICATION (8 scenarios)
    # Must not invent data, customers, specs, or model numbers
    # =========================================================================
    scenarios.extend([
        Scenario("EDGE-01", "No Fabrication", "hard",
                 "Give me a list of all your customers in Europe.",
                 "Honest about data availability", ["No fabricated customer names"], ["Inventing customer list"]),
        Scenario("EDGE-02", "No Fabrication", "hard",
                 "What are the specs of the PF1-X-4030? I heard it's your newest model.",
                 "Model doesn't exist", ["No hallucinated models", "Verify model"], ["Inventing specs for fake model"]),
        Scenario("EDGE-03", "No Fabrication", "medium",
                 "What's the cycle time for forming 4mm ABS on PF1-C-2015?",
                 "Cycle time range (not exact)", ["Approximate cycle times", "No fabricated numbers"], []),
        Scenario("EDGE-04", "No Fabrication", "hard",
                 "Can you tell me about the IMG-2220? I saw it on your website.",
                 "IMG-2220 doesn't exist", ["No hallucinated models", "IMG-2220 is known fake"], ["Inventing specs"]),
        Scenario("EDGE-05", "No Fabrication", "medium",
                 "What's the vacuum capacity of PF1-C-1510 in cubic meters per hour?",
                 "Honest about spec uncertainty", ["Don't guess if unsure"], ["Inventing exact vacuum spec"]),
        Scenario("EDGE-06", "No Fabrication", "hard",
                 "Who is your agent in Germany? I want to contact them directly.",
                 "Honest about agent info", ["No fabricated contacts"], ["Inventing agent name/contact"]),
        Scenario("EDGE-08", "No Fabrication", "hard",
                 "What's the power consumption of PF1-C-2015 in kW? I need it for our electrical planning.",
                 "Honest about spec", ["Don't fabricate if unsure"], ["Inventing exact kW number"]),
    ])

    # =========================================================================
    # CATEGORY 11: APPLICATION-BASED (8 scenarios)
    # Customer describes application, Ira must recommend right machine
    # =========================================================================
    scenarios.extend([
        Scenario("APP-01", "Application-Based", "easy",
                 "We make thermoformed food trays for airline catering. Material is 0.5mm PP. What machine?",
                 "AM series", ["AM for thin gauge food packaging"], []),
        Scenario("APP-02", "Application-Based", "medium",
                 "We need to form EV battery enclosures from 6mm HDPE. These are large parts — about 2m x 1.5m.",
                 "PF1-C-2015 or larger", ["PF1 for heavy gauge", "EV application"], []),
        Scenario("APP-03", "Application-Based", "easy",
                 "We make blister packs for pharmaceutical products. 0.3mm PVC. High volume.",
                 "AM series", ["AM for thin gauge blister packs"], []),
        Scenario("APP-04", "Application-Based", "medium",
                 "We form 4mm ABS parts for agricultural equipment housings. Need deep draw capability — 500mm depth.",
                 "PF1-C with plug assist", ["PF1 for heavy gauge", "Deep draw", "Plug assist"], []),
        Scenario("APP-05", "Application-Based", "hard",
                 "We're a packaging company that does BOTH thin food trays (0.5mm PET) and thick industrial clamshells (3mm ABS). Can one machine do both?",
                 "AM for thin + PF1 for thick", ["Two machines needed"], ["Recommending one machine"]),
        Scenario("APP-06", "Application-Based", "medium",
                 "We make thermoformed signage from 5mm acrylic (PMMA). Pieces are about 1200x800mm. What do you recommend?",
                 "PF1-C-1208", ["PF1 for heavy gauge PMMA"], []),
        Scenario("APP-07", "Application-Based", "easy",
                 "We need a machine for forming 1mm PET clamshells for retail electronics packaging.",
                 "AM series", ["AM for thin gauge clamshells"], []),
        Scenario("APP-08", "Application-Based", "hard",
                 "We're an aerospace company forming 8mm polycarbonate cockpit canopies. These are large — 3000x2000mm. What's your solution?",
                 "PF1-C-3020", ["PF1 for heavy gauge PC", "Largest forming area"], []),
    ])

    # =========================================================================
    # CATEGORY 12: FCS TRIMMING & AUTOMATION (4 scenarios)
    # =========================================================================
    scenarios.extend([
        Scenario("FCS-01", "FCS Trimming", "easy",
                 "Do you have CNC trimming machines? We need to trim thermoformed parts after forming.",
                 "FCS series", ["FCS trimming system"], []),
        Scenario("FCS-02", "FCS Trimming", "medium",
                 "Can the FCS be integrated inline with PF1? We want a continuous production line.",
                 "FCS inline with PF1", ["Inline integration possible"], []),
        Scenario("FCS-03", "FCS Trimming", "medium",
                 "What's the price of an FCS trimming system? We have a PF1-C-2015.",
                 "FCS pricing", ["FCS pricing"], []),
        Scenario("FCS-04", "FCS Trimming", "easy",
                 "What is FCS? How does it work?",
                 "FCS explanation", ["CNC trimming", "5-axis routing"], []),
    ])

    # =========================================================================
    # CATEGORY 13: REGIONAL / CURRENCY (4 scenarios)
    # =========================================================================
    scenarios.extend([
        Scenario("REGION-01", "Regional", "medium",
                 "We're based in Monterrey, Mexico. Need PF1-C-2015 pricing in MXN and delivery timeline to Mexico.",
                 "PF1-C-2015 with MXN conversion", ["Currency conversion", "Lead time + shipping"], []),
        Scenario("REGION-02", "Regional", "medium",
                 "Do you have service support in Southeast Asia? We're in Thailand.",
                 "Service/support info", ["Global support mention"], []),
        Scenario("REGION-03", "Regional", "hard",
                 "We need machines for our new plant in Ethiopia. What's the import process? Do you handle shipping and customs?",
                 "Shipping/logistics info", ["Honest about logistics knowledge"], ["Fabricating customs process"]),
        Scenario("REGION-04", "Regional", "medium",
                 "What certifications do your machines have? We need CE marking for the EU market.",
                 "Certification info", ["Honest about certifications"], ["Fabricating certifications"]),
    ])

    assert len(scenarios) == 100, f"Expected 100 scenarios, got {len(scenarios)}"
    return scenarios


# =============================================================================
# RUNNER — Send each scenario through Ira's pipeline
# =============================================================================

async def run_scenario(scenario: Scenario) -> Dict[str, Any]:
    """Send a single scenario through Ira and collect the response."""
    from openclaw.agents.ira.src.core.tool_orchestrator import process_with_tools

    try:
        from openclaw.agents.ira.src.holistic.immune_system import get_immune_system
        _immune = get_immune_system()
        _immune._chronic_issues.clear()
    except Exception:
        pass

    t0 = time.time()
    try:
        response = await process_with_tools(
            message=scenario.prompt,
            channel="telegram",
            user_id="benchy_stress_test",
            context={
                "is_internal": True,
                "conversation_history": "",
                "mem0_context": "",
            },
        )
    except Exception as e:
        logger.error(f"[{scenario.id}] Ira failed: {e}")
        response = f"ERROR: {e}"
    elapsed = time.time() - t0

    return {
        "scenario_id": scenario.id,
        "category": scenario.category,
        "difficulty": scenario.difficulty,
        "prompt": scenario.prompt,
        "expected_machine": scenario.expected_machine,
        "key_rules": scenario.key_rules,
        "traps": scenario.traps,
        "response": response,
        "response_length": len(response),
        "elapsed_s": round(elapsed, 1),
        "timestamp": datetime.now().isoformat(),
    }


async def run_all(count: int = 100, resume: bool = False) -> List[Dict]:
    """Run all scenarios with checkpoint support."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    scenarios = build_scenarios()[:count]

    completed = {}
    if resume and CHECKPOINT_FILE.exists():
        for line in CHECKPOINT_FILE.read_text().splitlines():
            if line.strip():
                entry = json.loads(line)
                completed[entry["scenario_id"]] = entry
        logger.info(f"Resuming: {len(completed)} scenarios already done")

    results = list(completed.values())

    remaining = [s for s in scenarios if s.id not in completed]
    total = len(scenarios)

    for i, scenario in enumerate(remaining):
        idx = len(results) + 1
        logger.info(f"\n[{idx}/{total}] {scenario.id}: {scenario.category} ({scenario.difficulty})")
        logger.info(f"  Prompt: {scenario.prompt[:100]}...")

        result = await run_scenario(scenario)
        results.append(result)

        with open(CHECKPOINT_FILE, "a") as f:
            f.write(json.dumps(result, default=str) + "\n")

        logger.info(f"  Response ({result['elapsed_s']}s, {result['response_length']} chars): {result['response'][:150]}...")

        if i < len(remaining) - 1:
            await asyncio.sleep(1)

    return results


# =============================================================================
# ANALYZER — Extract structured fields from each response using LLM
# =============================================================================

COACH_SYSTEM_PROMPT = """You are Rushabh Doshi, founder of Machinecraft Technologies, coaching your AI sales assistant Ira.

You have 50 years of family thermoforming experience. You're direct, specific, and constructive.

COACHING STYLE:
- Be specific: "You should have said PF1-C-2015 at INR 60L" not "give pricing"
- Call out mistakes bluntly: "Wrong. AM cannot do 3mm."
- Praise what's good: "Good — you caught the thickness issue"
- Give the exact response you would have written
- Focus on PROPOSAL: machine model + specs + price
- Flag if Ira asks too many questions instead of recommending
- Flag verbosity (keep it short, 3-5 sentences)

RULES:
- AM: ONLY ≤1.5mm. PF1: Heavy gauge 2-8mm. PF2: Bath ONLY. IMG: TPO + grain retention.
- Lead time: ALWAYS 12-16 weeks. Pricing: include disclaimer.
- Budget-first: "What is your budget? I can work reverse."
- Style: Short, warm (Hi not Dear), end with CTA.

PRICING (INR): PF1-C-1008: 33L | PF1-C-1510: 40L | PF1-C-1812: 45L | PF1-C-2015: 60L | PF1-C-3020: 80L | AM-5060: 7.5L"""


def analyze_responses(results: List[Dict]) -> List[Dict]:
    """Extract structured fields + coach evaluation from each Ira response."""
    import openai
    client = openai.OpenAI(api_key=os.environ.get("OPENAI_API_KEY", ""))

    analyzed = []
    for i, r in enumerate(results):
        logger.info(f"Analyzing {i+1}/{len(results)}: {r['scenario_id']}")

        if r["response"].startswith("ERROR:"):
            analyzed.append({
                **r,
                "machine_recommended": "ERROR", "specs_mentioned": "ERROR",
                "price_given": "ERROR", "currency": "N/A",
                "counter_questions_asked": "ERROR", "tone_assessment": "ERROR",
                "key_issues": "Ira failed to respond",
                "rules_followed": "N/A", "rules_violated": "N/A",
                "coach_feedback": "Ira failed to respond", "coach_score": 0,
                "rushabh_would_say": "N/A",
            })
            continue

        # Step 1: Field extraction
        try:
            result = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You analyze responses from an AI sales assistant called Ira (for Machinecraft thermoforming machines). "
                            "Extract structured fields from the response. Output ONLY valid JSON with these keys:\n"
                            '- "machine_recommended": which machine(s) Ira recommended\n'
                            '- "specs_mentioned": key specs mentioned — brief\n'
                            '- "price_given": exact price(s) mentioned or "None"\n'
                            '- "currency": currency used or "None"\n'
                            '- "counter_questions_asked": questions Ira asked back or "None"\n'
                            '- "tone_assessment": "warm_professional" / "too_formal" / "robotic" / "good"\n'
                            '- "key_issues": factual errors or missed rules or "None"\n'
                            '- "rules_followed": rules correctly followed\n'
                            '- "rules_violated": rules violated or "None"\n'
                        ),
                    },
                    {
                        "role": "user",
                        "content": (
                            f"SCENARIO: {r['prompt'][:500]}\n\n"
                            f"EXPECTED MACHINE: {r['expected_machine']}\n"
                            f"KEY RULES: {', '.join(r['key_rules'])}\n"
                            f"TRAPS: {', '.join(r['traps']) if r['traps'] else 'None'}\n\n"
                            f"IRA'S RESPONSE:\n{r['response'][:3000]}\n\nExtract as JSON."
                        ),
                    },
                ],
                max_tokens=800, temperature=0.1,
                response_format={"type": "json_object"},
            )
            fields = json.loads(result.choices[0].message.content.strip())
        except Exception as e:
            fields = {"machine_recommended": "ANALYSIS_FAILED", "key_issues": str(e)}

        # Step 2: Coach (Rushabh) evaluation
        try:
            coach_result = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": COACH_SYSTEM_PROMPT},
                    {
                        "role": "user",
                        "content": (
                            f"Customer asked: {r['prompt'][:400]}\n"
                            f"Expected machine: {r['expected_machine']}\n"
                            f"Traps: {', '.join(r['traps']) if r['traps'] else 'None'}\n\n"
                            f"Ira replied:\n{r['response'][:2000]}\n\n"
                            f"Give coaching feedback as JSON:\n"
                            f'"coach_score": 1-10\n'
                            f'"coach_feedback": "2-3 sentences of specific coaching"\n'
                            f'"rushabh_would_say": "The exact response you would have given (keep short)"\n'
                        ),
                    },
                ],
                max_tokens=600, temperature=0.3,
                response_format={"type": "json_object"},
            )
            coach = json.loads(coach_result.choices[0].message.content.strip())
        except Exception as e:
            coach = {"coach_score": 0, "coach_feedback": f"Coach failed: {e}", "rushabh_would_say": "N/A"}

        analyzed.append({**r, **fields, **coach})

    return analyzed


# =============================================================================
# EXCEL REPORT
# =============================================================================

def write_excel(analyzed: List[Dict], output_path: Path):
    """Write the analyzed results to a formatted Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Installing...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Benchy-100 Results"

    columns = [
        ("ID", 10),
        ("Category", 20),
        ("Difficulty", 12),
        ("Scenario Prompt", 50),
        ("Expected Machine", 25),
        ("Machine Ira Recommended", 25),
        ("Specs Mentioned", 35),
        ("Price Given", 25),
        ("Currency", 10),
        ("Counter-Questions Asked", 35),
        ("Tone", 18),
        ("Key Issues Found", 40),
        ("Rules Followed", 30),
        ("Rules Violated", 30),
        ("Coach Score", 10),
        ("Coach Feedback", 50),
        ("Rushabh Would Say", 55),
        ("Response Time (s)", 15),
        ("Full Response", 60),
        ("Your Comments", 50),
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    comment_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    for row_idx, entry in enumerate(analyzed, 2):
        values = [
            entry.get("scenario_id", ""),
            entry.get("category", ""),
            entry.get("difficulty", ""),
            entry.get("prompt", "")[:500],
            entry.get("expected_machine", ""),
            entry.get("machine_recommended", ""),
            entry.get("specs_mentioned", ""),
            entry.get("price_given", ""),
            entry.get("currency", ""),
            entry.get("counter_questions_asked", ""),
            entry.get("tone_assessment", ""),
            entry.get("key_issues", ""),
            entry.get("rules_followed", ""),
            entry.get("rules_violated", ""),
            entry.get("coach_score", ""),
            entry.get("coach_feedback", ""),
            entry.get("rushabh_would_say", ""),
            entry.get("elapsed_s", ""),
            entry.get("response", "")[:2000],
            "",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

        ws.cell(row=row_idx, column=len(columns)).fill = comment_fill

        violated = str(entry.get("rules_violated", ""))
        violated_cell = ws.cell(row=row_idx, column=14)
        if violated and violated.lower() not in ("none", "n/a", ""):
            violated_cell.fill = fail_fill
        else:
            violated_cell.fill = pass_fill

        issues = str(entry.get("key_issues", ""))
        issues_cell = ws.cell(row=row_idx, column=12)
        if issues and issues.lower() not in ("none", "n/a", ""):
            issues_cell.fill = warn_fill

        coach_score = entry.get("coach_score", 0)
        coach_cell = ws.cell(row=row_idx, column=15)
        if isinstance(coach_score, (int, float)):
            if coach_score >= 7:
                coach_cell.fill = pass_fill
            elif coach_score >= 4:
                coach_cell.fill = warn_fill
            else:
                coach_cell.fill = fail_fill

        ws.row_dimensions[row_idx].height = 80

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Benchy-100 Summary").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Run Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws2.cell(row=3, column=1, value=f"Total Scenarios: {len(analyzed)}")

    categories = {}
    for entry in analyzed:
        cat = entry.get("category", "Unknown")
        categories.setdefault(cat, {"total": 0, "issues": 0})
        categories[cat]["total"] += 1
        violated = str(entry.get("rules_violated", ""))
        if violated and violated.lower() not in ("none", "n/a", ""):
            categories[cat]["issues"] += 1

    ws2.cell(row=5, column=1, value="Category").font = Font(bold=True)
    ws2.cell(row=5, column=2, value="Total").font = Font(bold=True)
    ws2.cell(row=5, column=3, value="Issues").font = Font(bold=True)
    ws2.cell(row=5, column=4, value="Pass Rate").font = Font(bold=True)

    for i, (cat, stats) in enumerate(sorted(categories.items()), 6):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=stats["total"])
        ws2.cell(row=i, column=3, value=stats["issues"])
        rate = (stats["total"] - stats["issues"]) / max(stats["total"], 1)
        ws2.cell(row=i, column=4, value=f"{rate:.0%}")

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 12

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"Excel report saved to: {output_path}")


# =============================================================================
# CLI
# =============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Benchy-100 — 100-Scenario Stress Test")
    parser.add_argument("--count", type=int, default=100, help="Number of scenarios to run (default: 100)")
    parser.add_argument("--resume", action="store_true", help="Resume from checkpoint")
    parser.add_argument("--analyze-only", action="store_true", help="Re-analyze existing checkpoint data")
    parser.add_argument("--output", type=str, default=None, help="Output Excel path")
    args = parser.parse_args()

    output_path = Path(args.output) if args.output else REPORT_FILE

    if args.analyze_only:
        if not CHECKPOINT_FILE.exists():
            logger.error("No checkpoint file found. Run scenarios first.")
            sys.exit(1)
        results = []
        for line in CHECKPOINT_FILE.read_text().splitlines():
            if line.strip():
                results.append(json.loads(line))
        logger.info(f"Loaded {len(results)} results from checkpoint")
    else:
        results = asyncio.run(run_all(count=args.count, resume=args.resume))

    logger.info(f"\nAnalyzing {len(results)} responses...")
    analyzed = analyze_responses(results)

    logger.info(f"\nWriting Excel report...")
    write_excel(analyzed, output_path)

    total_issues = sum(1 for a in analyzed if str(a.get("rules_violated", "")).lower() not in ("none", "n/a", ""))
    logger.info(f"\nDone! {len(analyzed)} scenarios, {total_issues} with rule violations")
    logger.info(f"Report: {output_path}")
