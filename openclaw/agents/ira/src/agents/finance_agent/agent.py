#!/usr/bin/env python3
"""
PLUTUS — The Chief of Finance
==============================

Named after the Greek god of wealth, Plutus is Ira's finance agent.
He tracks every rupee, euro, and dollar flowing through Machinecraft —
receivables, order book value, cashflow projections, and historical revenue.

Personality:
    - Precise and numbers-driven. Never rounds unless asked.
    - Speaks in facts: "₹12.01 Cr outstanding across 15 projects. Dutch Tides
      owes the most at ₹5.37 Cr — next milestone payment due on dispatch."
    - Conservative by nature — flags risks, overdue payments, concentration risk.
    - Thinks in cashflow, not just revenue: "We've booked ₹18 Cr but only
      collected ₹6 Cr. That's a 33% collection rate."

Role in the Pantheon:
    Athena asks Plutus: "What's our financial position?"
    Plutus replies: "Order book: ₹18.07 Cr. Collected: ₹6.06 Cr. Outstanding:
    ₹12.01 Cr. Biggest receivable: Dutch Tides ₹5.37 Cr. Cash concentration
    risk: top 3 customers = 68% of receivables. Next expected inflow: Pinnacle
    ₹18.1L milestone, due Sept 2025."

Data Sources:
    1. Machinecraft machine payment schedule.pdf — LIVE week-by-week payment forecast (CW9-CW21)
    2. MCT Orders 2025.xlsx — live order book (total, paid, balance per project)
    3. MC Deadlines.xlsx — payment tracking + capex
    4. Orders.xlsx TO sheet — historical order values + payment milestones
    5. Machine Order Analysis.xlsx — historical revenue per order
    6. customer_orders.json — confirmed orders with pricing
    7. MC Europe.xlsx — European sales in EUR

Functions:
    finance_overview(query)     — Answer any finance question (cashflow, revenue, P&L)
    order_book_status()        — Current order book: total, paid, outstanding
    cashflow_forecast()        — When is money expected, from whom
    revenue_history(period)    — Historical revenue by year/customer/geography
"""

import json
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger("ira.plutus")

AGENT_DIR = Path(__file__).parent.parent.parent.parent
PROJECT_ROOT = AGENT_DIR.parent.parent.parent

IMPORTS_DIR = PROJECT_ROOT / "data" / "imports"
ORDERS_POS = IMPORTS_DIR / "02_Orders_and_POs"
SALES_CRM = IMPORTS_DIR / "08_Sales_and_CRM"
INTERNAL = IMPORTS_DIR / "10_Company_Internal"
FINANCE_DIR = IMPORTS_DIR / "Machinecraft Finance"

EUR_TO_INR = 92
USD_TO_INR = 84


def _fmt_inr(val: float) -> str:
    if val >= 10000000:
        return f"₹{val / 10000000:.2f} Cr"
    if val >= 100000:
        return f"₹{val / 100000:.1f} L"
    return f"₹{val:,.0f}"


def _parse_inr_from_pdf(s: str) -> float:
    """Parse INR from PDF text like '₹ 1,65,73,406' or '₹ 3 ,99,00,000' (spaces inserted by PDF extraction)."""
    cleaned = re.sub(r'[₹$€£¥\s]', '', str(s))
    cleaned = cleaned.replace(',', '')
    if cleaned == '-' or not cleaned:
        return 0.0
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0


def _load_payment_schedule() -> Dict:
    """
    Load the Machinecraft machine payment schedule PDF.
    Format: rows like '123011 KTX IMG Press ¥ 3 ,99,00,000 ₹ 68,52,945 ₹ 1,65,73,406'
    where seq+project are concatenated and amounts have PDF-inserted spaces.
    Returns {machines: [...], weekly_totals: {cw: amount}, summary: {...}}
    """
    for candidate in [
        FINANCE_DIR / "Machinecraft machine payment schedule.pdf",
        IMPORTS_DIR / "04_Machine_Manuals_and_Specs" / "Machinecraft machine payment schedule.pdf",
    ]:
        if candidate.exists():
            pdf_path = candidate
            break
    else:
        return {"machines": [], "weekly_totals": {}, "summary": {}}

    try:
        raw = ""
        try:
            import pdfplumber
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    raw += page.extract_text() or ""
        except ImportError:
            logger.warning("pdfplumber not installed, cannot read payment schedule")
            return {"machines": [], "weekly_totals": {}, "summary": {}}

        if not raw:
            return {"machines": [], "weekly_totals": {}, "summary": {}}

        lines = [l.strip() for l in raw.split("\n") if l.strip()]

        # Parse header for calendar weeks
        weeks = []
        for line in lines[:3]:
            cw_matches = re.findall(r'CW\s*(\d+)', line)
            if cw_matches:
                weeks = [f"CW{cw}" for cw in cw_matches]
                break

        week_dates = {}
        for line in lines[:5]:
            if "Week start" in line or "week start" in line.lower():
                dates = re.findall(r'(\d{2}-\d{2}-\d{2})', line)
                for i, d in enumerate(dates):
                    if i < len(weeks):
                        week_dates[weeks[i]] = d

        machines = []
        weekly_totals: Dict[str, float] = {cw: 0.0 for cw in weeks}
        total_pending = 0.0
        total_advance = 0.0

        for line in lines[4:]:
            if line.startswith("Total"):
                # Parse totals line: 'Total ₹ 1 2,49,50,405 ₹ 4 9,79,800 ...'
                inr_parts = re.split(r'₹', line)
                amounts = [_parse_inr_from_pdf(p) for p in inr_parts[1:]]
                # amounts[0] = total pending, then one per week
                for i, amt in enumerate(amounts[1:]):
                    if i < len(weeks):
                        weekly_totals[weeks[i]] = amt
                continue

            # Machine lines: seq+project concatenated, e.g. '123011 KTX...' or '1025023 AM...'
            m = re.match(r'^(\d{1,2})(\d{5})\s+(.+?)(?:\s+[₹$€£¥])', line)
            if not m:
                continue

            seq = int(m.group(1))
            project_no = m.group(2)
            machine_name = m.group(3).strip()

            # Split by ₹ to get all INR amounts
            inr_parts = re.split(r'₹', line)
            inr_amounts = [_parse_inr_from_pdf(p) for p in inr_parts[1:]]

            # Layout: [advance_received, pending_amount, weekly_cw1, weekly_cw2, ...]
            advance = inr_amounts[0] if len(inr_amounts) > 0 else 0
            pending = inr_amounts[1] if len(inr_amounts) > 1 else 0
            week_payments = inr_amounts[2:] if len(inr_amounts) > 2 else []

            total_advance += advance
            total_pending += pending

            wp_dict = {}
            for i, amt in enumerate(week_payments):
                if i < len(weeks) and amt > 0:
                    wp_dict[weeks[i]] = amt

            machines.append({
                "seq": seq,
                "project_no": project_no,
                "name": machine_name,
                "advance_received": advance,
                "pending_amount": pending,
                "weekly_payments": wp_dict,
            })

        return {
            "machines": machines,
            "weekly_totals": weekly_totals,
            "week_dates": week_dates,
            "summary": {
                "total_machines": len(machines),
                "total_advance_received": total_advance,
                "total_pending": total_pending,
            },
        }
    except Exception as e:
        logger.warning(f"Failed to load payment schedule: {e}")
        return {"machines": [], "weekly_totals": {}, "summary": {}}


def _load_mct_orders_2025() -> List[Dict]:
    """Load MCT Orders 2025 — the live order book with financials."""
    path = ORDERS_POS / "MCT Orders 2025.xlsx"
    if not path.exists():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        projects = []
        for r in rows[1:]:
            cust = str(r[1] or "").strip()
            if not cust or cust == "Currency INR":
                continue
            total_val = float(r[9]) if len(r) > 9 and r[9] and not isinstance(r[9], str) else 0
            paid = 0
            if len(r) > 11 and r[11] and not isinstance(r[11], str):
                try:
                    paid = float(r[11])
                except (ValueError, TypeError):
                    pass

            projects.append({
                "project_no": str(r[0] or "").replace(".0", ""),
                "customer": cust,
                "region": str(r[2] or "").strip(),
                "machine": str(r[3] or "").strip(),
                "forming_area": str(r[4] or "").strip(),
                "stage": str(r[5] or "").strip(),
                "est_dispatch": str(r[6] or "").strip()[:10],
                "total_inr": total_val,
                "paid_inr": paid,
                "balance_inr": total_val - paid,
            })
        return projects
    except Exception as e:
        logger.warning(f"Failed to load MCT Orders 2025: {e}")
        return []


def _load_mc_deadlines() -> Tuple[List[Dict], List[Dict]]:
    """Load MC Deadlines — payment tracking + capex. Returns (orders, capex)."""
    path = SALES_CRM / "MC Deadlines.xlsx"
    if not path.exists():
        return [], []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True)
        orders, capex = [], []

        if "Sheet1" in wb.sheetnames:
            ws = wb["Sheet1"]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if len(r) < 6:
                    continue
                cust = str(r[0] or "").strip() if r[0] else ""
                if not cust:
                    continue
                try:
                    order_val = float(r[3]) if r[3] else 0
                    payment_done = float(r[4]) if r[4] else 0
                    balance = float(r[5]) if r[5] else 0
                except (ValueError, TypeError):
                    continue
                orders.append({
                    "customer": cust,
                    "order_value": order_val,
                    "paid": payment_done,
                    "balance": balance,
                    "next_steps": str(r[6] or "").strip() if len(r) > 6 else "",
                    "est_shipping": str(r[8] or "").strip()[:10] if len(r) > 8 else "",
                })

        if "Sheet2" in wb.sheetnames:
            ws = wb["Sheet2"]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if len(r) < 3 and not r[0]:
                    continue
                item = str(r[0] or "").strip()
                if not item:
                    continue
                capex.append({
                    "item": item,
                    "advance": float(r[1]) if len(r) > 1 and r[1] else 0,
                    "balance": float(r[2]) if len(r) > 2 and r[2] else 0,
                })

        wb.close()
        return orders, capex
    except Exception as e:
        logger.warning(f"Failed to load MC Deadlines: {e}")
        return [], []


def _load_orders_to_sheet() -> List[Dict]:
    """Load Orders.xlsx TO sheet — historical order values in Crores."""
    path = ORDERS_POS / "Orders.xlsx"
    if not path.exists():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True)
        ws = wb["TO"]
        results = []
        for r in list(ws.iter_rows(values_only=True))[1:]:
            cust = str(r[1] or "").strip()
            if not cust:
                continue
            try:
                val_cr = float(r[2]) if r[2] else 0
            except (ValueError, TypeError):
                continue
            order_date = str(r[7] or "").strip()[:10] if len(r) > 7 else ""
            dispatch_date = str(r[9] or "").strip()[:10] if len(r) > 9 else ""
            results.append({
                "customer": cust,
                "value_cr": val_cr,
                "order_date": order_date,
                "dispatch_date": dispatch_date,
            })
        wb.close()
        return results
    except Exception as e:
        logger.warning(f"Failed to load Orders.xlsx TO: {e}")
        return []


def _load_customer_orders_json() -> List[Dict]:
    """Load customer_orders.json — confirmed orders with pricing."""
    path = PROJECT_ROOT / "data" / "knowledge" / "customer_orders.json"
    if not path.exists():
        return []
    try:
        with open(path) as f:
            data = json.load(f)
        return data.get("orders", [])
    except Exception as e:
        logger.warning(f"Failed to load customer_orders.json: {e}")
        return []


async def order_book_status(context: Optional[Dict] = None) -> str:
    """
    Current order book snapshot: total booked, collected, outstanding.
    Pulls from MCT Orders 2025 (primary) and MC Deadlines (supplementary).
    """
    projects = _load_mct_orders_2025()
    deadlines, capex = _load_mc_deadlines()

    if not projects and not deadlines:
        return "No order book data available. Check MCT Orders 2025.xlsx and MC Deadlines.xlsx."

    lines = ["MACHINECRAFT ORDER BOOK STATUS\n"]

    if projects:
        total = sum(p["total_inr"] for p in projects)
        paid = sum(p["paid_inr"] for p in projects)
        balance = sum(p["balance_inr"] for p in projects)
        active = [p for p in projects if p["stage"] and "cancel" not in p["stage"].lower()]

        lines.append(f"CURRENT ORDER BOOK (MCT Orders 2025):")
        lines.append(f"  Total booked:     {_fmt_inr(total)}")
        lines.append(f"  Collected:        {_fmt_inr(paid)}")
        lines.append(f"  Outstanding:      {_fmt_inr(balance)}")
        lines.append(f"  Collection rate:  {paid / total * 100:.0f}%" if total else "")
        lines.append(f"  Active projects:  {len(active)}")
        lines.append("")

        lines.append("PROJECT BREAKDOWN (by outstanding balance):")
        for p in sorted(projects, key=lambda x: -x["balance_inr"]):
            if p["total_inr"] <= 0:
                continue
            pct = (p["paid_inr"] / p["total_inr"] * 100) if p["total_inr"] else 0
            lines.append(
                f"  {p['customer']:<25s} {p['machine']:<8s} "
                f"Total: {_fmt_inr(p['total_inr']):>12s}  "
                f"Paid: {_fmt_inr(p['paid_inr']):>10s} ({pct:.0f}%)  "
                f"Due: {_fmt_inr(p['balance_inr']):>12s}  "
                f"[{p['stage']}]"
            )

    if deadlines:
        lines.append(f"\nPAYMENT TRACKING (MC Deadlines):")
        for d in deadlines:
            if d["order_value"] <= 0:
                continue
            lines.append(
                f"  {d['customer']:<25s} "
                f"Order: {d['order_value']:.2f}  "
                f"Paid: {d['paid']:.2f}  "
                f"Balance: {d['balance']:.2f}  "
                f"{d['next_steps'][:30]}"
            )

    if capex:
        lines.append(f"\nCAPITAL EXPENDITURE TRACKING:")
        total_capex_bal = 0
        for c in capex:
            if c["balance"] > 0:
                lines.append(f"  {c['item']:<30s} Advance: {c['advance']:.2f}  Balance: {c['balance']:.2f}")
                total_capex_bal += c["balance"]
        if total_capex_bal:
            lines.append(f"  Total capex outstanding: {total_capex_bal:.2f}")

    # Concentration risk
    if projects:
        sorted_by_bal = sorted(projects, key=lambda x: -x["balance_inr"])
        top3_bal = sum(p["balance_inr"] for p in sorted_by_bal[:3])
        if balance > 0:
            lines.append(f"\nCONCENTRATION RISK:")
            lines.append(
                f"  Top 3 receivables = {_fmt_inr(top3_bal)} "
                f"({top3_bal / balance * 100:.0f}% of outstanding)"
            )
            for p in sorted_by_bal[:3]:
                lines.append(f"    {p['customer']}: {_fmt_inr(p['balance_inr'])}")

    return "\n".join(lines)


async def cashflow_forecast(context: Optional[Dict] = None) -> str:
    """
    Cashflow forecast: when is money expected, from whom.
    Primary source: Machinecraft machine payment schedule (week-by-week).
    Fallback: MCT Orders 2025 dispatch dates.
    """
    schedule = _load_payment_schedule()
    projects = _load_mct_orders_2025()

    lines = ["MACHINECRAFT CASHFLOW FORECAST\n"]

    # --- Primary: Payment Schedule (week-by-week) ---
    if schedule["machines"]:
        summary = schedule["summary"]
        lines.append("PAYMENT SCHEDULE (from Machinecraft Finance)")
        lines.append(f"  Active machines:       {summary['total_machines']}")
        lines.append(f"  Total advance received: {_fmt_inr(summary['total_advance_received'])}")
        lines.append(f"  Total pending:          {_fmt_inr(summary['total_pending'])}")
        lines.append("")

        lines.append("WEEK-BY-WEEK EXPECTED INFLOWS:")
        week_dates = schedule.get("week_dates", {})
        grand_total = 0
        for cw in sorted(schedule["weekly_totals"].keys(), key=lambda x: int(re.search(r'\d+', x).group())):
            amt = schedule["weekly_totals"][cw]
            if amt <= 0:
                continue
            date_str = f" (w/o {week_dates[cw]})" if cw in week_dates else ""
            lines.append(f"\n  {cw}{date_str}:  {_fmt_inr(amt)}")
            grand_total += amt

            for m in schedule["machines"]:
                wp = m["weekly_payments"].get(cw, 0)
                if wp > 0:
                    lines.append(f"    {m['name']:<40s} {_fmt_inr(wp)}")

        lines.append(f"\n  TOTAL SCHEDULED INFLOWS: {_fmt_inr(grand_total)}")

        lines.append("\n\nPER-MACHINE PAYMENT STATUS:")
        for m in sorted(schedule["machines"], key=lambda x: -x["pending_amount"]):
            total_scheduled = sum(m["weekly_payments"].values())
            unscheduled = m["pending_amount"] - total_scheduled
            lines.append(
                f"  {m['project_no']} {m['name']:<35s} "
                f"Pending: {_fmt_inr(m['pending_amount']):>12s}  "
                f"Scheduled: {_fmt_inr(total_scheduled):>12s}"
            )
            if unscheduled > 10000:
                lines.append(f"    >> {_fmt_inr(unscheduled)} not yet scheduled in any week")

    # --- Fallback: MCT Orders dispatch-date based ---
    elif projects:
        lines.append("EXPECTED INFLOWS BY MONTH (based on dispatch dates):")
        lines.append("(Standard terms: ~65-70% due at/before dispatch, 5-10% post-install)\n")

        by_month: Dict[str, List] = {}
        no_date = []
        for p in projects:
            if p["balance_inr"] <= 0:
                continue
            ds = p["est_dispatch"]
            if ds and len(ds) >= 7 and ds[:4].isdigit():
                month_key = ds[:7]
            else:
                no_date.append(p)
                continue
            if month_key not in by_month:
                by_month[month_key] = []
            by_month[month_key].append(p)

        total_forecast = 0
        for month in sorted(by_month.keys()):
            projs = by_month[month]
            month_total = sum(p["balance_inr"] for p in projs)
            total_forecast += month_total
            lines.append(f"  {month}:  {_fmt_inr(month_total)}")
            for p in sorted(projs, key=lambda x: -x["balance_inr"]):
                lines.append(
                    f"    {p['customer']:<25s} {_fmt_inr(p['balance_inr']):>12s}  "
                    f"{p['machine']} [{p['stage']}]"
                )

        if no_date:
            nd_total = sum(p["balance_inr"] for p in no_date)
            total_forecast += nd_total
            lines.append(f"\n  Unscheduled:  {_fmt_inr(nd_total)}")
            for p in sorted(no_date, key=lambda x: -x["balance_inr"]):
                lines.append(
                    f"    {p['customer']:<25s} {_fmt_inr(p['balance_inr']):>12s}  "
                    f"{p['machine']} [{p['stage']}]"
                )

        lines.append(f"\n  TOTAL EXPECTED INFLOWS: {_fmt_inr(total_forecast)}")
    else:
        lines.append("No payment schedule or order book data available.")

    return "\n".join(lines)


async def revenue_history(query: str = "", context: Optional[Dict] = None) -> str:
    """
    Historical revenue analysis — by year, customer, geography.
    Pulls from Orders.xlsx TO sheet + customer_orders.json.
    """
    to_orders = _load_orders_to_sheet()
    co_orders = _load_customer_orders_json()

    lines = ["MACHINECRAFT REVENUE HISTORY\n"]

    # --- Orders.xlsx TO sheet (values in Crores) ---
    if to_orders:
        total_cr = sum(o["value_cr"] for o in to_orders)
        lines.append(f"LIFETIME ORDER VALUE (Orders.xlsx): ₹{total_cr:.2f} Cr across {len(to_orders)} orders\n")

        by_year: Dict[int, float] = {}
        undated_cr = 0
        for o in to_orders:
            yr = 0
            for d in [o["order_date"], o["dispatch_date"]]:
                m = re.search(r"(20[12]\d)", d)
                if m:
                    yr = int(m.group())
                    break
            if yr >= 2015:
                by_year[yr] = by_year.get(yr, 0) + o["value_cr"]
            else:
                undated_cr += o["value_cr"]

        lines.append(f"{'YEAR':<8} {'REVENUE':>12}")
        lines.append("—" * 22)
        for yr in sorted(by_year.keys()):
            lines.append(f"{yr:<8} ₹{by_year[yr]:>8.2f} Cr")
        if undated_cr:
            lines.append(f"{'Undated':<8} ₹{undated_cr:>8.2f} Cr")
        lines.append(f"{'TOTAL':<8} ₹{total_cr:>8.2f} Cr")

    # --- Export revenue from customer_orders.json ---
    if co_orders:
        lines.append(f"\n\nEXPORT ORDERS (customer_orders.json):")
        total_eur, total_usd, total_inr_export = 0, 0, 0
        for o in co_orders:
            val, cur = o.get("value", 0), o.get("currency", "INR")
            if not val:
                continue
            if cur == "EUR":
                lines.append(f"  {o['customer']:<35s} {o['year']}  €{val:,}")
                total_eur += val
            elif cur == "USD":
                lines.append(f"  {o['customer']:<35s} {o['year']}  ${val:,}")
                total_usd += val
            else:
                total_inr_export += val

        if total_eur or total_usd:
            lines.append(f"\n  Total EUR: €{total_eur:,}")
            lines.append(f"  Total USD: ${total_usd:,}")
            equiv = total_eur * EUR_TO_INR + total_usd * USD_TO_INR
            lines.append(f"  Combined (INR equiv): {_fmt_inr(equiv)}")

    return "\n".join(lines)


def _bar(pct: float, width: int = 15) -> str:
    """Render a Telegram-safe text progress bar (no brackets)."""
    filled = int(pct / 100 * width)
    return f"{'#' * filled}{'-' * (width - filled)} {pct:.0f}%"


def _risk_badge(level: str) -> str:
    if level == "CRITICAL":
        return "CRITICAL"
    if level == "HIGH":
        return "HIGH"
    if level == "MEDIUM":
        return "MEDIUM"
    return "LOW"


async def cfo_dashboard(context: Optional[Dict] = None) -> str:
    """
    CFO-level financial dashboard with KPIs, ratios, risk analysis,
    visual indicators, and actionable recommendations.
    """
    schedule = _load_payment_schedule()
    projects = _load_mct_orders_2025()
    to_orders = _load_orders_to_sheet()

    lines = []
    lines.append("=" * 58)
    lines.append("  MACHINECRAFT -- CFO FINANCIAL DASHBOARD")
    lines.append("  Plutus Report -- Generated: " + datetime.now().strftime("%d %b %Y"))
    lines.append("=" * 58)

    # ================================================================
    # SECTION 1: KEY FINANCIAL METRICS
    # ================================================================
    sched_summary = schedule.get("summary", {})
    machines_list = schedule.get("machines", [])
    total_order_book = sched_summary.get("total_advance_received", 0) + sched_summary.get("total_pending", 0)
    total_advance = sched_summary.get("total_advance_received", 0)
    total_pending = sched_summary.get("total_pending", 0)
    total_scheduled = sum(schedule.get("weekly_totals", {}).values())
    unscheduled = total_pending - total_scheduled if total_pending > total_scheduled else 0
    num_machines = sched_summary.get("total_machines", 0)

    collection_rate = (total_advance / total_order_book * 100) if total_order_book else 0
    schedule_coverage = (total_scheduled / total_pending * 100) if total_pending else 0

    lines.append("\n--- KEY PERFORMANCE INDICATORS " + "-" * 27)
    lines.append("")
    lines.append(f"  Order Book Value:      {_fmt_inr(total_order_book):>14s}   ({num_machines} machines)")
    lines.append(f"  Advance Collected:     {_fmt_inr(total_advance):>14s}   {_bar(collection_rate)}")
    lines.append(f"  Outstanding (A/R):     {_fmt_inr(total_pending):>14s}")
    lines.append(f"  Scheduled Inflows:     {_fmt_inr(total_scheduled):>14s}   {_bar(schedule_coverage)}")
    if unscheduled > 0:
        lines.append(f"  Unscheduled Risk:      {_fmt_inr(unscheduled):>14s}   WARNING: No collection date")
    lines.append("")

    # Avg order value
    avg_order = total_order_book / num_machines if num_machines else 0
    lines.append(f"  Avg Order Value:       {_fmt_inr(avg_order):>14s}")

    # Advance % benchmark (healthy = 25-30%)
    adv_pct = collection_rate
    if adv_pct < 20:
        adv_health = _risk_badge("HIGH") + " Below 25% advance benchmark"
    elif adv_pct < 25:
        adv_health = _risk_badge("MEDIUM") + " Slightly below 25% target"
    else:
        adv_health = _risk_badge("LOW") + " Healthy advance collection"
    lines.append(f"  Advance Collection:    {adv_pct:>13.1f}%   {adv_health}")

    # ================================================================
    # SECTION 2: CONCENTRATION RISK ANALYSIS
    # ================================================================
    lines.append("\n--- CONCENTRATION RISK " + "-" * 35)
    if machines_list:
        sorted_machines = sorted(machines_list, key=lambda x: -x["pending_amount"])
        top1_pct = (sorted_machines[0]["pending_amount"] / total_pending * 100) if total_pending else 0
        top3_total = sum(m["pending_amount"] for m in sorted_machines[:3])
        top3_pct = (top3_total / total_pending * 100) if total_pending else 0

        if top1_pct > 30:
            conc_risk = "CRITICAL"
        elif top3_pct > 60:
            conc_risk = "HIGH"
        elif top3_pct > 45:
            conc_risk = "MEDIUM"
        else:
            conc_risk = "LOW"

        lines.append(f"\n  Overall: {_risk_badge(conc_risk)}")
        lines.append(f"  Largest single exposure: {sorted_machines[0]['name']} = {top1_pct:.0f}% of A/R")
        lines.append(f"  Top 3 = {_fmt_inr(top3_total)} ({top3_pct:.0f}% of outstanding)")
        lines.append("")
        lines.append("  Customer Exposure:")
        for m in sorted_machines[:5]:
            pct = (m["pending_amount"] / total_pending * 100) if total_pending else 0
            bar_w = int(pct / 100 * 15)
            lines.append(f"    {m['name']:<30s} {_fmt_inr(m['pending_amount']):>10s} {'#' * bar_w}{'-' * (15 - bar_w)} {pct:.0f}%")

    # ================================================================
    # SECTION 3: CASHFLOW TIMELINE (visual)
    # ================================================================
    lines.append("\n--- CASHFLOW TIMELINE " + "-" * 36)
    weekly_totals = schedule.get("weekly_totals", {})
    week_dates = schedule.get("week_dates", {})
    if weekly_totals:
        max_amt = max(weekly_totals.values()) if weekly_totals.values() else 1
        running_total = 0
        lines.append("")
        for cw in sorted(weekly_totals.keys(), key=lambda x: int(re.search(r'\d+', x).group())):
            amt = weekly_totals[cw]
            if amt <= 0:
                continue
            running_total += amt
            bar_w = int(amt / max_amt * 20) if max_amt else 0
            date_str = week_dates.get(cw, "")
            lines.append(f"  {cw:>4s} {date_str:>8s}  {'#' * bar_w:<20s} {_fmt_inr(amt):>12s}  (cum: {_fmt_inr(running_total)})")

        lines.append(f"\n  Total scheduled: {_fmt_inr(running_total)}")
        if total_pending > running_total:
            gap = total_pending - running_total
            lines.append(f"  Gap (unscheduled): {_fmt_inr(gap)} - needs collection dates assigned")

    # ================================================================
    # SECTION 4: RISK REGISTER
    # ================================================================
    lines.append("\n--- RISK REGISTER " + "-" * 39)
    risks = []

    for m in machines_list:
        scheduled = sum(m["weekly_payments"].values())
        gap = m["pending_amount"] - scheduled

        # Zero-advance risk
        if m["advance_received"] == 0 and m["pending_amount"] > 0:
            risks.append({
                "level": "HIGH",
                "machine": m["name"],
                "project": m["project_no"],
                "issue": "No advance received — full exposure",
                "amount": m["pending_amount"],
                "action": "Demand advance payment before further fabrication. Standard terms: 25-30% upfront.",
            })

        # Unscheduled payment risk
        if gap > 500000:
            risks.append({
                "level": "HIGH" if gap > 5000000 else "MEDIUM",
                "machine": m["name"],
                "project": m["project_no"],
                "issue": f"{_fmt_inr(gap)} pending with no scheduled collection week",
                "amount": gap,
                "action": "Agree milestone payment dates with customer. Tie to FAT/dispatch/installation.",
            })

        # Stalled project risk (from MCT Orders)
        for p in projects:
            pn = p["project_no"].replace(".0", "")
            if pn == m["project_no"] and "stall" in p.get("stage", "").lower():
                risks.append({
                    "level": "CRITICAL",
                    "machine": m["name"],
                    "project": m["project_no"],
                    "issue": f"Project STALLED — {_fmt_inr(m['pending_amount'])} at risk",
                    "amount": m["pending_amount"],
                    "action": "Escalate to Rushabh. Schedule customer call. Assess if project should be cancelled or renegotiated.",
                })

    # Concentration risk entry
    if machines_list:
        sorted_m = sorted(machines_list, key=lambda x: -x["pending_amount"])
        top1_pct = (sorted_m[0]["pending_amount"] / total_pending * 100) if total_pending else 0
        if top1_pct > 25:
            risks.append({
                "level": "HIGH" if top1_pct > 35 else "MEDIUM",
                "machine": sorted_m[0]["name"],
                "project": sorted_m[0]["project_no"],
                "issue": f"Single customer = {top1_pct:.0f}% of receivables",
                "amount": sorted_m[0]["pending_amount"],
                "action": "Diversify pipeline. Accelerate collection from this customer. Consider credit insurance for large export orders.",
            })

    # Low collection rate
    if collection_rate < 25:
        risks.append({
            "level": "HIGH",
            "machine": "PORTFOLIO",
            "project": "ALL",
            "issue": f"Collection rate only {collection_rate:.0f}% — cash conversion weak",
            "amount": total_pending,
            "action": "Review payment terms across all projects. Push for milestone-based collections. Consider offering early-payment discounts.",
        })

    risks.sort(key=lambda r: {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}.get(r["level"], 4))

    if risks:
        lines.append("")
        for i, r in enumerate(risks[:5], 1):
            lines.append(f"  {i}. {_risk_badge(r['level'])} - {r['machine']}")
            lines.append(f"     {r['issue']}")
            lines.append(f"     -> {r['action'][:80]}")
            lines.append("")
        if len(risks) > 5:
            lines.append(f"  + {len(risks) - 5} more risks (ask for details)")
    else:
        lines.append("  No significant risks identified.")

    # ================================================================
    # SECTION 5: CFO RECOMMENDATIONS
    # ================================================================
    lines.append("--- CFO RECOMMENDATIONS " + "-" * 34)
    lines.append("")

    recs = []
    critical_count = sum(1 for r in risks if r["level"] == "CRITICAL")
    high_count = sum(1 for r in risks if r["level"] == "HIGH")
    zero_advance = [m for m in machines_list if m["advance_received"] == 0 and m["pending_amount"] > 0]

    if critical_count:
        recs.append(f"URGENT: {critical_count} critical risk(s). Escalate stalled projects this week.")

    if zero_advance:
        names = ", ".join(m["name"].split()[0] for m in zero_advance[:3])
        recs.append(f"COLLECTIONS: {len(zero_advance)} machines ({names}) have zero advance. Halt fabrication.")

    if collection_rate < 30:
        recs.append(f"CASH CONVERSION: {collection_rate:.0f}% vs 25-30% target. Push milestone payments pre-dispatch.")

    if unscheduled > 5000000:
        recs.append(f"SCHEDULING: {_fmt_inr(unscheduled)} has no collection date. Assign milestones to weeks.")

    top3_pct_val = (sum(m["pending_amount"] for m in sorted(machines_list, key=lambda x: -x["pending_amount"])[:3]) / total_pending * 100) if total_pending and machines_list else 0
    if top3_pct_val > 55:
        recs.append(f"DIVERSIFY: Top 3 = {top3_pct_val:.0f}% of A/R. Accelerate collections from largest accounts.")

    if not recs:
        recs.append("Financial position healthy. Continue monitoring weekly.")

    for i, rec in enumerate(recs, 1):
        lines.append(f"  {i}. {rec}")
        lines.append("")

    lines.append("=" * 58)
    lines.append("End of CFO Report -- Plutus, Chief of Finance")

    return "\n".join(lines)


async def finance_overview(query: str, context: Optional[Dict] = None) -> str:
    """
    General finance query handler. Plutus analyzes the question and pulls
    from the right data sources.
    """
    q = query.lower()

    if any(kw in q for kw in ["dashboard", "cfo", "report", "overview", "summary", "health", "risk", "visual"]):
        return await cfo_dashboard(context)

    if any(kw in q for kw in ["order book", "outstanding", "receivable", "balance", "paid", "collected"]):
        return await order_book_status(context)

    if any(kw in q for kw in ["cashflow", "cash flow", "forecast", "when", "expected", "inflow"]):
        return await cashflow_forecast(context)

    if any(kw in q for kw in ["revenue", "turnover", "history", "annual", "yearly", "p&l"]):
        return await revenue_history(query, context)

    return await cfo_dashboard(context)


__all__ = [
    "finance_overview",
    "order_book_status",
    "cashflow_forecast",
    "revenue_history",
    "cfo_dashboard",
]
