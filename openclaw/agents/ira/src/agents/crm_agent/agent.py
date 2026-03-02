#!/usr/bin/env python3
"""
MNEMOSYNE — The Keeper of Relationships
========================================

Named after the Greek goddess of memory and mother of the Muses,
Mnemosyne is Ira's CRM agent. She remembers every contact, every
conversation, every deal, every email thread. She is the institutional
memory of Machinecraft's sales operation.

Personality:
    - Meticulous and organized, but warm — she cares about people, not just data
    - She remembers details others forget: "Last time we spoke, Hans mentioned
      his daughter was starting university"
    - Protective of relationships — she'll push back if Ira tries to email
      someone too aggressively
    - She has opinions: "This lead has gone cold. Try a different angle."
    - She speaks in facts, not fluff: "3 emails sent, 0 replies, 47 days silent"

Role in the Pantheon:
    Athena asks Mnemosyne: "Who should we email today?"
    Mnemosyne replies: "Here are 5 leads ready for drip. TSN is critical —
    they haven't replied to 2 emails. I'd suggest a completely different
    approach. Parat replied last week, they're warm — follow up on the
    quote they asked about."

Functions:
    lookup_contact(query)     — Find a contact by name, email, or company
    get_lead_brief(email)     — Full brief on a lead (for Calliope to write emails)
    get_drip_candidates()     — Who's ready for the next drip email
    get_pipeline_overview()   — Full pipeline health for Athena
    record_interaction(...)   — Log an email, call, or meeting
    suggest_next_action(email) — What should we do next with this lead?

Usage:
    from openclaw.agents.ira.src.agents.crm_agent.agent import (
        lookup_contact, get_lead_brief, get_drip_candidates,
        get_pipeline_overview, suggest_next_action,
    )

    # Athena asks: "Tell me about TSN"
    brief = await lookup_contact("TSN")

    # Athena asks: "Who should we email today?"
    candidates = await get_drip_candidates()
"""

import json
import logging
import os
import sys
from dataclasses import dataclass, asdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger("ira.mnemosyne")

AGENT_DIR = Path(__file__).parent.parent.parent.parent
PROJECT_ROOT = AGENT_DIR.parent.parent.parent
sys.path.insert(0, str(AGENT_DIR / "src" / "crm"))

try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

MNEMOSYNE_SYSTEM_PROMPT = """You are Mnemosyne, the Keeper of Relationships at Machinecraft Technologies.
You are part of Ira's pantheon of agents. You manage the CRM — every contact,
lead, conversation, and deal flows through you.

Your personality:
- Meticulous with data but warm about people. You remember the human details.
- You speak in concrete facts: numbers, dates, names. Never vague.
- You have opinions about sales strategy based on the data you see.
- You're protective of relationships — you'll flag if someone is being over-contacted.
- You think in terms of pipeline health, not just individual leads.

When asked about a contact or lead, give a brief that includes:
1. Who they are (name, company, country, role)
2. Relationship history (emails exchanged, replies, last contact)
3. Deal status (stage, value if known)
4. Your recommendation (what to do next)

Keep responses concise and actionable. You're briefing Athena, not writing an essay."""


@dataclass
class LeadBrief:
    """Mnemosyne's brief on a lead — everything needed to write a personalized email."""
    email: str
    name: str
    company: str
    country: str
    title: str
    industry: str
    priority: str
    deal_stage: str
    drip_stage: int
    emails_sent: int
    emails_received: int
    last_email_sent: Optional[str]
    last_reply_at: Optional[str]
    reply_quality: str
    days_since_contact: int
    conversation_summary: Optional[str]
    notes: str
    recommendation: str

    def to_dict(self) -> Dict:
        return asdict(self)

    def to_text(self) -> str:
        lines = [
            f"{self.name} at {self.company} ({self.country})",
            f"Priority: {self.priority} | Stage: {self.deal_stage} | Drip: {self.drip_stage}/5",
            f"Emails: {self.emails_sent} sent, {self.emails_received} received",
        ]
        if self.last_email_sent:
            lines.append(f"Last email: {self.last_email_sent} ({self.days_since_contact}d ago)")
        if self.reply_quality:
            lines.append(f"Last reply: {self.reply_quality}")
        if self.conversation_summary:
            lines.append(f"History: {self.conversation_summary}")
        if self.notes:
            lines.append(f"Notes: {self.notes[:200]}")
        lines.append(f"Recommendation: {self.recommendation}")
        return "\n".join(lines)


def _get_crm():
    """Lazy-load the CRM."""
    try:
        from ira_crm import get_crm
        return get_crm()
    except ImportError:
        logger.warning("Ira CRM not available")
        return None


async def lookup_contact(query: str, context: Optional[Dict] = None) -> str:
    """
    Look up a contact, lead, or company in Ira's CRM.

    Mnemosyne searches by name, email, or company and returns
    a human-readable brief with relationship history and recommendations.
    """
    crm = _get_crm()
    if not crm:
        return "CRM not available. I can't look up contacts right now."

    results = crm.search_contacts(query, limit=5)
    if not results:
        return f"I don't have anyone matching '{query}' in my records."

    briefs = []
    for contact in results:
        lead = crm.get_lead(contact.email)
        conv_summary = crm.get_conversation_summary(contact.email)

        if lead:
            days_since = 0
            if lead.last_email_sent:
                try:
                    last = datetime.fromisoformat(lead.last_email_sent)
                    days_since = (datetime.now() - last).days
                except (ValueError, TypeError):
                    pass

            recommendation = _generate_recommendation(lead, days_since, conv_summary)

            brief = LeadBrief(
                email=contact.email,
                name=contact.name or f"{contact.first_name} {contact.last_name}".strip(),
                company=contact.company or lead.company,
                country=contact.country or lead.country,
                title=contact.title,
                industry=contact.industry or lead.industry,
                priority=lead.priority,
                deal_stage=lead.deal_stage,
                drip_stage=lead.drip_stage,
                emails_sent=lead.emails_sent,
                emails_received=lead.emails_received,
                last_email_sent=lead.last_email_sent,
                last_reply_at=lead.last_reply_at,
                reply_quality=lead.reply_quality,
                days_since_contact=days_since,
                conversation_summary=conv_summary,
                notes=lead.notes,
                recommendation=recommendation,
            )
            briefs.append(brief.to_text())
        else:
            briefs.append(
                f"{contact.name or contact.email} at {contact.company} ({contact.country})\n"
                f"  Contact only — not in sales pipeline yet.\n"
                f"  Notes: {contact.notes[:150] if contact.notes else 'None'}"
            )

    return "\n\n---\n\n".join(briefs)


async def get_lead_brief(email: str, context: Optional[Dict] = None) -> str:
    """Get a full brief on a specific lead for email personalization."""
    crm = _get_crm()
    if not crm:
        return "CRM not available."

    lead = crm.get_lead(email)
    if not lead:
        return f"No lead found for {email}."

    contact = crm.get_contact(email)
    conv_summary = crm.get_conversation_summary(email)
    conversations = crm.get_conversation_context(email, limit=5)

    days_since = 0
    if lead.last_email_sent:
        try:
            days_since = (datetime.now() - datetime.fromisoformat(lead.last_email_sent)).days
        except (ValueError, TypeError):
            pass

    recommendation = _generate_recommendation(lead, days_since, conv_summary)

    parts = [
        f"LEAD BRIEF: {contact.name if contact else email}",
        f"Company: {lead.company} | Country: {lead.country}",
        f"Priority: {lead.priority} | Deal: {lead.deal_stage}",
        f"Drip stage: {lead.drip_stage}/5 | Emails: {lead.emails_sent} sent, {lead.emails_received} replies",
    ]

    if days_since > 0:
        parts.append(f"Last contact: {days_since} days ago")
    if lead.reply_quality:
        parts.append(f"Reply quality: {lead.reply_quality}")
    if conv_summary:
        parts.append(f"\nConversation history: {conv_summary}")
    if conversations:
        parts.append("\nRecent threads:")
        for c in conversations[:3]:
            arrow = "<-" if c.direction == "inbound" else "->"
            parts.append(f"  {arrow} {c.subject[:60]} ({c.date[:10] if c.date else '?'})")
    if contact and contact.notes:
        parts.append(f"\nNotes: {contact.notes[:300]}")

    parts.append(f"\nMnemosyne's recommendation: {recommendation}")

    return "\n".join(parts)


async def get_drip_candidates(context: Optional[Dict] = None) -> str:
    """
    Get leads ready for the next drip email.

    Mnemosyne returns a prioritized list with her recommendations
    for each lead.
    """
    crm = _get_crm()
    if not crm:
        return "CRM not available."

    leads = crm.get_leads_ready_for_drip(max_results=15)
    if not leads:
        return "No leads are due for drip emails right now. Everyone is either too recent or has replied."

    lines = [f"DRIP CANDIDATES: {len(leads)} leads ready\n"]

    for lead in leads:
        days_since = 0
        if lead.last_email_sent:
            try:
                days_since = (datetime.now() - datetime.fromisoformat(lead.last_email_sent)).days
            except (ValueError, TypeError):
                pass

        conv = crm.get_conversation_summary(lead.email)
        rec = _generate_recommendation(lead, days_since, conv)

        lines.append(
            f"[{lead.priority.upper():8s}] {lead.company:35s} "
            f"drip={lead.drip_stage} sent={lead.emails_sent} "
            f"{'REPLIED' if lead.emails_received > 0 else f'{days_since}d silent'}"
        )
        lines.append(f"           -> {rec}")

    return "\n".join(lines)


async def get_pipeline_overview(context: Optional[Dict] = None) -> str:
    """Full pipeline health overview for Athena."""
    crm = _get_crm()
    if not crm:
        return "CRM not available."

    stats = crm.get_pipeline_stats()
    drip = crm.get_drip_stats()

    lines = [
        "PIPELINE OVERVIEW",
        f"Total leads: {stats['total_leads']}",
        f"Reply rate: {stats['reply_rate']:.1%}",
        f"Emails sent: {stats['total_emails_sent']} | Replies: {stats['total_replies']}",
        "",
        "By stage:",
    ]
    for stage, count in sorted(stats.get("by_stage", {}).items()):
        lines.append(f"  {stage or 'new':15s} {count}")

    lines.extend([
        "",
        "By priority:",
    ])
    for pri, count in sorted(stats.get("by_priority", {}).items()):
        lines.append(f"  {pri:15s} {count}")

    lines.extend([
        "",
        f"Drip ready: {drip['ready_for_drip']}",
        f"Engaged: {drip['engaged']}",
        f"Bounced: {drip['bounced']}",
        f"Unsubscribed: {drip['unsubscribed']}",
    ])

    return "\n".join(lines)


async def list_all_customers(context: Optional[Dict] = None) -> str:
    """
    List confirmed Machinecraft customers — companies that actually BOUGHT machines.

    Threads together ALL customer data sources:
    1. MCT Orders 2025.xlsx — active 2024-2025 orders with project numbers
    2. Machinecraft Nov 2025 Order Book.xlsx — current order book with statuses
    3. customer_orders.json — confirmed orders with pricing/specs (2020-2025)
    4. Machinecraft Machine Order Analysis.xlsx — historical orders since 2015
    5. Clients MC EUROPE.xlsx — European customers who bought machines
    6. List of Customers - Machinecraft.xlsx — India/global customers 2014-2017
    7. Mem0 machinecraft_customers — any additional known customers
    """
    by_country: Dict[str, list] = {}
    seen_companies: set = set()
    imports_dir = PROJECT_ROOT / "data" / "imports"

    _country_normalize = {
        "SWEDEN": "Sweden", "NORWAY": "Norway", "ROMANIA": "Romania",
        "RUSSIA": "Russia", "UK": "UK", "United Kingdom": "UK",
        "Delhi": "India", "Pune": "India", "Mumbai": "India",
        "Hyderabad": "India", "Bangalore": "India", "London": "UK",
        "Norrkoping": "Sweden", "Amsterdam": "Netherlands",
        "Tamil Nadu": "India", "India ": "India",
    }

    def _normalize_key(name: str) -> str:
        """Collapse variants like 'Anatomic SITT AB' / 'Anatomic SITT' / 'Anatomic' to one key."""
        import re
        k = name.lower().strip()
        k = re.sub(r'\s*(pvt\.?\s*ltd\.?|ltd\.?|ab|a/s|s\.?r\.?l\.?|gmbh|kft\.?|co\.?|fzco)\s*$', '', k, flags=re.I)
        k = re.sub(r'\s*(x\s*\d+)$', '', k)
        k = re.sub(r'[^a-z0-9]', '', k)
        return k

    def _add(company: str, country: str, detail: str):
        key = _normalize_key(company)
        if not key or key in seen_companies:
            return
        seen_companies.add(key)
        country = _country_normalize.get(country.strip(), country.strip()) or "Unknown"
        if country not in by_country:
            by_country[country] = []
        by_country[country].append(f"{company} — {detail}")

    # --- Source 1: MCT Orders 2025 (active orders with project numbers) ---
    try:
        import openpyxl
        path = imports_dir / "02_Orders_and_POs" / "MCT Orders 2025.xlsx"
        if path.exists():
            wb = openpyxl.load_workbook(str(path), read_only=True)
            ws = wb[wb.sheetnames[0]]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if r[1] and str(r[1]).strip() and str(r[1]).strip() != "Currency INR":
                    customer = str(r[1]).strip()
                    region = str(r[2] or "Unknown").strip()
                    machine = str(r[3] or "").strip()
                    stage = str(r[5] or "").strip()
                    proj = str(r[0] or "").strip()
                    detail = machine
                    if proj:
                        detail = f"[#{proj}] {machine}"
                    if stage:
                        detail += f" ({stage})"
                    _add(customer, region, detail)
            wb.close()
    except Exception as e:
        logger.debug(f"MCT Orders 2025 load failed: {e}")

    # --- Source 2: Nov 2025 Order Book (current projects) ---
    # Columns: 0=No, 1=Project Code, 2=None, 3=Description, 4=Customer, 5=None, 6=Region, 7=Status
    try:
        path = imports_dir / "02_Orders_and_POs" / "Machinecraft Nov 2025 Order Book.xlsx"
        if path.exists():
            wb = openpyxl.load_workbook(str(path), read_only=True)
            ws = wb["Projects"]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if len(r) < 8:
                    continue
                customer = str(r[4] or "").strip()
                if not customer or customer == "Next in Line":
                    continue
                region = str(r[6] or "Unknown").strip()
                desc = str(r[3] or "").strip()
                status = str(r[7] or "").strip()
                proj = str(r[1] or "").strip()
                detail = f"[#{proj}] {desc}" if proj else desc
                if status:
                    detail += f" ({status})"
                _add(customer, region, detail)
            wb.close()
    except Exception as e:
        logger.debug(f"Nov 2025 Order Book load failed: {e}")

    # --- Source 3: customer_orders.json (confirmed with pricing) ---
    try:
        orders_file = PROJECT_ROOT / "data" / "knowledge" / "customer_orders.json"
        if orders_file.exists():
            with open(orders_file) as f:
                orders_data = json.load(f)
            for order in orders_data.get("orders", []):
                company = order.get("customer", "")
                country = order.get("country", "Unknown")
                machine = order.get("machine_model", "")
                year = order.get("year", "")
                application = order.get("application", "")
                detail = f"{machine} ({year})"
                if application:
                    detail += f" [{application}]"
                _add(company, country, detail)
    except Exception as e:
        logger.debug(f"customer_orders.json load failed: {e}")

    # --- Source 4: Machine Order Analysis (historical since 2015) ---
    # Columns: 0=Sr, 1=PO Date, 2=Bill Date, 3=Lifecycle, 4=Company, 5=City, 6=Country,
    #          7-9=Forming Size, 10=Optional, 11=Cost INR, 12=Cost EUR, 13=Product
    try:
        path = imports_dir / "02_Orders_and_POs" / "Machinecraft Machine Order Analysis.xlsx"
        if path.exists():
            wb = openpyxl.load_workbook(str(path), read_only=True)
            ws = wb["Analysis"]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if len(r) < 7:
                    continue
                company = str(r[4] or "").strip()
                if not company:
                    continue
                city = str(r[5] or "").strip()
                country = str(r[6] or "Unknown").strip()
                product = str(r[13] or "").strip() if len(r) > 13 else ""
                po_date = str(r[1] or "").strip()
                detail = product[:60] if product and not product.startswith("=") else "Machine order"
                if city:
                    detail += f" ({city})"
                if po_date and not po_date.startswith("="):
                    detail += f" PO:{po_date[:10]}"
                _add(company, country, detail)
            wb.close()
    except Exception as e:
        logger.debug(f"Machine Order Analysis load failed: {e}")

    # --- Source 5: Clients MC EUROPE (European customers) ---
    # Mixed layout: rows ~19-24 = [None, Company, Country, Machine, Year]
    #               rows ~37+  = [Company, Country, Website, Contact, Email, MachineInfo]
    try:
        path = imports_dir / "07_Leads_and_Contacts" / "Clients MC EUROPE.xlsx"
        if path.exists():
            wb = openpyxl.load_workbook(str(path), read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            for r in rows:
                cols = list(r) + [None] * 8
                c0 = str(cols[0] or "").strip()
                c1 = str(cols[1] or "").strip()
                c5 = str(cols[5] or "").strip()

                # Format 1: [None, Company, Country, Machine, Year] (recent EU deliveries)
                if not c0 and c1 and len(c1) > 1:
                    country = str(cols[2] or "").strip()
                    machine = str(cols[3] or "").strip()
                    year = str(cols[4] or "").strip()
                    if machine and ("PF" in machine or "ATF" in machine or "IMG" in machine or "AM" in machine):
                        detail = f"{machine} ({year})" if year else machine
                        _add(c1, country, detail)

                # Format 2: [Company, Country, Website, Contact, Email, MachineInfo]
                if c0 and c1 and c5:
                    if "Vacuum Forming" in c5 or "INLINE" in c5 or "PF" in c5:
                        _add(c0, c1, c5[:60])
            wb.close()
    except Exception as e:
        logger.debug(f"Clients MC EUROPE load failed: {e}")

    # --- Source 6: List of Customers - Machinecraft.xlsx (India 2014-2017) ---
    try:
        path = imports_dir / "07_Leads_and_Contacts" / "List of Customers - Machinecraft.xlsx"
        if path.exists():
            wb = openpyxl.load_workbook(str(path), read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                ncols = len(rows[0])
                for r in rows[1:]:
                    r = list(r)[:ncols]
                    if len(r) >= 10 and r[0] and r[1]:
                        type_val = str(r[9] or "")
                        if "Machine" not in type_val:
                            continue
                        company = str(r[1]).strip()
                        region = str(r[2] or "").strip()
                        goods = str(r[7] or "").strip()
                        country = region.split(",")[-1].strip() if "," in region else region
                        _add(company, country, f"{goods[:60]} (FY {sheet_name})")
            wb.close()
    except Exception as e:
        logger.debug(f"Customer list Excel load failed: {e}")

    # --- Source 7: Mem0 (additional intelligence) ---
    try:
        from openclaw.agents.ira.src.memory.mem0_memory import get_mem0_service
        mem0 = get_mem0_service()
        memories = mem0.search(
            "Machinecraft confirmed customers who bought machines",
            "machinecraft_customers", limit=20,
        )
        for m in memories:
            text = m.memory
            if any(kw in text.lower() for kw in ["order", "bought", "sold", "supplied", "installed", "confirmed"]):
                if "Additional intelligence" not in by_country:
                    by_country["Additional intelligence"] = []
                by_country["Additional intelligence"].append(f"[mem0] {text[:150]}")
    except Exception:
        pass

    if not by_country:
        return "No confirmed customer data found. Check data/imports/ folders."

    total = sum(len(v) for v in by_country.values())
    lines = [
        f"MACHINECRAFT CONFIRMED CUSTOMERS — {total} entries",
        "(Companies that BOUGHT machines — not leads or prospects)",
        f"Sources: MCT Orders 2025, Order Book, Order Analysis, MC Europe, Historical list, Mem0\n",
    ]
    for country in sorted(by_country.keys()):
        lines.append(f"\n{country} ({len(by_country[country])}):")
        for entry in sorted(by_country[country]):
            lines.append(f"  • {entry}")

    return "\n".join(lines)


async def suggest_next_action(email: str, context: Optional[Dict] = None) -> str:
    """Mnemosyne suggests what to do next with a specific lead."""
    crm = _get_crm()
    if not crm:
        return "CRM not available."

    lead = crm.get_lead(email)
    if not lead:
        return f"No lead found for {email}."

    days_since = 0
    if lead.last_email_sent:
        try:
            days_since = (datetime.now() - datetime.fromisoformat(lead.last_email_sent)).days
        except (ValueError, TypeError):
            pass

    conv = crm.get_conversation_summary(email)
    return _generate_recommendation(lead, days_since, conv)


def _generate_recommendation(lead, days_since: int, conv_summary: Optional[str]) -> str:
    """Mnemosyne's opinion on what to do next."""

    if lead.reply_quality == "bounce":
        return "Email bounced. Verify the address or find an alternative contact at this company."

    if lead.reply_quality == "unsubscribe":
        return "They asked to stop. Respect it. Remove from drip."

    if lead.reply_quality == "engaged":
        if days_since > 7:
            return "They showed interest but it's been a week. Follow up with specifics — quote, specs, or meeting invite."
        return "Warm lead! They're engaged. Move to proposal stage if ready, or keep the conversation going."

    if lead.reply_quality == "polite_decline":
        if days_since > 60:
            return "They declined before but it's been 2+ months. Try a fresh angle — new product, case study, or industry news."
        return "They politely declined recently. Give them space. Try again in a couple months with something new."

    # No reply yet
    if lead.emails_sent == 0:
        return "Fresh lead. Send the first drip email — personalized intro with a specific hook."

    if lead.emails_sent >= 3 and lead.emails_received == 0:
        if days_since > 30:
            return (
                f"3+ emails, zero replies, {days_since}d silent. "
                "This approach isn't working. Try something radically different: "
                "shorter email, different subject line, or try reaching them on LinkedIn."
            )
        return "Multiple emails with no response. Space it out and try a different angle next time."

    if lead.emails_sent >= 1 and lead.emails_received == 0:
        if days_since > 14:
            return "Sent an email 2+ weeks ago, no reply. Time for the next drip stage — try a value-prop or case study angle."
        return "Recently emailed, waiting for response. Give it a few more days."

    if conv_summary:
        return f"Has history: {conv_summary} Continue the conversation thread."

    return "Standard drip sequence. Send the next stage email."


# Expose for __init__.py
__all__ = [
    "lookup_contact",
    "list_all_customers",
    "get_lead_brief",
    "get_drip_candidates",
    "get_pipeline_overview",
    "suggest_next_action",
    "LeadBrief",
]
