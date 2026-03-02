#!/usr/bin/env python3
"""
Deep Company Research - Thorough web research for thermoforming market analysis.

This script performs comprehensive research on each company:
1. Scrapes multiple pages (homepage, about, services, products, capabilities)
2. Uses GPT-4o for intelligent extraction of thermoforming-specific data
3. Searches for additional information when primary sources are incomplete
4. Validates and cross-references extracted data

Usage:
    python deep_research.py --limit 50 --workers 5
"""

import os
import sys
import json
import re
import time
import asyncio
import logging
import warnings
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass, field, asdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse, urljoin
from pathlib import Path

import requests
import psycopg2
import psycopg2.extras
from openai import OpenAI

# Suppress warnings
warnings.filterwarnings('ignore')
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Import from centralized config
MARKET_RESEARCH_DIR = Path(__file__).parent
SKILLS_DIR = MARKET_RESEARCH_DIR.parent
AGENT_DIR = SKILLS_DIR.parent
sys.path.insert(0, str(AGENT_DIR))

try:
    from config import DATABASE_URL, OPENAI_API_KEY
except ImportError:
    DATABASE_URL = os.getenv('DATABASE_URL', '')
    OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')

JINA_API_KEY = os.getenv('JINA_API_KEY', '')

# Timeouts
SCRAPE_TIMEOUT = 30
MAX_CONTENT_LENGTH = 50000  # Max chars to send to LLM

@dataclass
class CompanyResearch:
    """Comprehensive company research data."""
    name: str
    website: str
    
    # Basic info
    country: Optional[str] = None
    city: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    employees: Optional[int] = None
    established_year: Optional[int] = None
    revenue: Optional[str] = None
    
    # Thermoforming specifics
    thermoforming_services: List[str] = field(default_factory=list)
    forming_types: List[str] = field(default_factory=list)  # vacuum, pressure, twin-sheet, etc.
    machine_brands: List[str] = field(default_factory=list)
    machine_models: List[str] = field(default_factory=list)
    machine_count: Optional[int] = None
    max_sheet_size: Optional[str] = None
    max_forming_area: Optional[str] = None
    max_draw_depth: Optional[str] = None
    
    # Materials & capabilities
    materials: List[str] = field(default_factory=list)
    material_thicknesses: Optional[str] = None
    
    # Markets
    industries: List[str] = field(default_factory=list)
    applications: List[str] = field(default_factory=list)
    certifications: List[str] = field(default_factory=list)
    
    # Additional services
    secondary_services: List[str] = field(default_factory=list)  # CNC, assembly, etc.
    
    # Research metadata
    summary: Optional[str] = None
    confidence: float = 0.0
    sources_scraped: List[str] = field(default_factory=list)
    research_notes: Optional[str] = None


class JinaReader:
    """Jina Reader API for clean web scraping."""
    
    BASE_URL = "https://r.jina.ai/"
    
    def __init__(self, api_key: str = ""):
        self.api_key = api_key
        self.session = requests.Session()
        self.session.headers.update({
            'Accept': 'text/plain',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        })
        if api_key:
            self.session.headers['Authorization'] = f'Bearer {api_key}'
    
    def read(self, url: str, timeout: int = SCRAPE_TIMEOUT) -> Tuple[str, bool]:
        """Read a URL and return clean markdown content."""
        try:
            response = self.session.get(
                f"{self.BASE_URL}{url}",
                timeout=timeout
            )
            if response.status_code == 200:
                content = response.text
                # Clean up the content
                content = self._clean_content(content)
                return content, True
            else:
                return f"Error: {response.status_code}", False
        except requests.Timeout:
            return "Timeout", False
        except Exception as e:
            return str(e), False
    
    def _clean_content(self, content: str) -> str:
        """Clean up scraped content."""
        # Remove excessive whitespace
        content = re.sub(r'\n{3,}', '\n\n', content)
        content = re.sub(r' {2,}', ' ', content)
        # Remove common boilerplate
        content = re.sub(r'Cookie Policy.*?Accept', '', content, flags=re.DOTALL | re.IGNORECASE)
        content = re.sub(r'Privacy Policy.*?Terms', '', content, flags=re.DOTALL | re.IGNORECASE)
        return content.strip()


class DeepResearcher:
    """Deep research engine for thermoforming companies."""
    
    # Pages to scrape for each company
    PAGE_SUFFIXES = [
        '',  # Homepage
        '/about', '/about-us', '/about-us/', '/company', '/company/',
        '/services', '/services/', '/capabilities', '/capabilities/',
        '/products', '/products/', '/thermoforming', '/thermoforming/',
        '/equipment', '/equipment/', '/machines', '/machinery',
        '/materials', '/industries', '/markets', '/applications',
        '/contact', '/contact-us',
    ]
    
    THERMOFORMING_EXTRACTION_PROMPT = """You are a market research analyst specializing in the thermoforming/plastic forming industry.

Analyze the following website content and extract detailed information about this company's thermoforming capabilities.

IMPORTANT: Only extract information that is EXPLICITLY stated or strongly implied in the content. Do not make assumptions.

Extract the following (return null/empty if not found):

1. **Basic Company Info:**
   - Country/location
   - City
   - Full address
   - Phone number
   - Email
   - Number of employees (estimate if range given)
   - Year established/founded
   - Revenue (if mentioned)

2. **Thermoforming Services** (list all that apply):
   - Vacuum forming
   - Pressure forming
   - Twin-sheet forming
   - Heavy gauge forming
   - Thin gauge forming
   - Drape forming
   - Matched mold forming
   - Custom thermoforming
   - Prototype thermoforming
   - Production thermoforming

3. **Equipment/Machines:**
   - Machine brands (e.g., Kiefel, Illig, Geiss, Brown, Sencorp, ZMD, Cannon, etc.)
   - Specific machine models
   - Number of machines (if mentioned)
   - Maximum sheet size they can handle
   - Maximum forming area
   - Maximum draw depth

4. **Materials they work with:**
   - ABS, HDPE, PETG, PC, PP, PVC, Acrylic, HIPS, etc.
   - Material thickness range they can handle

5. **Industries/Markets served:**
   - Automotive, Medical, Aerospace, Food packaging, Electronics, etc.

6. **Applications/Products:**
   - What specific products do they make? (trays, enclosures, panels, etc.)

7. **Certifications:**
   - ISO, FDA, automotive certifications, etc.

8. **Secondary Services:**
   - CNC trimming, assembly, painting, silk screening, etc.

9. **Summary:**
   - Write a 2-3 sentence summary of this company's thermoforming capabilities

10. **Confidence Score:**
    - Rate 0.0-1.0 how confident you are in the extracted data based on content quality

Return as JSON with this exact structure:
{
    "country": "string or null",
    "city": "string or null", 
    "address": "string or null",
    "phone": "string or null",
    "email": "string or null",
    "employees": "number or null",
    "established_year": "number or null",
    "revenue": "string or null",
    "thermoforming_services": ["list of services"],
    "forming_types": ["vacuum forming", "pressure forming", etc.],
    "machine_brands": ["list of brands"],
    "machine_models": ["list of models"],
    "machine_count": "number or null",
    "max_sheet_size": "string or null",
    "max_forming_area": "string or null",
    "max_draw_depth": "string or null",
    "materials": ["list of materials"],
    "material_thicknesses": "string or null",
    "industries": ["list of industries"],
    "applications": ["list of applications/products"],
    "certifications": ["list of certifications"],
    "secondary_services": ["list of services"],
    "summary": "string",
    "confidence": 0.0-1.0,
    "research_notes": "any important notes about limitations or uncertainties"
}

WEBSITE CONTENT:
"""

    def __init__(self):
        self.reader = JinaReader(JINA_API_KEY)
        self.client = OpenAI(api_key=OPENAI_API_KEY)
        self.conn = psycopg2.connect(DATABASE_URL)
    
    def get_pending_companies(self, limit: int = 100) -> List[Dict]:
        """Get companies that need research."""
        cur = self.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT company_id, name, website, country, email
            FROM market_research.companies
            WHERE website IS NOT NULL 
              AND website NOT LIKE '%%youtube%%'
              AND website NOT LIKE '%%linkedin%%'
              AND website NOT LIKE '%%facebook%%'
              AND (last_researched IS NULL 
                   OR (thermoforming_services IS NULL OR thermoforming_services = '[]'::jsonb))
            ORDER BY 
                CASE WHEN last_researched IS NULL THEN 0 ELSE 1 END,
                name
            LIMIT %s
        """, (limit,))
        return cur.fetchall()
    
    def scrape_company_pages(self, base_url: str) -> Tuple[str, List[str]]:
        """Scrape multiple pages from a company website."""
        all_content = []
        scraped_urls = []
        
        # Normalize base URL
        if not base_url.startswith('http'):
            base_url = 'https://' + base_url
        
        parsed = urlparse(base_url)
        base = f"{parsed.scheme}://{parsed.netloc}"
        
        # Try different page URLs
        urls_to_try = set()
        urls_to_try.add(base_url)  # Original URL
        urls_to_try.add(base)  # Just the domain
        
        for suffix in self.PAGE_SUFFIXES:
            urls_to_try.add(urljoin(base, suffix))
        
        # Scrape each URL
        for url in list(urls_to_try)[:15]:  # Limit to 15 pages max
            content, success = self.reader.read(url)
            if success and len(content) > 200:
                # Avoid duplicate content
                content_hash = hash(content[:500])
                if not any(hash(c[:500]) == content_hash for c in all_content):
                    all_content.append(f"\n\n=== PAGE: {url} ===\n\n{content}")
                    scraped_urls.append(url)
                    logger.debug(f"  ✓ Scraped: {url} ({len(content)} chars)")
            
            # Small delay between requests
            time.sleep(0.5)
        
        combined = "\n".join(all_content)
        # Truncate if too long
        if len(combined) > MAX_CONTENT_LENGTH:
            combined = combined[:MAX_CONTENT_LENGTH] + "\n\n[Content truncated...]"
        
        return combined, scraped_urls
    
    def extract_with_llm(self, content: str, company_name: str) -> Dict[str, Any]:
        """Use GPT-4o to extract structured data from content."""
        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": "You are a precise data extraction assistant. Extract only factual information from the provided content. Return valid JSON only."
                    },
                    {
                        "role": "user", 
                        "content": f"{self.THERMOFORMING_EXTRACTION_PROMPT}\n\nCompany Name: {company_name}\n\n{content}"
                    }
                ],
                temperature=0.1,
                max_tokens=2000,
                response_format={"type": "json_object"}
            )
            
            result = json.loads(response.choices[0].message.content)
            return result
            
        except Exception as e:
            logger.error(f"LLM extraction error: {e}")
            return {}
    
    def research_company(self, company: Dict) -> Optional[CompanyResearch]:
        """Perform deep research on a single company."""
        company_id = company['company_id']
        name = company['name']
        website = company['website']
        
        logger.info(f"🔍 Researching: {name}")
        logger.info(f"   Website: {website}")
        
        # Step 1: Scrape multiple pages
        logger.info(f"   Scraping pages...")
        content, scraped_urls = self.scrape_company_pages(website)
        
        if not content or len(content) < 500:
            logger.warning(f"   ⚠ Insufficient content scraped")
            return None
        
        logger.info(f"   Scraped {len(scraped_urls)} pages, {len(content)} chars total")
        
        # Step 2: Extract with LLM
        logger.info(f"   Extracting data with GPT-4o...")
        extracted = self.extract_with_llm(content, name)
        
        if not extracted:
            logger.warning(f"   ⚠ LLM extraction failed")
            return None
        
        # Step 3: Build research object
        research = CompanyResearch(
            name=name,
            website=website,
            country=extracted.get('country') or company.get('country'),
            city=extracted.get('city'),
            address=extracted.get('address'),
            phone=extracted.get('phone'),
            email=extracted.get('email') or company.get('email'),
            employees=extracted.get('employees'),
            established_year=extracted.get('established_year'),
            revenue=extracted.get('revenue'),
            thermoforming_services=extracted.get('thermoforming_services', []),
            forming_types=extracted.get('forming_types', []),
            machine_brands=extracted.get('machine_brands', []),
            machine_models=extracted.get('machine_models', []),
            machine_count=extracted.get('machine_count'),
            max_sheet_size=extracted.get('max_sheet_size'),
            max_forming_area=extracted.get('max_forming_area'),
            max_draw_depth=extracted.get('max_draw_depth'),
            materials=extracted.get('materials', []),
            material_thicknesses=extracted.get('material_thicknesses'),
            industries=extracted.get('industries', []),
            applications=extracted.get('applications', []),
            certifications=extracted.get('certifications', []),
            secondary_services=extracted.get('secondary_services', []),
            summary=extracted.get('summary'),
            confidence=extracted.get('confidence', 0.5),
            sources_scraped=scraped_urls,
            research_notes=extracted.get('research_notes')
        )
        
        # Log results
        services_count = len(research.thermoforming_services)
        machines_count = len(research.machine_brands) + len(research.machine_models)
        logger.info(f"   ✓ Extracted: {services_count} services, {machines_count} machine refs, confidence: {research.confidence:.1%}")
        
        return research
    
    def save_research(self, company_id: str, research: CompanyResearch):
        """Save research results to database."""
        cur = self.conn.cursor()
        
        cur.execute("""
            UPDATE market_research.companies SET
                country = COALESCE(%s, country),
                employees = COALESCE(%s, employees),
                established_year = COALESCE(%s, established_year),
                revenue = COALESCE(%s, revenue),
                thermoforming_services = %s::jsonb,
                machine_models = %s::jsonb,
                machine_count = COALESCE(%s, machine_count),
                max_part_size = COALESCE(%s, max_part_size),
                materials = %s::jsonb,
                industries = %s::jsonb,
                applications = %s::jsonb,
                research_summary = %s,
                research_confidence = %s,
                research_sources = %s::jsonb,
                last_researched = NOW(),
                updated_at = NOW()
            WHERE company_id = %s
        """, (
            research.country,
            research.employees,
            research.established_year,
            research.revenue,
            json.dumps(research.thermoforming_services + research.forming_types),
            json.dumps(research.machine_brands + research.machine_models),
            research.machine_count,
            research.max_sheet_size or research.max_forming_area,
            json.dumps(research.materials),
            json.dumps(research.industries),
            json.dumps(research.applications),
            research.summary,
            research.confidence,
            json.dumps(research.sources_scraped),
            company_id
        ))
        
        self.conn.commit()
    
    def run_batch(self, limit: int = 50, workers: int = 3):
        """Run batch research with parallel processing."""
        companies = self.get_pending_companies(limit)
        
        if not companies:
            logger.info("No companies pending research")
            return
        
        logger.info(f"\n{'='*60}")
        logger.info(f"DEEP RESEARCH: {len(companies)} companies with {workers} workers")
        logger.info(f"{'='*60}\n")
        
        success_count = 0
        fail_count = 0
        start_time = time.time()
        
        def research_single(company):
            try:
                research = self.research_company(company)
                if research:
                    self.save_research(company['company_id'], research)
                    return True, company['name'], research.confidence
                return False, company['name'], 0
            except Exception as e:
                logger.error(f"Error researching {company['name']}: {e}")
                return False, company['name'], 0
        
        # Use ThreadPoolExecutor for parallel processing
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(research_single, c): c for c in companies}
            
            for i, future in enumerate(as_completed(futures), 1):
                success, name, confidence = future.result()
                if success:
                    success_count += 1
                    logger.info(f"[{i}/{len(companies)}] ✓ {name} (confidence: {confidence:.0%})")
                else:
                    fail_count += 1
                    logger.info(f"[{i}/{len(companies)}] ✗ {name}")
        
        elapsed = time.time() - start_time
        
        logger.info(f"\n{'='*60}")
        logger.info(f"COMPLETE")
        logger.info(f"  Success: {success_count}/{len(companies)}")
        logger.info(f"  Failed: {fail_count}")
        logger.info(f"  Time: {elapsed:.1f}s ({elapsed/len(companies):.1f}s per company)")
        logger.info(f"{'='*60}\n")


def export_to_excel(output_path: str = None):
    """Export research data to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Market Research"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    data_fill = PatternFill(start_color="E8F4E8", end_color="E8F4E8", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    headers = [
        "Company Name", "Website", "Country", "Employees", "Established",
        "Thermoforming Services", "Machine Brands/Models", "Machine Count",
        "Max Size", "Materials", "Industries", "Applications",
        "Summary", "Confidence", "Last Researched"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30
    
    # Fetch data
    cur.execute('''
        SELECT name, website, country, employees, established_year,
               thermoforming_services, machine_models, machine_count,
               max_part_size, materials, industries, applications,
               research_summary, research_confidence, last_researched
        FROM market_research.companies
        ORDER BY 
            CASE WHEN research_confidence > 0.5 THEN 0 
                 WHEN thermoforming_services IS NOT NULL AND thermoforming_services != '[]'::jsonb THEN 1 
                 ELSE 2 END,
            research_confidence DESC NULLS LAST,
            name
    ''')
    companies = cur.fetchall()
    
    def join_list(arr):
        if arr and isinstance(arr, list):
            return ", ".join(str(x) for x in arr[:10])  # Limit to 10 items
        return ""
    
    for row_num, c in enumerate(companies, 2):
        has_data = c['thermoforming_services'] and c['thermoforming_services'] != []
        
        ws.cell(row=row_num, column=1, value=c['name'])
        ws.cell(row=row_num, column=2, value=c['website'])
        ws.cell(row=row_num, column=3, value=c['country'])
        ws.cell(row=row_num, column=4, value=c['employees'])
        ws.cell(row=row_num, column=5, value=c['established_year'])
        ws.cell(row=row_num, column=6, value=join_list(c['thermoforming_services']))
        ws.cell(row=row_num, column=7, value=join_list(c['machine_models']))
        ws.cell(row=row_num, column=8, value=c['machine_count'])
        ws.cell(row=row_num, column=9, value=c['max_part_size'])
        ws.cell(row=row_num, column=10, value=join_list(c['materials']))
        ws.cell(row=row_num, column=11, value=join_list(c['industries']))
        ws.cell(row=row_num, column=12, value=join_list(c['applications']))
        ws.cell(row=row_num, column=13, value=c['research_summary'])
        ws.cell(row=row_num, column=14, value=f"{c['research_confidence']*100:.0f}%" if c['research_confidence'] else "")
        ws.cell(row=row_num, column=15, value=c['last_researched'].strftime('%Y-%m-%d') if c['last_researched'] else "Pending")
        
        for col in range(1, 16):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if has_data:
                cell.fill = data_fill
            elif row_num % 2 == 0:
                cell.fill = alt_fill
    
    # Column widths
    col_widths = [25, 35, 12, 10, 10, 45, 35, 10, 15, 35, 35, 35, 50, 10, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = "Market Research Summary"
    ws2['A1'].font = Font(bold=True, size=16)
    
    cur.execute('SELECT COUNT(*) FROM market_research.companies')
    total = cur.fetchone()['count']
    cur.execute('SELECT COUNT(*) FROM market_research.companies WHERE last_researched IS NOT NULL')
    researched = cur.fetchone()['count']
    cur.execute("SELECT COUNT(*) FROM market_research.companies WHERE research_confidence > 0.5")
    high_conf = cur.fetchone()['count']
    cur.execute("SELECT COUNT(*) FROM market_research.companies WHERE thermoforming_services IS NOT NULL AND thermoforming_services != '[]'::jsonb")
    with_data = cur.fetchone()['count']
    
    stats = [
        ("Total Companies", total),
        ("Researched", researched),
        ("High Confidence (>50%)", high_conf),
        ("With Thermoforming Data", with_data),
        ("Export Date", datetime.now().strftime('%Y-%m-%d %H:%M')),
    ]
    
    for i, (label, value) in enumerate(stats, 3):
        ws2[f'A{i}'] = label
        ws2[f'A{i}'].font = Font(bold=True)
        ws2[f'B{i}'] = value
    
    conn.close()
    
    if not output_path:
        output_path = os.path.expanduser('~/Desktop/market_research_deep.xlsx')
    
    wb.save(output_path)
    print(f"\n✅ Exported to: {output_path}")
    print(f"   Total: {total}")
    print(f"   Researched: {researched}")
    print(f"   High confidence: {high_conf}")
    print(f"   With data: {with_data}")
    
    return output_path


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Deep company research for thermoforming market')
    parser.add_argument('--limit', type=int, default=50, help='Number of companies to research')
    parser.add_argument('--workers', type=int, default=3, help='Parallel workers')
    parser.add_argument('--export', action='store_true', help='Export to Excel after research')
    parser.add_argument('--export-only', action='store_true', help='Only export, no research')
    
    args = parser.parse_args()
    
    if args.export_only:
        export_to_excel()
        return
    
    researcher = DeepResearcher()
    researcher.run_batch(limit=args.limit, workers=args.workers)
    
    if args.export:
        export_to_excel()


if __name__ == '__main__':
    main()
