#!/usr/bin/env python3
"""
Document Re-Indexer - Voyage AI
================================

Re-indexes documents from data/imports with Voyage AI embeddings.

Usage:
    python reindex_docs.py
    python reindex_docs.py --dry-run
"""

import hashlib
import json
import logging
import os
import sys
import time
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional

import warnings
warnings.filterwarnings("ignore")

# Import from centralized config
BRAIN_DIR = Path(__file__).parent
SKILLS_DIR = BRAIN_DIR.parent
AGENT_DIR = SKILLS_DIR.parent
sys.path.insert(0, str(AGENT_DIR))

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

try:
    from config import (
        QDRANT_URL, DATABASE_URL, VOYAGE_API_KEY,
        COLLECTIONS, EMBEDDING_MODEL_VOYAGE, PROJECT_ROOT,
    )
except ImportError:
    PROJECT_ROOT = Path(__file__).parent.parent.parent.parent.parent.parent
    env_file = PROJECT_ROOT / ".env"
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            if line.strip() and not line.startswith('#') and '=' in line:
                key, _, value = line.partition('=')
                os.environ.setdefault(key.strip(), value.strip().strip('"'))
    QDRANT_URL = os.environ.get("QDRANT_URL", "http://localhost:6333")
    DATABASE_URL = os.environ.get("DATABASE_URL")
    VOYAGE_API_KEY = os.environ.get("VOYAGE_API_KEY")
    COLLECTIONS = {"chunks_voyage": "ira_chunks_v4_voyage"}
    EMBEDDING_MODEL_VOYAGE = "voyage-3"

IMPORTS_DIR = PROJECT_ROOT / "data" / "imports"
COLLECTION_NAME = COLLECTIONS.get("chunks_voyage", "ira_chunks_v4_voyage")
EMBEDDING_DIM = 1024
VOYAGE_MODEL = EMBEDDING_MODEL_VOYAGE


class DocumentIndexer:
    def __init__(self, batch_size: int = 128):
        self.batch_size = batch_size
        self._qdrant = None
        self._voyage = None
    
    def _get_qdrant(self):
        """Get Qdrant client using centralized config."""
        if self._qdrant is None:
            try:
                from config import get_qdrant_client
                self._qdrant = get_qdrant_client()
            except ImportError:
                from qdrant_client import QdrantClient
                self._qdrant = QdrantClient(url=QDRANT_URL)
        return self._qdrant
    
    def _get_voyage(self):
        """Get Voyage client using centralized config."""
        if self._voyage is None:
            try:
                from config import get_voyage_client
                self._voyage = get_voyage_client()
            except ImportError:
                import voyageai
                self._voyage = voyageai.Client(api_key=VOYAGE_API_KEY)
        return self._voyage
    
    def _get_db_connection(self):
        """Get database connection context manager."""
        try:
            from config import get_db_connection
            return get_db_connection()
        except ImportError:
            import psycopg2
            from contextlib import contextmanager
            
            @contextmanager
            def _fallback():
                conn = psycopg2.connect(DATABASE_URL)
                try:
                    yield conn
                    conn.commit()
                except Exception:
                    conn.rollback()
                    raise
                finally:
                    conn.close()
            
            return _fallback()
    
    def ensure_collection(self):
        """Create Qdrant collection if it doesn't exist."""
        from qdrant_client.models import Distance, VectorParams
        
        qdrant = self._get_qdrant()
        
        try:
            qdrant.get_collection(COLLECTION_NAME)
            count = qdrant.count(COLLECTION_NAME).count
            logger.info(f"Collection {COLLECTION_NAME} exists with {count} vectors")
        except Exception:
            logger.info(f"Creating collection {COLLECTION_NAME}")
            qdrant.create_collection(
                collection_name=COLLECTION_NAME,
                vectors_config=VectorParams(size=EMBEDDING_DIM, distance=Distance.COSINE)
            )
    
    def ensure_schema(self):
        """Create PostgreSQL schema if needed."""
        with self._get_db_connection() as conn:
            cur = conn.cursor()
            
            cur.execute("""
                CREATE SCHEMA IF NOT EXISTS ira_knowledge;
                
                CREATE TABLE IF NOT EXISTS ira_knowledge.chunks_v4 (
                    chunk_id TEXT PRIMARY KEY,
                    doc_id TEXT,
                    filename TEXT,
                    raw_text TEXT,
                    contextual_text TEXT,
                    doc_type TEXT,
                    machines TEXT[],
                    specs JSONB,
                    coherence_score FLOAT,
                    token_count INT,
                    embedding_provider TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                );
            """)
    
    def extract_text_from_pdf(self, filepath: Path) -> str:
        """Extract text from PDF file using shared DocumentExtractor."""
        # Use shared document extractor (has fallback chain: PyMuPDF → pdfplumber → pypdf)
        try:
            from document_extractor import extract_pdf
            text = extract_pdf(filepath)
            if text:
                return text
        except ImportError:
            logger.warning("document_extractor not available, falling back to pdfplumber")
        except Exception as e:
            logger.warning(f"Shared extractor failed: {e}, falling back to pdfplumber")
        
        # Fallback to direct pdfplumber
        try:
            import pdfplumber
        except ImportError:
            logger.error("pdfplumber not installed")
            return ""
        
        text_parts = []
        try:
            with pdfplumber.open(filepath) as pdf:
                for i, page in enumerate(pdf.pages):
                    page_text = page.extract_text() or ""
                    if page_text.strip():
                        text_parts.append(f"--- Page {i+1} ---\n{page_text}")
        except Exception as e:
            logger.warning(f"PDF extraction failed for {filepath}: {e}")
        
        return "\n\n".join(text_parts)
    
    def extract_text_from_excel(self, filepath: Path) -> str:
        """Extract text from Excel file."""
        from openpyxl import load_workbook
        
        text_parts = []
        try:
            wb = load_workbook(filepath, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(c) if c else "" for c in row)
                    if row_text.strip(" |"):
                        rows.append(row_text)
                if rows:
                    text_parts.append(f"Sheet: {sheet_name}\n" + "\n".join(rows[:100]))
        except Exception as e:
            logger.warning(f"Excel extraction failed for {filepath}: {e}")
        
        return "\n\n".join(text_parts)
    
    def chunk_text(self, text: str, chunk_size: int = 500, overlap: int = 50) -> List[Dict]:
        """Split text into chunks."""
        import re
        
        sentences = re.split(r'(?<=[.!?])\s+', text)
        
        chunks = []
        current_chunk = []
        current_length = 0
        
        for sentence in sentences:
            sentence = sentence.strip()
            if not sentence:
                continue
            
            word_count = len(sentence.split())
            
            if current_length + word_count > chunk_size and current_chunk:
                chunk_text = ' '.join(current_chunk)
                chunks.append({'text': chunk_text, 'tokens': current_length})
                
                overlap_sentences = current_chunk[-2:] if len(current_chunk) > 2 else []
                current_chunk = overlap_sentences + [sentence]
                current_length = sum(len(s.split()) for s in current_chunk)
            else:
                current_chunk.append(sentence)
                current_length += word_count
        
        if current_chunk:
            chunk_text = ' '.join(current_chunk)
            chunks.append({'text': chunk_text, 'tokens': current_length})
        
        return chunks
    
    def extract_machines(self, text: str) -> List[str]:
        """Extract machine model names from text."""
        import re
        
        patterns = [
            r'\bPF1[-\s]?[XCSPRA]?[-\s]?\d{2,4}(?:[xX]\d{2,4})?\b',
            r'\bPF2[-\s]?\d{2,4}(?:[xX]\d{2,4})?\b',
            r'\bIMG[-\s]?\d{3,4}\b',
            r'\bATF[-\s]?\d{3,4}\b',
        ]
        
        machines = set()
        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            machines.update(matches)
        
        return list(machines)[:10]
    
    def classify_document(self, filename: str, text: str) -> str:
        """Classify document type."""
        filename_lower = filename.lower()
        text_lower = text[:1000].lower()
        
        if 'quote' in filename_lower or 'quotation' in filename_lower:
            return 'quote'
        elif 'spec' in filename_lower or 'specification' in text_lower:
            return 'specification'
        elif 'price' in filename_lower or 'price list' in text_lower:
            return 'price_list'
        elif 'catalogue' in filename_lower or 'catalog' in filename_lower:
            return 'catalogue'
        elif 'presentation' in filename_lower or '.pptx' in filename_lower:
            return 'presentation'
        elif '.xlsx' in filename_lower or '.xls' in filename_lower:
            return 'spreadsheet'
        else:
            return 'document'
    
    def embed_texts(self, texts: List[str]) -> List[List[float]]:
        """Generate embeddings using Voyage AI."""
        voyage = self._get_voyage()
        
        clean_texts = []
        for t in texts:
            t = t.replace('\x00', '').strip()
            if not t:
                t = "empty"
            clean_texts.append(t[:16000])
        
        result = voyage.embed(clean_texts, model=VOYAGE_MODEL, input_type="document")
        return result.embeddings
    
    def store_chunks(self, chunks: List[Dict]):
        """Store chunks in Qdrant and PostgreSQL."""
        from qdrant_client.models import PointStruct
        
        qdrant = self._get_qdrant()
        
        def make_uuid(chunk_id: str) -> str:
            return str(uuid.uuid5(uuid.NAMESPACE_DNS, chunk_id))
        
        points = [
            PointStruct(
                id=make_uuid(c["chunk_id"]),
                vector=c["embedding"],
                payload={
                    "chunk_id": c["chunk_id"],
                    "filename": c["filename"],
                    "doc_type": c["doc_type"],
                    "raw_text": c["text"][:1000],
                    "machines": c["machines"],
                }
            )
            for c in chunks if c.get("embedding")
        ]
        
        for i in range(0, len(points), 100):
            batch = points[i:i+100]
            qdrant.upsert(collection_name=COLLECTION_NAME, points=batch)
        
        logger.info(f"Stored {len(points)} vectors in Qdrant")
        
        # Store in PostgreSQL
        def sanitize(text):
            return text.replace('\x00', '') if text else text
        
        with self._get_db_connection() as conn:
            cur = conn.cursor()
            
            for c in chunks:
                cur.execute("""
                    INSERT INTO ira_knowledge.chunks_v4 
                    (chunk_id, doc_id, filename, raw_text, doc_type, machines, embedding_provider)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (chunk_id) DO UPDATE SET
                        raw_text = EXCLUDED.raw_text,
                        embedding_provider = EXCLUDED.embedding_provider
                """, (
                    c["chunk_id"], c.get("doc_id", ""), c["filename"],
                    sanitize(c["text"]), c["doc_type"], c["machines"], "voyage"
                ))
        
        logger.info(f"Stored {len(chunks)} chunks in PostgreSQL")
    
    def process_file(self, filepath: Path) -> List[Dict]:
        """Process a single file and return chunks."""
        filename = filepath.name
        
        # Extract text
        if filepath.suffix.lower() == '.pdf':
            text = self.extract_text_from_pdf(filepath)
        elif filepath.suffix.lower() in ['.xlsx', '.xls']:
            text = self.extract_text_from_excel(filepath)
        else:
            logger.warning(f"Unsupported file type: {filepath}")
            return []
        
        if not text or len(text) < 50:
            logger.warning(f"No text extracted from {filename}")
            return []
        
        # Generate doc ID
        doc_id = hashlib.md5(text.encode()).hexdigest()[:16]
        
        # Classify and extract entities
        doc_type = self.classify_document(filename, text)
        machines = self.extract_machines(text)
        
        # Chunk text
        raw_chunks = self.chunk_text(text)
        
        # Create chunk records
        chunks = []
        for i, rc in enumerate(raw_chunks):
            chunk_id = f"{doc_id}_{i:04d}"
            
            # Add context prefix
            context = f"Document: {filename}\nType: {doc_type}\n"
            if machines:
                context += f"Machines: {', '.join(machines[:5])}\n"
            context += "\n"
            
            chunks.append({
                "chunk_id": chunk_id,
                "doc_id": doc_id,
                "filename": filename,
                "text": rc["text"],
                "contextual_text": context + rc["text"],
                "doc_type": doc_type,
                "machines": machines,
                "tokens": rc["tokens"],
            })
        
        logger.info(f"  {filename}: {len(chunks)} chunks")
        return chunks
    
    def reindex_all(self, dry_run: bool = False):
        """Re-index all documents in imports folder."""
        if not IMPORTS_DIR.exists():
            logger.error(f"Imports directory not found: {IMPORTS_DIR}")
            return
        
        # Find files
        files = list(IMPORTS_DIR.glob("*.pdf")) + list(IMPORTS_DIR.glob("*.xlsx"))
        logger.info(f"Found {len(files)} files to process")
        
        if dry_run:
            for f in files[:10]:
                logger.info(f"  Would process: {f.name}")
            return
        
        self.ensure_collection()
        self.ensure_schema()
        
        # Process files
        all_chunks = []
        for filepath in files:
            chunks = self.process_file(filepath)
            all_chunks.extend(chunks)
        
        if not all_chunks:
            logger.warning("No chunks to index")
            return
        
        logger.info(f"Total chunks: {len(all_chunks)}")
        
        # Generate embeddings
        logger.info(f"Generating {len(all_chunks)} embeddings with Voyage AI...")
        texts = [c["contextual_text"] for c in all_chunks]
        
        embeddings = []
        for i in range(0, len(texts), self.batch_size):
            batch = texts[i:i+self.batch_size]
            batch_embeddings = self.embed_texts(batch)
            embeddings.extend(batch_embeddings)
        
        for chunk, emb in zip(all_chunks, embeddings):
            chunk["embedding"] = emb
        
        # Store
        self.store_chunks(all_chunks)
        
        logger.info("Re-indexing complete!")


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="Re-index documents with Voyage AI")
    parser.add_argument("--dry-run", action="store_true", help="Don't make changes")
    parser.add_argument("--batch-size", type=int, default=128, help="Batch size")
    args = parser.parse_args()
    
    if not VOYAGE_API_KEY:
        print("ERROR: VOYAGE_API_KEY not set in .env")
        sys.exit(1)
    
    print("=" * 60)
    print("  DOCUMENT RE-INDEXER (Voyage AI)")
    print("=" * 60)
    print(f"Collection: {COLLECTION_NAME}")
    print(f"Dimensions: {EMBEDDING_DIM}")
    print()
    
    indexer = DocumentIndexer(batch_size=args.batch_size)
    indexer.reindex_all(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
